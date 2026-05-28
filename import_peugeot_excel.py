"""
import_peugeot_excel.py
-----------------------
Direct import of Peugeot parts from unlocked Excel catalog.

Column layout (type F, RTL, data starts row 8):
  A(0) = vehicle model
  B(1) = stock (כן/לא)
  C(2) = price ILS incl. VAT
  D(3) = warranty text
  E(4) = importer (PEUGEOT FRANCE)
  F(5) = part type (ORIGINAL)
  G(6) = Hebrew description
  H(7) = OEM number (מק"\u05d8)

Usage:
  python3 /tmp/import_peugeot_excel.py --dry-run
  python3 /tmp/import_peugeot_excel.py
"""
import asyncio
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import asyncpg
import openpyxl

DB_URL = "postgresql://autospare:e4b79d75ca640dbe7f259618f078b82f21573e419308f668beed5e20b26b1d43@postgres_catalog:5432/autospare"
XLSX_FILE = Path("/app/uploads/peugeot_catalog.xlsx")
MANUFACTURER = "Peugeot"
START_ROW = 8

OEM_PAT = re.compile(r'^[A-Z0-9][A-Z0-9\-\.\/]{4,19}$', re.ASCII)
SKIP_TEXTS = {
    "PEUGEOT", "FRANCE", "ORIGINAL", "ERP", "HYBRID4", "BOXER3",
    "308GTI", "J16H4", "208L",
}

try:
    sys.path.insert(0, "/app")
    from categories import guess_category_by_text
    HAS_CATEGORIES = True
except Exception:
    HAS_CATEGORIES = False


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in {"-", "None", "nan", ""} else None


def parse_price(v):
    if v is None:
        return 0.0
    try:
        s = str(v).replace(",", "").replace("₪", "").strip()
        p = float(s)
        return round(p, 2) if p > 0 else 0.0
    except (ValueError, TypeError):
        return 0.0


def is_valid_oem(s: str) -> bool:
    if not s or len(s) < 5:
        return False
    s = s.strip().upper()
    if s in SKIP_TEXTS:
        return False
    if not OEM_PAT.match(s):
        return False
    if not any(c.isdigit() for c in s):
        return False
    return True


def parse_stock(v) -> tuple[bool, str]:
    s = clean(v)
    if s in ("כן", "yes", "in_stock", "זמין", "available"):
        return True, "in_stock"
    return False, "on_order"


async def run(dry_run: bool):
    if not XLSX_FILE.exists():
        print(f"ERROR: File not found: {XLSX_FILE}")
        print("Upload the unlocked Excel as /app/uploads/peugeot_catalog.xlsx first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "peugeot" in name.lower():
            sheet = wb[name]
            print(f"Using sheet: '{name}'")
            break
    if sheet is None:
        print(f"ERROR: No sheet with 'peugeot' in name. Available: {wb.sheetnames}")
        sys.exit(1)

    rows = []
    skipped_no_oem = 0
    skipped_bad_oem = 0
    skipped_no_name = 0
    model_counts: Counter = Counter()

    for row in sheet.iter_rows(min_row=START_ROW, values_only=True):
        if not row or all(v is None for v in row):
            continue

        vehicle   = clean(row[0]) if len(row) > 0 else None
        stock_raw = row[1]        if len(row) > 1 else None
        price_raw = row[2]        if len(row) > 2 else None
        ptype_raw = clean(row[5]) if len(row) > 5 else None
        name_he   = clean(row[6]) if len(row) > 6 else None
        oem_raw   = clean(row[7]) if len(row) > 7 else None

        if not oem_raw:
            skipped_no_oem += 1
            continue
        oem = oem_raw.strip().upper()
        if not is_valid_oem(oem):
            skipped_bad_oem += 1
            continue
        if not name_he:
            skipped_no_name += 1
            continue

        price = parse_price(price_raw)
        is_avail, av_code = parse_stock(stock_raw)
        part_type = "ORIGINAL" if ptype_raw and "original" in ptype_raw.lower() else (ptype_raw or "OEM")
        if vehicle:
            model_counts[vehicle] += 1

        rows.append({
            "oem":        oem,
            "name_he":    name_he,
            "price":      price,
            "vehicle":    vehicle,
            "is_avail":   is_avail,
            "av_code":    av_code,
            "part_type":  part_type,
        })

    wb.close()

    print(f"\nParsed rows    : {len(rows)}")
    print(f"Skipped no OEM : {skipped_no_oem}")
    print(f"Skipped bad OEM: {skipped_bad_oem}")
    print(f"Skipped no name: {skipped_no_name}")
    print(f"Unique models  : {len(model_counts)}")
    if rows:
        print(f"\nSample rows:")
        for r in rows[:5]:
            print(f"  OEM={r['oem']!s:20} | name={r['name_he']!s:30} | price={r['price']:10.2f} | model={r['vehicle']}")
        oem_lens = Counter(len(r["oem"]) for r in rows)
        print(f"\nOEM length distribution: {sorted(oem_lens.items())}")
        unique_oems = len({r["oem"] for r in rows})
        print(f"Unique OEM numbers: {unique_oems} / {len(rows)} total rows")
        top_models = model_counts.most_common(10)
        print(f"Top models: {top_models}")

    if dry_run:
        print("\n[DRY RUN] — no DB changes made.")
        return

    if not rows:
        print("No valid rows to import. Aborting.")
        return

    conn = await asyncpg.connect(DB_URL)
    try:
        before = await conn.fetchval(
            "SELECT COUNT(*) FROM parts_catalog WHERE LOWER(manufacturer)='peugeot' AND is_active=TRUE"
        )
        print(f"\nDB active Peugeot rows BEFORE: {before}")

        deact = await conn.execute(
            "UPDATE parts_catalog SET is_active=FALSE, updated_at=NOW() "
            "WHERE LOWER(manufacturer)='peugeot' AND is_active=TRUE"
        )
        print(f"Deactivated: {deact}")

        inserted = 0
        updated  = 0
        errors   = 0
        now = datetime.utcnow()

        grouped: dict[str, dict] = {}
        for r in rows:
            oem = r["oem"]
            if oem not in grouped:
                grouped[oem] = {**r, "models": []}
            if r["vehicle"] and r["vehicle"] not in grouped[oem]["models"]:
                grouped[oem]["models"].append(r["vehicle"])
            if r["price"] > grouped[oem]["price"]:
                grouped[oem]["price"] = r["price"]

        deduped = list(grouped.values())
        print(f"Deduped to {len(deduped)} unique OEM rows (from {len(rows)} total)")

        BATCH = 25
        for i in range(0, len(deduped), BATCH):
            batch = deduped[i:i + BATCH]
            for r in batch:
                try:
                    sku = f"PEUG-{r['oem']}"
                    if len(sku) > 100:
                        sku = sku[:100]

                    category = "unknown"
                    if HAS_CATEGORIES:
                        try:
                            category = guess_category_by_text(r["name_he"]) or "unknown"
                        except Exception:
                            pass

                    models_str = ", ".join(r["models"]) if r["models"] else None
                    compat = f"[{models_str}]" if models_str else None

                    new_id = str(uuid.uuid4())
                    await conn.execute("""
                        INSERT INTO parts_catalog
                          (id, sku, oem_number, name, manufacturer,
                           base_price, importer_price_ils, online_price_ils,
                           is_active, part_type, part_condition,
                           compatible_vehicles,
                           category, created_at, updated_at)
                        VALUES ($1,$2,$3,$4,$5, $6,$7,$8, TRUE,$9,$10, $11, $12,$13,$13)
                        ON CONFLICT (sku) DO UPDATE SET
                          oem_number          = EXCLUDED.oem_number,
                          name                = EXCLUDED.name,
                          base_price          = EXCLUDED.base_price,
                          importer_price_ils  = EXCLUDED.importer_price_ils,
                          online_price_ils    = EXCLUDED.online_price_ils,
                          part_type           = EXCLUDED.part_type,
                          part_condition      = EXCLUDED.part_condition,
                          compatible_vehicles = EXCLUDED.compatible_vehicles,
                          category            = EXCLUDED.category,
                          is_active           = TRUE,
                          updated_at          = NOW()
                    """,
                        new_id, sku, r["oem"], r["name_he"], MANUFACTURER,
                        r["price"], r["price"], r["price"],
                        r["part_type"], "original",
                        compat, category, now,
                    )
                    inserted += 1
                except Exception as e:
                    errors += 1
                    if errors <= 5:
                        print(f"  ERR {r['oem']}: {e}")

        after = await conn.fetchval(
            "SELECT COUNT(*) FROM parts_catalog WHERE LOWER(manufacturer)='peugeot' AND is_active=TRUE"
        )
        print(f"\nInserted/updated: {inserted}, errors: {errors}")
        print(f"DB active Peugeot rows AFTER: {after}")

        no_cat = await conn.fetchval(
            "SELECT COUNT(*) FROM parts_catalog WHERE LOWER(manufacturer)='peugeot' "
            "AND is_active=TRUE AND (category IS NULL OR category='unknown')"
        )
        no_price = await conn.fetchval(
            "SELECT COUNT(*) FROM parts_catalog WHERE LOWER(manufacturer)='peugeot' "
            "AND is_active=TRUE AND (base_price IS NULL OR base_price=0)"
        )
        print(f"Quality: no_category={no_cat}, zero_price={no_price}")

    finally:
        await conn.close()

    print("\nNext step: run scoped Meilisearch sync:")
    print("  docker exec autospare_backend python3 /app/meili_sync.py --manufacturer Peugeot --no-rebuild")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    asyncio.run(run(dry))
