"""
Import parts from 'parts data base.xlsx' into the parts_catalog PostgreSQL table.

Column layouts per sheet type (row 2 = headers, row 3+ = data):

Type A  – Chevrolet, Citroen, Peugeot  (7-9 cols)
  [0] importer  [1] catalog_num  [2] name  [3] part_type
  [4] stock     [5] price        [6] warranty  [7] vehicle

Type B  – GEN, Hyundai, JAECOO, Mercedes, Mitsubishi, ORA, Smart  (5-254 cols)
  [0] stock  [1] qty_or_num  [2] name  [3] manufacturer  [4] catalog_num

Type C  – Porsche  (6 cols)
  [0] price  [1] stock  [2] warranty  [3] part_type  [4] name  [5] catalog_num

Type D  – Suzuki  (7 cols)
  [0] price  [1] stock  [2] catalog_num  [3] manufacturer  [4] part_type  [5] name

Type E  – Renault  (6-7 cols)
  [0] date  [1] stock  [2] cat_code  [3] part_type  [4] name  [5] catalog_num
"""

import asyncio
import hashlib
import uuid
import json
import os
import asyncpg
import openpyxl
from datetime import datetime

XLSX_FILE = "parts data base.xlsx"
_raw_url = os.getenv("DATABASE_URL", "postgresql://autospare:autospare_dev@localhost:5432/autospare")
DB_URL = _raw_url.replace("postgresql+asyncpg://", "postgresql://").replace("+asyncpg", "")

SHEET_TYPES = {
    # Type A
    "Chevrolet": "A", "Citroen": "A2", "Peugeot": "A2",
    # Type B (single-block)
    "Hyundai": "B", "Mercedes": "B", "Mitsubishi": "B", "Smart": "B",
    # Type B-multi (horizontal multi-block layout)
    "GEN": "BM", "JAECOO": "BM", "ORA": "BM",
    # Type C
    "Porsche": "C",
    # Type D
    "Suzuki": "D",
    # Type E
    "Renault": "E",
}

# Sheets with horizontal multi-block layout (42 rows/block)
MULTIBLOCK_SHEETS = {"GEN", "JAECOO", "ORA"}


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s not in ('-', 'None') else None


def make_sku(brand: str, catalog_num, row_idx: int) -> str:
    prefix = brand[:4].upper().replace(' ', '_')
    cat = clean(catalog_num)
    if cat and str(cat).strip():
        cat_clean = str(cat).replace(' ', '').replace('/', '-').strip()
        raw_sku = f"{prefix}-{cat_clean}"
        if len(raw_sku) > 100:
            h = hashlib.md5(cat_clean.encode()).hexdigest()[:12]
            raw_sku = f"{prefix}-{h}"
    else:
        raw_sku = f"{prefix}-R{row_idx}"
    return raw_sku[:100]


def parse_price(val):
    if val is None:
        return 0.0
    try:
        # Remove currency symbols
        s = str(val).replace('₪', '').replace(',', '').strip()
        p = float(s)
        return min(max(p, 0.0), 99_999_999.99)
    except (ValueError, TypeError):
        return 0.0


def parse_vehicle(vehicle_str: str, brand: str) -> dict | None:
    import re
    if not vehicle_str:
        return None
    s = str(vehicle_str).strip()
    if not s or s in ('-', 'None', 'מק"ט', 'דגם'):
        return None
    years = re.findall(r'\b(?:19|20)\d{2}\b', s)
    year = int(years[-1]) if years else None
    model = re.sub(r'\b(?:19|20)\d{2}\b', '', s).strip()
    model = re.sub(r'\s*\bNEW\b\s*$', '', model).strip()
    if not model:
        model = s
    return {
        "manufacturer": brand,
        "model": model[:100],
        "year_from": year or 2000,
        "year_to": year,
    }


# Known header cell values to skip in multi-block sheets
_HEADER_VALS = {'זמינות מלאי', 'מחיר לצרכן', 'תיאור החלק', 'מותג', 'מספר קטלוגי'}


def _make_part_record(brand: str, avail, price, name, mfr, catalog, row_idx: int) -> dict | None:
    """Helper: build a part dict from extracted fields, or return None to skip."""
    catalog = clean(catalog)
    if not catalog or str(catalog) in _HEADER_VALS:
        return None
    name = clean(name) or "(ללא שם)"
    if name in _HEADER_VALS:
        return None
    stock = clean(avail)
    manufacturer = clean(mfr) or brand
    specs = {}
    if stock:
        specs["stock_status"] = stock
    specs["manufacturer_name"] = manufacturer
    sku = make_sku(brand, catalog, row_idx)
    return {
        "sku": sku,
        "name": name[:255],
        "category": brand[:100],
        "manufacturer": manufacturer[:100],
        "part_type": "unknown",
        "description": name[:500],
        "specifications": json.dumps(specs, ensure_ascii=False),
        "compatible_vehicles": "[]",
        "base_price": parse_price(price),
    }


def parse_multiblock_row(sheet_name: str, row: tuple, row_idx: int) -> list:
    """
    Parse a row from a multi-block sheet (GEN/JAECOO/ORA).

    Layout:
      Block 0 (cols 0-4):  avail, price, name, mfr, catalog
      Block 1 (cols 5-9):  avail, price, name, mfr, catalog
      Block 2+ (cols 5,8,X,X+1,X+2 for X=10,13,16,...): avail=col5, mfr=col8,
          then price/name/catalog rotate every 3 columns starting at col 10.
    Each 42-row chunk of data belongs to exactly one block.
    """
    brand = sheet_name.strip()
    row = list(row)
    max_col = len(row)

    def g(idx):
        return row[idx] if idx < max_col else None

    results = []

    # Block 0: cols 0-4
    if any(g(i) is not None for i in range(5)):
        r = _make_part_record(brand, g(0), g(1), g(2), g(3), g(4), row_idx)
        if r:
            results.append(r)

    # Block 1: cols 5-9 (full 5-col record)
    if any(g(i) is not None for i in range(5, 10)):
        # Validate: needs at least name or catalog
        if g(9) is not None or g(7) is not None:
            r = _make_part_record(brand, g(5), g(6), g(7), g(8), g(9), row_idx)
            if r:
                results.append(r)

    # Blocks 2+: avail shared at col 5, mfr shared at col 8,
    # then (price, name, catalog) in groups of 3 starting at col 10
    avail_common = g(5)
    mfr_common = g(8)
    for x in range(10, max_col - 1, 3):
        price_v = g(x)
        name_v = g(x + 1)
        catalog_v = g(x + 2) if x + 2 < max_col else None
        if any(v is not None for v in (price_v, name_v, catalog_v)):
            r = _make_part_record(brand, avail_common, price_v, name_v, mfr_common, catalog_v, row_idx)
            if r:
                results.append(r)

    return results


def parse_row(sheet_name: str, row: tuple, row_idx: int) -> dict | None:
    """Parse a row based on the sheet type. Returns a dict or None to skip."""
    if not any(row):
        return None

    stype = SHEET_TYPES.get(sheet_name, "A")
    brand = sheet_name.strip()

    def g(idx):
        return row[idx] if idx < len(row) else None

    if stype in ("A", "A2"):
        if stype == "A2":
            catalog_num = g(7)
            name = clean(g(6)) or "(ללא שם)"
            part_type = (clean(g(5)) or "unknown")[:50]
            stock = clean(g(1))
            price = parse_price(g(2))
            warranty = clean(g(3))
            vehicle = clean(g(0))
            importer = clean(g(4))
        else:
            catalog_num = g(1)
            name = clean(g(2)) or "(ללא שם)"
            part_type = (clean(g(3)) or "unknown")[:50]
            stock = clean(g(4))
            price = parse_price(g(5))
            warranty = clean(g(6))
            vehicle = clean(g(7))
            importer = clean(g(0))
        specs = {k: v for k, v in {"stock_status": stock, "warranty": warranty, "importer": importer}.items() if v}
        compat = [{"make": brand, "model_year": vehicle}] if vehicle else []

    elif stype in ("B", "BM"):
        catalog_num = g(4)
        name = clean(g(2)) or "(ללא שם)"
        part_type = "unknown"
        stock = clean(g(0))
        price = 0.0
        manufacturer_override = clean(g(3))
        specs = {k: v for k, v in {"stock_status": stock}.items() if v}
        if manufacturer_override:
            specs["manufacturer_name"] = manufacturer_override
        compat = []

    elif stype == "C":
        catalog_num = g(5)
        name = clean(g(4)) or "(ללא שם)"
        part_type = (clean(g(3)) or "unknown")[:50]
        stock = clean(g(1))
        price = parse_price(g(0))
        warranty = clean(g(2))
        specs = {k: v for k, v in {"stock_status": stock, "warranty": warranty}.items() if v}
        compat = []

    elif stype == "D":
        catalog_num = g(2)
        name = clean(g(5)) or clean(g(4)) or "(ללא שם)"
        part_type = (clean(g(4)) or "unknown")[:50]
        stock = clean(g(1))
        price = parse_price(g(0))
        specs = {k: v for k, v in {"stock_status": stock}.items() if v}
        compat = []

    elif stype == "E":
        catalog_num = g(5)
        name = clean(g(4)) or "(ללא שם)"
        part_type = (clean(g(3)) or "unknown")[:50]
        stock = clean(g(1))
        price = 0.0
        category_code = clean(g(2))
        specs = {k: v for k, v in {"stock_status": stock, "category_code": category_code}.items() if v}
        compat = []
    else:
        return None

    sku = make_sku(brand, catalog_num, row_idx)

    fitment = parse_vehicle(vehicle, brand) if stype in ("A", "A2") else None

    return {
        "sku": sku,
        "name": name[:255],
        "category": brand[:100],
        "manufacturer": brand[:100],
        "part_type": part_type[:50],
        "description": name[:500],
        "specifications": json.dumps(specs, ensure_ascii=False),
        "compatible_vehicles": json.dumps(compat, ensure_ascii=False),
        "base_price": price,
        "fitment": fitment,
    }


UPSERT_SQL = """
INSERT INTO parts_catalog
    (id, sku, name, category, manufacturer, part_type,
     description, specifications, compatible_vehicles,
     base_price, is_active, created_at, updated_at)
VALUES ($1, $2, $3, $4, $5, $6, $7, $8::jsonb, $9::jsonb, $10, true, NOW(), NOW())
ON CONFLICT (sku) DO UPDATE SET
    name = EXCLUDED.name,
    category = EXCLUDED.category,
    manufacturer = EXCLUDED.manufacturer,
    part_type = EXCLUDED.part_type,
    description = EXCLUDED.description,
    specifications = EXCLUDED.specifications,
    compatible_vehicles = EXCLUDED.compatible_vehicles,
    base_price = EXCLUDED.base_price,
    is_active = true,
    updated_at = NOW()
RETURNING id
"""

FITMENT_SQL = """
INSERT INTO part_vehicle_fitment
    (id, part_id, manufacturer, model, year_from, year_to, created_at)
VALUES ($1, $2, $3, $4, $5, $6, NOW())
ON CONFLICT DO NOTHING
"""

BATCH_SIZE = 500


async def import_parts():
    print(f"[{datetime.now():%H:%M:%S}] Opening workbook…")
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    conn = await asyncpg.connect(DB_URL)
    print("Connected to PostgreSQL\n")

    total_upserted = total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        stype = SHEET_TYPES.get(sheet_name, "A")
        is_multiblock = sheet_name in MULTIBLOCK_SHEETS
        print(f"  → {sheet_name} (type {stype})")
        sheet_rows = 0
        sheet_errors = 0
        batch = []

        # Multi-block sheets: start from row 1 (include all rows; header cols are
        # filtered by _HEADER_VALS inside the parser).
        # Regular sheets: start from row 3 (2-row header).
        start_row = 1 if is_multiblock else (8 if stype == "A2" else 3)

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), 1):
            if is_multiblock:
                parsed_list = parse_multiblock_row(sheet_name, row, row_idx)
            else:
                p = parse_row(sheet_name, row, row_idx)
                parsed_list = [p] if p else []

            for parsed in parsed_list:
                if parsed is None:
                    continue

                part_id = uuid.uuid4()
                batch.append({
                    "rec": (
                        part_id,
                        parsed["sku"], parsed["name"], parsed["category"],
                        parsed["manufacturer"], parsed["part_type"], parsed["description"],
                        parsed["specifications"], parsed["compatible_vehicles"],
                        parsed["base_price"],
                    ),
                    "fitment": parsed.get("fitment"),
                    "part_id": part_id,
                })

            if len(batch) >= BATCH_SIZE:
                for r in batch:
                    try:
                        row_result = await conn.fetchrow(UPSERT_SQL, *r["rec"])
                        sheet_rows += 1
                        if r["fitment"] and row_result:
                            actual_id = row_result["id"]
                            await conn.execute(FITMENT_SQL,
                                uuid.uuid4(), actual_id,
                                r["fitment"]["manufacturer"],
                                r["fitment"]["model"],
                                r["fitment"]["year_from"],
                                r["fitment"]["year_to"])
                    except Exception:
                        sheet_errors += 1
                batch = []

        # Flush remaining
        if batch:
            for r in batch:
                try:
                    row_result = await conn.fetchrow(UPSERT_SQL, *r["rec"])
                    sheet_rows += 1
                    if r["fitment"] and row_result:
                        actual_id = row_result["id"]
                        await conn.execute(FITMENT_SQL,
                            uuid.uuid4(), actual_id,
                            r["fitment"]["manufacturer"],
                            r["fitment"]["model"],
                            r["fitment"]["year_from"],
                            r["fitment"]["year_to"])
                except Exception:
                    sheet_errors += 1

        total_upserted += sheet_rows
        total_errors += sheet_errors
        print(f"     ✓ {sheet_rows:,} rows  ({sheet_errors} errors)")

    wb.close()

    final_count = await conn.fetchval("SELECT COUNT(*) FROM parts_catalog")
    await conn.close()

    print(f"\n{'='*50}")
    print(f"Import complete:")
    print(f"  Rows upserted : {total_upserted:,}")
    print(f"  Errors        : {total_errors:,}")
    print(f"  DB total now  : {final_count:,}")
    print(f"{'='*50}")


if __name__ == "__main__":
    asyncio.run(import_parts())

