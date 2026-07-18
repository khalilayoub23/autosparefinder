#!/usr/bin/env python3
"""
Import IL prices from parts_database.xlsx for all brands with priced rows.
The Renault sheet contains mixed brands (NI=Infiniti/Nissan, RE=Renault, etc.)
Prices are IL retail (ILS incl. VAT) stored in max_price_ils.
After import, normalize_base_price is called to set base_price = max_price_ils * 1.45
"""
import sys
import os
sys.path.insert(0, '/app')

import openpyxl
import asyncpg
import asyncio
import re

DB_URL = os.environ.get("DATABASE_URL", "postgresql://autospare:autospare@db:5432/autospare").replace("postgresql+asyncpg://", "postgresql://")
XLS_PATH = "/app/data/parts_database.xlsx"

# Map OEM prefixes to manufacturer names in DB
PREFIX_TO_MANUFACTURER = {
    "NI": ["Infiniti", "Nissan"],
    "RE": ["Renault"],
    "CH": ["Chrysler", "Jeep", "Dodge"],
    "HO": ["Honda"],
    "PO": ["Porsche"],
    # Add more as needed
}


def parse_doubled_string(val):
    """Handle duplicated strings like '101.31101.31' -> 101.31"""
    s = str(val).strip()
    # Try to find a numeric pattern repeated
    # e.g., "101.31101.31" or "1234.561234.56"
    mid = len(s) // 2
    if len(s) % 2 == 0 and s[:mid] == s[mid:]:
        s = s[:mid]
    # Now parse as float
    try:
        return float(s)
    except ValueError:
        # Try stripping currency symbols
        s = re.sub(r'[^\d.]', '', s)
        try:
            return float(s) if s else None
        except ValueError:
            return None


def clean_oem(oem_raw):
    """Clean and deduplicate OEM number strings like 'NI1234  NI1234 ' -> 'NI1234'"""
    s = str(oem_raw).strip()
    # Handle doubled strings
    words = s.split()
    if len(words) >= 2:
        # Check if first word repeated
        half = len(words) // 2
        if words[:half] == words[half:]:
            words = words[:half]
    oem = ' '.join(words).strip()
    # Also handle exact character doubling
    mid = len(oem) // 2
    if len(oem) % 2 == 0 and oem[:mid] == oem[mid:]:
        oem = oem[:mid]
    return oem.strip()


def load_priced_rows():
    """Load all rows from Renault sheet that have a price."""
    print(f"Loading {XLS_PATH} ...")
    wb = openpyxl.load_workbook(XLS_PATH, read_only=True, data_only=True)

    priced = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        sheet_priced = 0
        for row in rows_iter:
            if len(row) < 7:
                continue
            price_raw = row[5]
            oem_raw = row[6]
            name_raw = row[4]

            if price_raw is None or str(price_raw).strip() == '':
                continue

            price = parse_doubled_string(price_raw)
            if price is None or price <= 0:
                continue

            oem = clean_oem(oem_raw) if oem_raw else None
            if not oem or len(oem) < 3:
                continue

            name = str(name_raw).strip() if name_raw else ''
            # Deduplicate name too
            mid = len(name) // 2
            if len(name) % 2 == 0 and name[:mid] == name[mid:]:
                name = name[:mid]

            priced.append({
                'sheet': sheet_name,
                'oem': oem,
                'price': price,
                'name': name,
            })
            sheet_priced += 1

        if sheet_priced:
            print(f"  Sheet '{sheet_name}': {sheet_priced} priced rows")

    wb.close()
    print(f"Total priced rows across all sheets: {len(priced)}")
    return priced


async def import_prices(conn, priced_rows):
    """Match OEM numbers to parts_catalog and update max_price_ils."""
    updated = 0
    not_found = 0
    skipped = 0

    # Batch by OEM for efficiency
    for i, row in enumerate(priced_rows):
        oem = row['oem']
        price = row['price']
        sheet = row['sheet']

        # Normalize OEM: remove spaces, uppercase
        oem_clean = re.sub(r'\s+', '', oem).upper()

        matches = await conn.fetch("""
            SELECT id, oem_number, manufacturer, max_price_ils, base_price
            FROM parts_catalog
            WHERE UPPER(REPLACE(oem_number, ' ', '')) = $1
              AND is_active = TRUE
            LIMIT 5
        """, oem_clean)

        if not matches:
            not_found += 1
            continue

        for m in matches:
            old_max = m['max_price_ils']
            # Don't overwrite if existing IL price is already higher (better data source)
            if old_max and old_max > 0 and old_max > price * 1.5:
                skipped += 1
                continue

            await conn.execute("""
                UPDATE parts_catalog
                SET max_price_ils = $1,
                    updated_at = NOW()
                WHERE id = $2
            """, price, m['id'])
            updated += 1
            if updated <= 20 or updated % 1000 == 0:
                print(f"  [{sheet}] Updated {m['manufacturer']} | {m['oem_number']} | {row['name'][:35]} | {old_max} -> {price} ILS")

        if (i + 1) % 5000 == 0:
            print(f"  Progress: {i+1}/{len(priced_rows)} rows processed, {updated} updated")

    return updated, not_found, skipped


async def normalize_prices_after_import(conn, updated_count):
    """Set base_price for parts that now have max_price_ils but had no base_price."""
    # Policy Case 3: only max_price_ils available -> base_price = max_price_ils * 1.45
    result = await conn.fetch("""
        UPDATE parts_catalog
        SET base_price = ROUND(max_price_ils * 1.45, 2),
            updated_at = NOW()
        WHERE max_price_ils > 0
          AND (base_price IS NULL OR base_price = 0)
          AND is_active = TRUE
        RETURNING id
    """)
    print(f"Normalized base_price (×1.45) for {len(result)} parts")
    return len(result)


async def main():
    print("=== XLS Price Importer ===")
    priced_rows = load_priced_rows()

    if not priced_rows:
        print("No priced rows found, exiting.")
        return

    print(f"\nConnecting to DB...")
    conn = await asyncpg.connect(DB_URL)

    print(f"\nImporting {len(priced_rows)} priced rows...")
    updated, not_found, skipped = await import_prices(conn, priced_rows)

    print(f"\n=== Import Summary ===")
    print(f"  Parts updated: {updated}")
    print(f"  OEMs not found in DB: {not_found}")
    print(f"  Skipped (better price exists): {skipped}")

    if updated > 0:
        print(f"\nNormalizing base_price for newly priced parts...")
        await normalize_prices_after_import(conn, updated)

    await conn.close()
    print("Done.")


if __name__ == '__main__':
    asyncio.run(main())
