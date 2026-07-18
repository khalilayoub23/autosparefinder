#!/usr/bin/env python3
"""Export samelet brand catalogs to XLS files in /app/uploads/."""
import asyncio, asyncpg, json, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DB_URL = "postgresql://autospare:e4b79d75ca640dbe7f259618f078b82f21573e419308f668beed5e20b26b1d43@postgres_catalog:5432/autospare"
UPLOAD_DIR = "/app/uploads"

BRANDS = ["Alfa Romeo","Jeep","Fiat","RAM","Subaru","Abarth","Iveco","Hongqi","WEY"]
SINGLE = os.environ.get("SINGLE_BRAND","")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
COLS = ["SKU","Name","Name (Hebrew)","Category","Base Price (ILS)","Importer Price (ILS)",
        "Part Type","OEM Number","Compatible Vehicles","Updated At"]
COL_WIDTHS = [20,40,40,20,16,18,14,22,50,22]

async def export_brand(conn, brand):
    rows = await conn.fetch("""
        SELECT sku, name, name_he, category, base_price, importer_price_ils,
               part_type, oem_number, compatible_vehicles::text, updated_at
        FROM parts_catalog
        WHERE manufacturer=$1 AND is_active=TRUE
        ORDER BY category, name
    """, brand)
    if not rows:
        print(f"  {brand}: 0 parts, skipping")
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = brand[:31]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for ci, (col, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows, 2):
        cv_raw = row["compatible_vehicles"]
        if cv_raw:
            try:
                cv_list = json.loads(cv_raw)
                cv_str = "; ".join(
                    f"{v.get('brand','')} {v.get('model','')} {v.get('years','')}".strip()
                    for v in cv_list
                ) if isinstance(cv_list, list) else str(cv_list)
            except:
                cv_str = str(cv_raw)
        else:
            cv_str = ""

        vals = [
            row["sku"],
            row["name"] or "",
            row["name_he"] or "",
            row["category"] or "",
            float(row["base_price"] or 0),
            float(row["importer_price_ils"] or 0),
            row["part_type"] or "",
            row["oem_number"] or "",
            cv_str,
            str(row["updated_at"])[:19] if row["updated_at"] else "",
        ]
        for ci, val in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=val)

    slug = brand.lower().replace(" ","_").replace("/","_")
    filename = f"{slug}_parts_catalog.xlsx"
    path = os.path.join(UPLOAD_DIR, filename)
    wb.save(path)
    size_kb = os.path.getsize(path) // 1024
    print(f"  {brand}: {len(rows)} parts → {filename} ({size_kb}KB)")
    return len(rows)

async def main():
    conn = await asyncpg.connect(DB_URL)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    try:
        brands = [SINGLE] if SINGLE else BRANDS
        total = 0
        for brand in brands:
            c = await export_brand(conn, brand)
            total += c
        print(f"\nTotal exported: {total} parts across {len(brands)} brands")
    finally:
        await conn.close()

if __name__ == "__main__":
    asyncio.run(main())
