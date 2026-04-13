from __future__ import annotations

import asyncio
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List

import openpyxl
from sqlalchemy import text

from BACKEND_DATABASE_MODELS import async_session_factory
from manufacturer_normalization import normalize_manufacturer_name
from workbook_normalizer import iter_normalized_rows


OUTPUT_XLSX_FILE = Path(__file__).parent / "data" / "full car database.xlsx"
PENDING_JSON_FILE = Path(__file__).parent / "data" / "full_car_database.pending_enrichment.json"

EXPORT_HEADERS = [
    "source_sheet",
    "source_row",
    "catalog_num",
    "sku",
    "oem_number",
    "name",
    "part_type",
    "category",
    "manufacturer",
    "manufacturer_id",
    "base_price",
    "vehicle_manufacturer",
    "vehicle_model",
    "vehicle_submodel",
    "year_from",
    "year_to",
    "fitment_status",
    "enrichment_status",
    "needs_scraper",
    "needs_worker",
    "specifications_json",
    "compatible_vehicles_json",
    "notes",
]

SUMMARY_HEADERS = [
    "manufacturer",
    "total_rows",
    "rows_with_fitment",
    "rows_missing_fitment",
    "rows_missing_price",
    "sheet_name",
]

PENDING_HEADERS = EXPORT_HEADERS


def _lookup_token(value: Any) -> str:
    return str(value or "").strip().upper()


async def _load_catalog_fitment_lookup() -> Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]]:
    lookup: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]] = defaultdict(
        lambda: {"sku": {}, "oem": {}, "catalog": {}}
    )

    async with async_session_factory() as db:
        rows = (await db.execute(text("""
            SELECT manufacturer, sku, oem_number, compatible_vehicles
            FROM parts_catalog
            WHERE is_active = TRUE
              AND manufacturer IS NOT NULL
              AND compatible_vehicles IS NOT NULL
              AND jsonb_typeof(compatible_vehicles) = 'array'
              AND jsonb_array_length(compatible_vehicles) > 0
        """))).fetchall()

    for manufacturer, sku, oem_number, compatible_vehicles in rows:
        canonical_manufacturer = normalize_manufacturer_name(str(manufacturer or ""), str(manufacturer or ""))
        if not canonical_manufacturer:
            continue

        compat_list = compatible_vehicles if isinstance(compatible_vehicles, list) else _json_loads(compatible_vehicles, [])
        if not compat_list:
            continue

        bucket = lookup[canonical_manufacturer.casefold()]
        sku_key = _lookup_token(sku)
        oem_key = _lookup_token(oem_number)

        if sku_key and sku_key not in bucket["sku"]:
            bucket["sku"][sku_key] = compat_list
        if oem_key:
            if oem_key not in bucket["oem"]:
                bucket["oem"][oem_key] = compat_list
            if oem_key not in bucket["catalog"]:
                bucket["catalog"][oem_key] = compat_list

    return lookup


def _resolve_catalog_fitment(
    record: Dict[str, Any],
    fitment_lookup: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]],
) -> List[Dict[str, Any]]:
    canonical_manufacturer = normalize_manufacturer_name(
        str(record.get("manufacturer") or ""),
        str(record.get("source_sheet") or ""),
    )
    if not canonical_manufacturer:
        return []

    bucket = fitment_lookup.get(canonical_manufacturer.casefold())
    if not bucket:
        return []

    for family, value in (
        ("sku", record.get("sku")),
        ("oem", record.get("oem_number")),
        ("catalog", record.get("catalog_num")),
    ):
        token = _lookup_token(value)
        if token and token in bucket[family]:
            return bucket[family][token]
    return []


def _safe_sheet_name(value: str) -> str:
    invalid = '[]:*?/\\'
    cleaned = "".join("_" if ch in invalid else ch for ch in value).strip()
    cleaned = cleaned or "UNKNOWN"
    return cleaned[:31]


def _json_loads(value: Any, default: Any) -> Any:
    if not value:
        return default
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return default
    return value


def _header_like(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return text.casefold() in {
        "מחיר",
        "price",
        "מספר קטלוגי",
        "catalog",
        "catalog_num",
        "name",
        "תיאור החלק",
    }


def _expand_record(
    record: Dict[str, Any],
    fitment_lookup: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]] | None = None,
) -> Dict[str, Any]:
    compat_list = _json_loads(record.get("compatible_vehicles"), [])
    fitment_source = "source_workbook"
    if not compat_list and fitment_lookup:
        compat_list = _resolve_catalog_fitment(record, fitment_lookup)
        if compat_list:
            fitment_source = "parts_catalog"
    compat = compat_list[0] if compat_list else {}
    specs = _json_loads(record.get("specifications"), {})

    vehicle_manufacturer = compat.get("manufacturer") or compat.get("make") or record.get("manufacturer") or ""
    vehicle_model = compat.get("model") or compat.get("model_year") or ""
    vehicle_submodel = compat.get("sub_model") or compat.get("trim") or compat.get("generation") or ""
    year_from = compat.get("year_from") or compat.get("year") or ""
    year_to = compat.get("year_to") or compat.get("year") or year_from

    fitment_status = "ready" if vehicle_model else "missing_fitment"
    enrichment_status = (
        "ready"
        if fitment_source == "source_workbook" and vehicle_model
        else "worker_db_matched"
        if vehicle_model
        else "pending_scraper_worker"
    )
    needs_scraper = "yes" if not vehicle_model else "no"
    needs_worker = "yes" if not vehicle_model else "no"

    note_parts: List[str] = []
    if not vehicle_model:
        note_parts.append("Missing fitment in source workbook")
    elif fitment_source == "parts_catalog":
        note_parts.append("Fitment pulled from parts_catalog compatible_vehicles")
    if not record.get("base_price"):
        note_parts.append("Missing price in source workbook")
    if specs.get("stock_status"):
        note_parts.append(f"stock={specs['stock_status']}")

    return {
        "source_sheet": record.get("source_sheet") or "",
        "source_row": record.get("source_row") or "",
        "catalog_num": record.get("catalog_num") or "",
        "sku": record.get("sku") or "",
        "oem_number": record.get("oem_number") or "",
        "name": record.get("name") or "",
        "part_type": record.get("part_type") or "",
        "category": record.get("category") or "",
        "manufacturer": record.get("manufacturer") or record.get("source_sheet") or "UNKNOWN",
        "manufacturer_id": record.get("manufacturer_id") if record.get("manufacturer_id") is not None else "",
        "base_price": record.get("base_price") or "",
        "vehicle_manufacturer": vehicle_manufacturer,
        "vehicle_model": vehicle_model,
        "vehicle_submodel": vehicle_submodel,
        "year_from": year_from,
        "year_to": year_to,
        "fitment_status": fitment_status,
        "enrichment_status": enrichment_status,
        "needs_scraper": needs_scraper,
        "needs_worker": needs_worker,
        "specifications_json": json.dumps(specs, ensure_ascii=False),
        "compatible_vehicles_json": json.dumps(compat_list, ensure_ascii=False),
        "notes": "; ".join(note_parts),
    }


def _iter_export_rows(
    fitment_lookup: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]] | None = None,
) -> Iterable[Dict[str, Any]]:
    for record in iter_normalized_rows():
        if _header_like(record.get("catalog_num")) or _header_like(record.get("name")):
            continue
        yield _expand_record(record, fitment_lookup=fitment_lookup)


def build_full_car_database(
    output_path: Path = OUTPUT_XLSX_FILE,
    pending_output_path: Path = PENDING_JSON_FILE,
    fitment_lookup: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]] | None = None,
) -> Path:
    rows_by_manufacturer: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    pending_rows: List[Dict[str, Any]] = []
    fitment_lookup = fitment_lookup or asyncio.run(_load_catalog_fitment_lookup())

    for row in _iter_export_rows(fitment_lookup=fitment_lookup):
        manufacturer = str(row.get("manufacturer") or "UNKNOWN").strip() or "UNKNOWN"
        rows_by_manufacturer[manufacturer].append(row)
        if row["fitment_status"] != "ready":
            pending_rows.append(row)

    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(SUMMARY_HEADERS)

    pending_sheet = workbook.create_sheet("Pending Enrichment")
    pending_sheet.append(PENDING_HEADERS)

    for row in pending_rows:
        pending_sheet.append([row.get(header, "") for header in PENDING_HEADERS])

    for manufacturer in sorted(rows_by_manufacturer):
        sheet_name = _safe_sheet_name(manufacturer)
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(EXPORT_HEADERS)

        manufacturer_rows = rows_by_manufacturer[manufacturer]
        with_fitment = 0
        missing_fitment = 0
        missing_price = 0

        for row in manufacturer_rows:
            if row["fitment_status"] == "ready":
                with_fitment += 1
            else:
                missing_fitment += 1
            if not row.get("base_price"):
                missing_price += 1
            sheet.append([row.get(header, "") for header in EXPORT_HEADERS])

        summary_sheet.append([
            manufacturer,
            len(manufacturer_rows),
            with_fitment,
            missing_fitment,
            missing_price,
            sheet_name,
        ])

        widths = {
            "A": 14,
            "B": 10,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 36,
            "G": 14,
            "H": 16,
            "I": 18,
            "J": 14,
            "K": 12,
            "L": 18,
            "M": 20,
            "N": 18,
            "O": 10,
            "P": 10,
            "Q": 16,
            "R": 24,
            "S": 12,
            "T": 12,
            "U": 40,
            "V": 44,
            "W": 36,
        }
        for column_name, width in widths.items():
            sheet.column_dimensions[column_name].width = width
        sheet.freeze_panes = "A2"

    summary_sheet.freeze_panes = "A2"
    pending_sheet.freeze_panes = "A2"
    for sheet in (summary_sheet, pending_sheet):
        for column_name in [chr(code) for code in range(ord("A"), ord("W") + 1)]:
            if column_name in sheet.column_dimensions:
                continue
            sheet.column_dimensions[column_name].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    pending_payload = {
        "workbook": str(output_path),
        "pending_sheet": "Pending Enrichment",
        "pending_rows": len(pending_rows),
        "manufacturers": {
            manufacturer: {
                "pending_rows": sum(1 for row in manufacturer_rows if row["fitment_status"] != "ready"),
                "sample_catalog_numbers": [
                    row["catalog_num"]
                    for row in manufacturer_rows
                    if row["fitment_status"] != "ready"
                ][:20],
            }
            for manufacturer, manufacturer_rows in sorted(rows_by_manufacturer.items())
            if any(row["fitment_status"] != "ready" for row in manufacturer_rows)
        },
    }
    pending_output_path.write_text(
        json.dumps(pending_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


if __name__ == "__main__":
    path = build_full_car_database()
    print(path)