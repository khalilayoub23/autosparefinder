#!/usr/bin/env python3
"""Import merged_all_final workbook into parts_catalog/supplier_parts/part_vehicle_fitment.

Usage:
  python scripts/import_merged_catalog.py --file /path/to/merged_all_final.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values


SUPPLIER_NAME = "Official Manufacturer Sites"
BATCH_SIZE_DEFAULT = 500
PROGRESS_EVERY = 10_000
ILS_TO_USD = Decimal("3.70")

YEAR_RANGE_RE = re.compile(r"\b(19\d{2}|20\d{2})\s*[-/]\s*(19\d{2}|20\d{2})\b")
YEAR_RE = re.compile(r"\b(19\d{2}|20\d{2})\b")


@dataclass
class NormalizedRow:
    sku: str
    name: str
    manufacturer: str
    model: str | None
    year_from: int | None
    year_to: int | None
    price_ils: Decimal | None
    part_condition: str
    is_available: bool
    importer_info: str | None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value != value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in {"nan", "none", "null", "-"}:
        return None
    return text


def parse_price_ils(value: Any) -> Decimal | None:
    text = clean_text(value)
    if not text:
        return None
    normalized = re.sub(r"[^0-9.\-]", "", text.replace(",", ""))
    if not normalized:
        return None
    try:
        price = Decimal(normalized)
    except InvalidOperation:
        return None
    if price <= 0:
        return None
    return price.quantize(Decimal("0.01"))


def normalize_manufacturer(value: Any, fallback: str) -> str:
    text = clean_text(value) or fallback
    return text.strip().title()


def normalize_part_condition(value: Any) -> str:
    text = (clean_text(value) or "").strip()
    upper = text.upper()
    if text == "מקורי" or upper in {"ORIGINAL", "OEM"}:
        return "OEM"
    return "aftermarket"


def normalize_availability(value: Any) -> bool:
    text = (clean_text(value) or "").strip().lower()
    return text in {"זמין במלאי", "כן", "זמין", "yes", "in stock", "available", "true", "1"}


def parse_model_and_year(model_value: Any) -> tuple[str | None, int | None, int | None]:
    model_text = clean_text(model_value)
    if not model_text:
        return None, None, None

    year_from = None
    year_to = None
    clean_model = model_text

    range_match = YEAR_RANGE_RE.search(model_text)
    if range_match:
        y1 = int(range_match.group(1))
        y2 = int(range_match.group(2))
        year_from = min(y1, y2)
        year_to = max(y1, y2)
        clean_model = YEAR_RANGE_RE.sub(" ", clean_model, count=1)
    else:
        year_match = YEAR_RE.search(model_text)
        if year_match:
            year_from = int(year_match.group(1))
            year_to = year_from
            clean_model = YEAR_RE.sub(" ", clean_model, count=1)

    clean_model = re.sub(r"[_/\-]+", " ", clean_model)
    clean_model = re.sub(r"\s+", " ", clean_model).strip()
    if not clean_model:
        clean_model = model_text

    return clean_model, year_from, year_to


def get_database_url() -> str:
    raw = os.getenv("DATABASE_URL", "").strip()
    if not raw:
        raise RuntimeError("DATABASE_URL environment variable is required")
    return raw.replace("postgresql+asyncpg://", "postgresql://").replace("+asyncpg", "")


def ensure_supporting_indexes(conn: psycopg2.extensions.connection) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS uq_supplier_parts_part_supplier
            ON supplier_parts (part_id, supplier_id)
            """
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS uix_pvf_part_mfr_model_year_from
            ON part_vehicle_fitment (part_id, manufacturer, model, year_from)
            """
        )


def ensure_temp_table(conn: psycopg2.extensions.connection) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TEMP TABLE IF NOT EXISTS tmp_import_parts (
                sku TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                manufacturer TEXT NOT NULL,
                model TEXT NULL,
                year_from INTEGER NULL,
                year_to INTEGER NULL,
                price_ils NUMERIC(10,2) NULL,
                part_condition TEXT NOT NULL,
                is_available BOOLEAN NOT NULL,
                importer_info TEXT NULL
            ) ON COMMIT PRESERVE ROWS
            """
        )


def get_supplier_id(conn: psycopg2.extensions.connection) -> str:
    with conn.cursor() as cur:
        cur.execute("SELECT id::text FROM suppliers WHERE name = %s", (SUPPLIER_NAME,))
        row = cur.fetchone()
    if not row:
        raise RuntimeError(f"Supplier '{SUPPLIER_NAME}' not found")
    return row[0]


def normalize_row(row: dict[str, Any], sheet_name: str) -> NormalizedRow | None:
    sku = clean_text(row.get("מספר קטלוגי"))
    if not sku:
        return None

    name = clean_text(row.get("תיאור החלק")) or sku
    manufacturer = normalize_manufacturer(row.get("מותג"), sheet_name)
    model, year_from, year_to = parse_model_and_year(row.get("דגם רכב"))
    price_ils = parse_price_ils(row.get("מחיר לצרכן"))
    condition = normalize_part_condition(row.get("סוג מוצר"))
    is_available = normalize_availability(row.get("זמינות מלאי"))
    importer_info = clean_text(row.get("פרטי היבואן"))

    return NormalizedRow(
        sku=sku.strip(),
        name=name,
        manufacturer=manufacturer,
        model=model,
        year_from=year_from,
        year_to=year_to,
        price_ils=price_ils,
        part_condition=condition,
        is_available=is_available,
        importer_info=importer_info,
    )


def load_workbook_rows(file_path: str):
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    print(f"Workbook loaded: {file_path}")
    print(f"Sheets detected: {len(workbook.sheetnames)}")

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        total_rows = max((worksheet.max_row or 1) - 1, 0)
        print(f"Reading sheet: {sheet_name} ({total_rows} rows)")

        row_iter = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iter)
        except StopIteration:
            continue

        headers = [str(col).strip() if col is not None else "" for col in raw_headers]
        for row_idx, row_values in enumerate(row_iter, start=2):
            row_dict: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row_values[idx] if idx < len(row_values) else None
                row_dict[header] = value
            yield sheet_name, row_idx, row_dict

    workbook.close()


def process_batch(
    conn: psycopg2.extensions.connection,
    supplier_id: str,
    batch_rows: list[NormalizedRow],
) -> tuple[int, int, int, int, int]:
    if not batch_rows:
        return 0, 0, 0, 0, 0

    deduped_by_sku: dict[str, NormalizedRow] = {}
    for item in batch_rows:
        deduped_by_sku[item.sku] = item
    rows = list(deduped_by_sku.values())

    values = [
        (
            r.sku,
            r.name,
            r.manufacturer,
            r.model,
            r.year_from,
            r.year_to,
            r.price_ils,
            r.part_condition,
            r.is_available,
            r.importer_info,
        )
        for r in rows
    ]

    with conn.cursor() as cur:
        cur.execute("TRUNCATE TABLE tmp_import_parts")
        execute_values(
            cur,
            """
            INSERT INTO tmp_import_parts (
                sku, name, manufacturer, model, year_from, year_to,
                price_ils, part_condition, is_available, importer_info
            ) VALUES %s
            """,
            values,
            page_size=500,
        )

        cur.execute(
            """
            WITH upserted AS (
                INSERT INTO parts_catalog (
                    id, sku, name, manufacturer, oem_number,
                    online_price_ils, min_price_ils, max_price_ils,
                    part_condition, is_active, needs_oem_lookup,
                    created_at, updated_at
                )
                SELECT
                    gen_random_uuid(),
                    t.sku,
                    t.name,
                    t.manufacturer,
                    t.sku,
                    t.price_ils,
                    t.price_ils,
                    t.price_ils,
                    t.part_condition,
                    TRUE,
                    FALSE,
                    NOW(),
                    NOW()
                FROM tmp_import_parts t
                ON CONFLICT (sku)
                DO UPDATE SET
                    online_price_ils = EXCLUDED.online_price_ils,
                    min_price_ils = EXCLUDED.min_price_ils,
                    max_price_ils = EXCLUDED.max_price_ils,
                    part_condition = EXCLUDED.part_condition,
                    updated_at = NOW()
                RETURNING (xmax = 0) AS inserted
            )
            SELECT
                COALESCE(SUM(CASE WHEN inserted THEN 1 ELSE 0 END), 0)::int,
                COALESCE(SUM(CASE WHEN NOT inserted THEN 1 ELSE 0 END), 0)::int
            FROM upserted
            """
        )
        parts_inserted, parts_updated = cur.fetchone()

        cur.execute(
            """
            WITH src AS (
                SELECT
                    p.id AS part_id,
                    t.sku,
                    t.price_ils,
                    t.is_available
                FROM tmp_import_parts t
                JOIN parts_catalog p ON p.sku = t.sku
            ),
            upserted AS (
                INSERT INTO supplier_parts (
                    id, part_id, supplier_id, supplier_sku,
                    price_usd, price_ils,
                    is_available, availability,
                    last_checked_at, created_at, updated_at
                )
                SELECT
                    gen_random_uuid(),
                    s.part_id,
                    %s::uuid,
                    s.sku,
                    ROUND(COALESCE(s.price_ils, 0) / %s, 2),
                    s.price_ils,
                    s.is_available,
                    CASE WHEN s.is_available THEN 'in_stock' ELSE 'out_of_stock' END,
                    NOW(),
                    NOW(),
                    NOW()
                FROM src s
                ON CONFLICT (part_id, supplier_id)
                DO UPDATE SET
                    supplier_sku = EXCLUDED.supplier_sku,
                    price_usd = EXCLUDED.price_usd,
                    price_ils = EXCLUDED.price_ils,
                    is_available = EXCLUDED.is_available,
                    availability = EXCLUDED.availability,
                    last_checked_at = NOW(),
                    updated_at = NOW()
                RETURNING (xmax = 0) AS inserted
            )
            SELECT
                COALESCE(SUM(CASE WHEN inserted THEN 1 ELSE 0 END), 0)::int,
                COALESCE(SUM(CASE WHEN NOT inserted THEN 1 ELSE 0 END), 0)::int
            FROM upserted
            """,
            (supplier_id, ILS_TO_USD),
        )
        supplier_inserted, supplier_updated = cur.fetchone()

        cur.execute(
            """
            WITH src AS (
                SELECT
                    p.id AS part_id,
                    t.manufacturer,
                    t.model,
                    t.year_from,
                    t.year_to
                FROM tmp_import_parts t
                JOIN parts_catalog p ON p.sku = t.sku
                WHERE t.model IS NOT NULL
                  AND t.year_from IS NOT NULL
            ),
            inserted_fitment AS (
                INSERT INTO part_vehicle_fitment (
                    id, part_id, manufacturer, model, year_from, year_to
                )
                SELECT
                    gen_random_uuid(),
                    s.part_id,
                    s.manufacturer,
                    s.model,
                    s.year_from,
                    s.year_to
                FROM src s
                ON CONFLICT (part_id, manufacturer, model, year_from)
                DO NOTHING
                RETURNING 1
            )
            SELECT COUNT(*)::int FROM inserted_fitment
            """
        )
        fitment_inserted = cur.fetchone()[0]

    return parts_inserted, parts_updated, supplier_inserted, supplier_updated, fitment_inserted


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import merged catalog workbook into catalog tables")
    parser.add_argument("--file", required=True, help="Path to merged_all_final.xlsx")
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE_DEFAULT, help="Batch size (default: 500)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    file_path = args.file

    if not os.path.exists(file_path):
        print(f"ERROR: file not found: {file_path}")
        return 1

    database_url = get_database_url()

    processed = 0
    skipped = 0
    errors = 0
    parts_inserted = 0
    parts_updated = 0
    supplier_inserted = 0
    supplier_updated = 0
    fitment_inserted = 0

    pending_batch: list[NormalizedRow] = []

    conn = psycopg2.connect(database_url)
    try:
        conn.autocommit = False
        ensure_supporting_indexes(conn)
        ensure_temp_table(conn)
        supplier_id = get_supplier_id(conn)
        conn.commit()

        print(f"Connected. Supplier ID: {supplier_id}")
        print(f"Batch size: {args.batch_size}")

        for sheet_name, sheet_row_number, raw_row in load_workbook_rows(file_path):
            processed += 1
            try:
                normalized = normalize_row(raw_row, sheet_name)
                if normalized is None:
                    skipped += 1
                    continue

                pending_batch.append(normalized)

                if len(pending_batch) >= args.batch_size:
                    p_ins, p_upd, s_ins, s_upd, f_ins = process_batch(conn, supplier_id, pending_batch)
                    conn.commit()
                    parts_inserted += p_ins
                    parts_updated += p_upd
                    supplier_inserted += s_ins
                    supplier_updated += s_upd
                    fitment_inserted += f_ins
                    pending_batch.clear()

                if processed % PROGRESS_EVERY == 0:
                    print(
                        f"Progress {processed}: "
                        f"parts inserted={parts_inserted}, parts updated={parts_updated}, "
                        f"supplier inserted={supplier_inserted}, supplier updated={supplier_updated}, "
                        f"fitment inserted={fitment_inserted}, skipped={skipped}, errors={errors}"
                    )

            except Exception as row_exc:
                errors += 1
                print(f"Row error at sheet={sheet_name}, row={sheet_row_number}: {row_exc}")

        if pending_batch:
            p_ins, p_upd, s_ins, s_upd, f_ins = process_batch(conn, supplier_id, pending_batch)
            conn.commit()
            parts_inserted += p_ins
            parts_updated += p_upd
            supplier_inserted += s_ins
            supplier_updated += s_upd
            fitment_inserted += f_ins
            pending_batch.clear()

        print("Import completed.")
        print(
            "SUMMARY "
            f"processed={processed} "
            f"parts_inserted={parts_inserted} "
            f"parts_updated={parts_updated} "
            f"supplier_inserted={supplier_inserted} "
            f"supplier_updated={supplier_updated} "
            f"fitment_inserted={fitment_inserted} "
            f"skipped={skipped} "
            f"errors={errors}"
        )
        return 0

    except Exception as exc:
        conn.rollback()
        print(f"FATAL: {exc}")
        return 2

    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())