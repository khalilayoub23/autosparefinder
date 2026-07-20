"""
DB Update Agent — backend/db_update_agent.py

Runs a set of autonomous cleaning / normalisation tasks against the catalogue DB.
Each task is idempotent: re-running it is always safe.

Tasks
-----
1.  clean_part_names          – strip trailing car-model suffixes from part names
2.  normalize_part_types      – unify to "Original" / "OEM" / "Aftermarket"
3.  normalize_categories      – map variants to shared canonical Hebrew categories
4.  normalize_availability    – unify to "in_stock" / "out_of_stock" / "on_order"
5.  fix_base_prices           – ensure base_price = supplier min + 18 % VAT markup
6.  flag_fake_skus            – set needs_oem_lookup=True for auto-generated SKUs
7.  fill_car_brands           – seed il_importer / warranty_* for known makes
8.  sync_manufacturer_registries – keep car/truck brand registries clean + logos
9.  run_all_tasks             – orchestrator that runs core tasks and returns a report dict

On-demand tasks (NOT in run_all_tasks — must be triggered explicitly):
9.  populate_supplier_parts   – link every active part to every active supplier
                                (ON CONFLICT DO NOTHING; safe to re-run)
10. validate_migrations       – pre-flight safety scan of all Alembic migration
                                files for patterns that can cause downtime

Admin endpoints call run_all_tasks or individual tasks through get_db.
run_agent_background_loop()  – optional periodic loop (disabled by default).
"""
# DATA QUALITY PIPELINE OWNER: DB Update Agent — normalises and enriches parts_catalog

from __future__ import annotations

import asyncio
import contextlib
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
import httpx
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from currency_rate import get_usd_to_ils_rate
from resilience import job_registry_start, job_registry_finish, job_heartbeat
from manufacturer_normalization import PARTS_BRANDS, canonicalize_vehicle_model_for_manufacturer
from manufacturer_normalization import normalize_vehicle_model_name, normalize_vehicle_submodel_name
from manufacturer_normalization import normalize_manufacturer_name
from categories import CATEGORY_MAP as SHARED_CATEGORY_MAP
from agent_todo_utils import get_active_agent_todos, extract_todo_task_names

logger = logging.getLogger("db_update_agent")
# Ensure the logger outputs to stderr (docker logs captures it) if no handler
# is configured on the root logger. This is a no-op when uvicorn/gunicorn sets
# up its own logging.
if not logging.root.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        stream=__import__("sys").stderr,
    )

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VAT = 0.18          # Israeli VAT rate
ILS_PER_USD = 3.72  # fallback – overridden at runtime from system_settings

# Canonical categories come from the shared taxonomy module.
CANONICAL_CATEGORIES: List[str] = list(SHARED_CATEGORY_MAP.keys()) + ["כללי"]

_HE_CHAR_RE = re.compile(r"[֐-׿]")
_ALIAS_MIN_SCORE = float(os.getenv("HE_ALIAS_MIN_SCORE", "0.88"))
_ALIAS_MIN_MARGIN = float(os.getenv("HE_ALIAS_MIN_MARGIN", "0.06"))


def _real_data_only_enabled() -> bool:
    """True when synthetic data generation is blocked (AI-generated parts, fake supplier links)."""
    env_name = (os.getenv("ENVIRONMENT", "development") or "development").strip().lower()
    default_flag = "1" if env_name == "production" else "0"
    raw = (os.getenv("REAL_DATA_ONLY", default_flag) or default_flag).strip().lower()
    return raw in {"1", "true", "yes", "on"}


def _rex_harvest_enabled() -> bool:
    """True when REX may trigger real-source discovery. Separate from synthetic-data guard."""
    raw = (os.getenv("REX_HARVEST_ENABLED", "true") or "true").strip().lower()
    return raw in {"1", "true", "yes", "on"}


_HE_TO_LATIN = {
    "א": "a", "ב": "b", "ג": "g", "ד": "d", "ה": "h", "ו": "v", "ז": "z", "ח": "h", "ט": "t",
    "י": "y", "כ": "k", "ך": "k", "ל": "l", "מ": "m", "ם": "m", "נ": "n", "ן": "n", "ס": "s",
    "ע": "a", "פ": "p", "ף": "p", "צ": "ts", "ץ": "ts", "ק": "k", "ר": "r", "ש": "sh", "ת": "t",
    " ": " ", "׳": "", "'": "",
}

_MANUAL_HEBREW_CANDIDATES: Dict[str, List[str]] = {
    "chevrolet": ["שברולט"],
    "lamborghini": ["למבורגיני"],
    "porsche": ["פורשה"],
    "rollsroyce": ["רולס רויס"],
    "mclaren": ["מקלארן"],
    "hummer": ["האמר"],
    "lincoln": ["לינקולן"],
    "lotus": ["לוטוס"],
    "mini": ["מיני"],
    "morgan": ["מורגן"],
    "pontiac": ["פונטיאק"],
    "piaggio": ["פיאגו"],
    "polaris": ["פולריס"],
    "haval": ["האבל"],
    "changan": ["צ אנגאן", "צאנגאן"],
    "dongfeng": ["דונגפנג"],
    "foton": ["פוטון"],
    "jac": ["ג אק", "גאק", "ג אקו", "גאקו"],
    "jmc": ["ג יי אם סי", "גי אם סי"],
    "baic": ["באיק"],
    "hongqi": ["הונג צ י", "הונגצי"],
    "voyah": ["וויה"],
    "wey": ["ווי"],
    "ora": ["אורה"],
    "genesis": ["ג נסיס", "גנסיס"],
    "liauto": ["לי אוטו"],
    "jaecoo": ["ג ייקו", "גייקו"],
    "leapmotor": ["ליפמוטור"],
    "skywell": ["סקייוול"],
    "seres": ["סרס"],
    "omoda": ["אומודה"],
    "neta": ["נטא"],
    "lti": ["אל טי איי"],
    "im": ["איי אם"],
    "aiways": ["איווייז"],
    "arcfox": ["ארקפוקס"],
    "aion": ["איון"],
    "changyang": ["צ אנג יאנג", "צאנג יאנג"],
    "dayon": ["דאיון"],
    "dfsk": ["די אף אס קיי"],
    "detomaso": ["דה תומאסו"],
    "exlantix": ["אקסלנטיקס"],
    "farizon": ["פאריזון"],
    "forthing": ["פורתינג"],
    "giayuan": ["גיאיוואן"],
    "ineos": ["אינאוס"],
    "karma": ["קארמה"],
    "levc": ["אל אי וי סי"],
    "lynkco": ["לינק אנד קו"],
    "lynxis": ["לינקסיס"],
    "movimatic": ["מובימטיק"],
    "sentro": ["סנטרו"],
    "modern": ["מודרן"],
    "yudo": ["יודו"],
    "swm": ["אס דאבליו אם"],
    "tic": ["טי איי סי"],
    "wmi": ["דאבליו אם איי"],
    "xev": ["אקס אי וי"],
    "gms": ["גי אם אס"],
    "mcc": ["אם סי סי"],
}

_TRANSPORT_FREQ_PATHS = [
    Path(__file__).parent / "data" / "rex_transport_manufacturer_frequency.json",
    Path("/app/data/rex_transport_manufacturer_frequency.json"),
]


# Mapping of synonyms → canonical category
CATEGORY_MAP: Dict[str, str] = {
    # brakes
    "brakes": "בלמים",
    "brake": "בלמים",
    "בלם": "בלמים",
    # wheels / tyres
    "wheels": "גלגלים וצמיגים",
    "tyres": "גלגלים וצמיגים",
    "tires": "גלגלים וצמיגים",
    "גלגלים": "גלגלים וצמיגים",
    "צמיגים": "גלגלים וצמיגים",
    # fuel
    "fuel": "דלק",
    "fuel system": "דלק",
    "מערכת דלק": "דלק",
    # steering
    "steering": "היגוי",
    # electrical
    "electrical": "חשמל רכב",
    "electric": "חשמל רכב",
    "electronics": "חשמל רכב",
    "חשמל": "חשמל רכב",
    # general / misc
    "general": "כללי",
    "misc": "כללי",
    "miscellaneous": "כללי",
    "other": "כללי",
    "אחר": "כללי",
    # wipers
    "wipers": "מגבים",
    "wiper": "מגבים",
    "מגב": "מגבים",
    # ac / climate
    "ac": "מיזוג",
    "air conditioning": "מיזוג",
    "climate": "מיזוג",
    "hvac": "מיזוג",
    # engine
    "engine": "מנוע",
    "motor": "מנוע",
    # suspension
    "suspension": "מתלה",
    "שוקים": "מתלה",
    # body
    "body": "פחיין ומרכב",
    "bodywork": "פחיין ומרכב",
    "מרכב": "פחיין ומרכב",
    # interior
    "interior": "ריפוד ופנים",
    "upholstery": "ריפוד ופנים",
    # chains / belts
    "belts": "שרשראות ורצועות",
    "chains": "שרשראות ורצועות",
    "belt": "שרשראות ורצועות",
    "timing": "שרשראות ורצועות",
    "רצועות": "שרשראות ורצועות",
    # lighting
    "lighting": "תאורה",
    "lights": "תאורה",
    "light": "תאורה",
    "lamps": "תאורה",
    "תאור": "תאורה",

    # reference taxonomy (Autodoc-like families)
    "tyres and related products": "גלגלים וצמיגים",
    "tires and related products": "גלגלים וצמיגים",
    "brake system": "בלמים",
    "filters": "כללי",
    "oils and fluids": "כללי",
    "body": "פחיין ומרכב",
    "suspension and arms": "מתלה",
    "turbocharger": "מנוע",
    "air conditioning": "מיזוג",
    "fuel supply system": "דלק",
    "steering": "היגוי",
    "transmission": "כללי",
    "fasteners": "כללי",
    "pipes and hoses": "מנוע",
    "gaskets and sealing rings": "מנוע",
    "damping": "מתלה",
    "windscreen cleaning system": "מגבים",
    "exhaust system": "כללי",
    "accessories": "כללי",
    "ignition and glowplug system": "חשמל רכב",
    "tuning": "כללי",
    "interior and comfort": "ריפוד ופנים",
    "belts, chains, rollers": "שרשראות ורצועות",
    "exhaust gas recirculation": "מנוע",
    "towbar / parts": "פחיין ומרכב",
    "towbar": "פחיין ומרכב",
    "heater": "מיזוג",
    "bearings": "מתלה",
    "air suspension": "מתלה",
    "sensors, relays, control units": "חשמל רכב",
    "repair kits": "כללי",
    "propshafts and differentials": "מתלה",
    "electrics": "חשמל רכב",
    "engine cooling system": "מנוע",
    "clutch / parts": "מנוע",
    "drive shaft and cv joint": "מתלה",
    "auto detailing & car care": "כללי",
    "tools": "כללי",

    # non-car families -> keep out of core car categories
    "motorcycle accessories": "כללי",
    "motorcycle clothing": "כללי",
    "motorcycle helmets": "כללי",

    # car-parts.ie slug-style categories (from importer pipeline)
    "service-general": "כללי",
    "body-exterior": "גוף הרכב",
    "electrical-sensors": "חשמל ואלקטרוניקה",
    "air-conditioning-heating": "מזגן וחימום",
    "suspension-steering": "מתלה",
    "interior-comfort": "פנים הרכב",
    "fuel-air": "מערכת דלק",
    "wheels-bearings": "גלגלים וצמיגים",
    "clutch-drivetrain": "מצמד",
    "wipers-washers": "שמשות ומגבים",
    "cooling": "מערכת קירור",
    "exhaust": "מערכת פליטה",
    "gearbox": "תיבת הילוכים",
    "belts-chains": "רצועות תזמון",
    "safety-systems": "בלמים",
    "fluids": "כללי",
    "filters": "כללי",

    # title-case variants from oempartsonline / champion motors importers
    "General Parts": "כללי",
    "Body Parts": "גוף הרכב",
    "Engine Parts": "מנוע",
    "Electrical": "חשמל ואלקטרוניקה",
    "Brakes": "בלמים",
    "Suspension": "מתלה",
    "Service & General": "כללי",
    "Fuel System": "מערכת דלק",
    "Accessories": "כללי",
    "Steering": "היגוי",
    "Transmission": "תיבת הילוכים",
    "Cooling": "מערכת קירור",
    "Exhaust": "מערכת פליטה",
    "Interior": "פנים הרכב",
    "Lighting": "תאורה",
    "Filters": "כללי",
    "Fluids": "כללי",
    "Belts & Chains": "רצועות תזמון",
    "Wheels & Bearings": "גלגלים וצמיגים",
    "Air Conditioning": "מזגן וחימום",
    "Body & Exterior": "גוף הרכב",
    "Clutch & Drivetrain": "מצמד",
    "Wipers & Washers": "שמשות ומגבים",
    "Safety Systems": "בלמים",
}

# Legacy category aliases remapped to the shared 28-category taxonomy
CATEGORY_NAME_REMAP: Dict[str, str] = {
    "דלק": "מערכת דלק",
    "חשמל רכב": "חשמל ואלקטרוניקה",
    "מגבים": "שמשות ומגבים",
    "מיזוג": "מזגן וחימום",
    "פחיין ומרכב": "גוף הרכב",
    "ריפוד ופנים": "פנים הרכב",
    "שרשראות ורצועות": "רצועות תזמון",
}

# Normalisation map for part_type
PART_TYPE_MAP: Dict[str, str] = {
    "original": "Original",
    "oem_original": "Original",
    "genuine": "Original",
    "מקורי": "Original",
    "מקורימקורי": "Original",
    "oem": "OEM",
    "oem_equivalent": "OEM",
    "oe": "OEM",
    "aftermarket": "Aftermarket",
    "after market": "Aftermarket",
    "generic": "Aftermarket",
    "third party": "Aftermarket",
    "תחליפי": "Aftermarket",
    "חליפיחליפי": "Aftermarket",
    "תחליפיתחליפי": "Aftermarket",
    "שוק משני": "Aftermarket",
}

# Normalisation map for availability
AVAILABILITY_MAP: Dict[str, str] = {
    "in stock": "in_stock",
    "instock": "in_stock",
    "in-stock": "in_stock",
    "available": "in_stock",
    "במלאי": "in_stock",
    "out of stock": "out_of_stock",
    "outofstock": "out_of_stock",
    "out-of-stock": "out_of_stock",
    "unavailable": "out_of_stock",
    "אזל": "out_of_stock",
    "אין במלאי": "out_of_stock",
    "on order": "on_order",
    "onorder": "on_order",
    "on-order": "on_order",
    "order": "on_order",
    "להזמנה": "on_order",
    "בהזמנה": "on_order",
}

# Known Israeli importers + warranty info  {brand_lower: (importer, years, km, notes)}
BRAND_IMPORTER_MAP: Dict[str, Tuple[str, int, Optional[int], str]] = {
    "toyota":   ("Champion Motors",        3, 100_000, ""),
    "lexus":    ("Champion Motors",        3, 100_000, ""),
    "bmw":      ("PREMIUM MOTORS",         2, None,    "Unlimited km"),
    "mini":     ("PREMIUM MOTORS",         2, None,    "Unlimited km"),
    "volkswagen": ("Dubek",                2, None,    "Unlimited km"),
    "vw":       ("Dubek",                  2, None,    "Unlimited km"),
    "audi":     ("Dubek",                  2, None,    "Unlimited km"),
    "skoda":    ("Dubek",                  2, None,    "Unlimited km"),
    "škoda":    ("Dubek",                  2, None,    "Unlimited km"),
    "seat":     ("Dubek",                  2, None,    "Unlimited km"),
    "hyundai":  ("Colmobil",               5, 150_000, ""),
    "kia":      ("Colmobil",               5, 150_000, ""),
    "mazda":    ("Delek Motors",           3, None,    "Unlimited km"),
    "honda":    ("Car Trading Company",    3, 100_000, ""),
    "nissan":   ("Carasso Motors",         3, None,    "Unlimited km"),
    "infiniti": ("Carasso Motors",         3, None,    "Unlimited km"),
    "ford":     ("Shlomo Sixt",            3, 100_000, ""),
    "subaru":   ("Inovision",              3, 100_000, ""),
    "mitsubishi": ("Inbar Motors",         3, 100_000, ""),
    "jeep":     ("Auto Hadar",             3, 60_000,  ""),
    "chrysler": ("Auto Hadar",             3, 60_000,  ""),
    "dodge":    ("Auto Hadar",             3, 60_000,  ""),
    "mercedes-benz": ("Authorized Importers", 2, None, "Unlimited km"),
    "mercedes": ("Authorized Importers",   2, None,    "Unlimited km"),
    "volvo":    ("Volvo Cars Israel",      3, None,    "Unlimited km"),
    "peugeot":  ("Citroen Israel",         2, 60_000,  ""),
    "citroen":  ("Citroen Israel",         2, 60_000,  ""),
    "renault":  ("Renault Israel",         2, 60_000,  ""),
    "opel":     ("General Motors Israel",  2, 60_000,  ""),
    "chevrolet":("General Motors Israel",  2, 60_000,  ""),
    "fiat":     ("Fiat Israel",            2, 60_000,  ""),
    "suzuki":   ("Sela Motors",            3, 100_000, ""),
    "daihatsu": ("Sela Motors",            3, 100_000, ""),
    "tesla":    ("Tesla Israel",           4, None,    "Unlimited km / 8yr battery"),
}

# Deterministic fallback for active brands not yet curated in BRAND_IMPORTER_MAP.
# Keeps metadata non-null for search/admin UX while signaling that values need verification.
DEFAULT_BRAND_METADATA: Tuple[str, int, Optional[int], str] = (
    "Pending Importer Verification",
    2,
    100_000,
    "Auto-filled default metadata; verify official importer and warranty policy",
)

# Curated generation/platform year ranges used for strict workbook fitment.
GENERATION_YEAR_RULES: Dict[Tuple[str, str, str], Tuple[int, int]] = {
    ("citroen", "berlingo", "b9"): (2008, 2018),
    ("citroen", "berlingo", "k9"): (2018, 2027),
    ("citroen", "berlingo", "k9 acc"): (2018, 2027),
    ("peugeot", "partner", "b9"): (2008, 2018),
    ("peugeot", "partner", "k9"): (2018, 2027),
    ("peugeot", "partner", "k9 acc"): (2018, 2027),
}

# Regex patterns that identify auto-generated / fake SKUs
_FAKE_SKU_PATTERNS = [
    re.compile(r"^[A-Z]{2,6}-[A-Z]{2,6}-\d{3,6}$"),        # OIL-TOY-001
    re.compile(r"^PART-\d+$"),                                # PART-12345
    re.compile(r"^AUTO-[A-Z]+-\d+$"),                        # AUTO-BRK-001
    re.compile(r"^TEMP[-_]\d+$", re.IGNORECASE),             # TEMP-001
    re.compile(r"^TBD[-_]?\d*$", re.IGNORECASE),             # TBD / TBD-1
    re.compile(r"^SKU[-_]\d+$", re.IGNORECASE),              # SKU-001
    re.compile(r"^[A-Z]{1,3}\d{1,3}$"),                      # A1 / AB12 (too short to be real)
    re.compile(r"^0+$"),                                      # 000
]

# Pattern to strip trailing car-model suffix from part names
# e.g., " - Toyota Corolla 2015-2023"  /  " (Ford Focus 2018)"
_NAME_SUFFIX_RE = re.compile(
    r"\s*[-–(]\s*"                                   # separator
    r"([A-Za-zא-ת]+\s+[A-Za-z0-9 À-öø-ÿ]+)"        # brand + model
    r"(?:\s+\d{4}(?:\s*[-–]\s*\d{4})?)?[)]*\s*$",  # optional year range
    re.UNICODE,
)
_TRAILING_PAREN_RE = re.compile(r"\s*\(([^()]*)\)\s*$")
_TRAILING_DASH_RE = re.compile(r"\s*[-–]\s*([^()]{2,120})\s*$")
_YEAR_TOKEN_RE = re.compile(r"\b(?:19|20)\d{2}(?:\s*[-–/]\s*(?:19|20)\d{2})?\b")
_COMPAT_TOKEN_RE = re.compile(
    r"\b(?:RH|LH|LT|RT|ACC|AT|A/T|4X4|AWD|FWD|RWD|"
    r"[A-Z]{1,4}\d{2,}[A-Z0-9-]*|[A-Z]{2,6})\b",
    re.IGNORECASE,
)
_LEADING_TAG_RE = re.compile(r"^\(\s*([A-Za-zא-ת0-9+./\\-]{1,24})\s*\)\s+")
_EMPTY_PARENS_RE = re.compile(r"\(\s*\)")
_NOISE_QUOTE_RE = re.compile(r"[׳’`]+")
_PARENS_GROUP_RE = re.compile(r"\(\s*([^()]{1,120})\s*\)")
_DASH_SEP_RE = re.compile(r"\s[-–]\s")
_TRAILING_TOKEN_DASH_RE = re.compile(r"\b([A-Za-zא-ת0-9]{1,24})-(?=\s|$)")
_YEAR_DANGLING_RANGE_RE = re.compile(r"\b((?:19|20)\d{2})\s*[-–]\s*(?=$|\b)")

# Task 6 dictionary of known reversed French automotive terms.
REVERSED_FRENCH_TERMS: Dict[str, str] = {
    "TNIOJ": "JOINT",
    "SIALER": "RELAIS",
    "EFARG": "AGRAFE",
    "EFARGA": "AGRAFE",
    "EHCUOTRAC": "CARTOUCHE",
    "RUETARUTBO": "OBTURATEUR",
    "SRUETARUTBO": "OBTURATEURS",
    "TENISSUOC": "COUSSINET",
    "RUETCETORP": "PROTECTEUR",
    "TEHCORC": "CROCHET",
    "ERIASSECEN": "NECESSAIRE",
    "ELLIRG": "GRILLE",
    "TOPAC": "CAPOT",
    "EGUAJ": "JAUGE",
    "LIUH": "HUILE",
    "TNAYOV": "VOYANT",
    "ELLEPUOC": "COUPELLE",
    "OCED": "DECO",
    "NONRAHC": "CHARNON",
    "IUQOC": "COQUI",
    "REITIO": "OITIER",
    "REITIOB": "BOITIER",
    "EDNA": "ANED",
    "TCELLOC": "COLLECT",
    "ELUSPA": "CAPSULE",
    "ETRO": "ORTE",
    "ETROP": "PORTE",
    "LCAR": "RACL",
    "PILC": "CLIP",
    "RTED": "DETR",
    "NECS": "SCEN",
    "EGRU": "URGE",
    "ELGNIRT": "TRINGLER",
    "ETNED": "DENTE",
    "TCES": "SECT",
    "TNARI": "IRANT",
    "ELLUG": "GULLE",
    "ERTIV": "VITRE",
    "EBUT": "TUBE",
    "EUGAB": "BAGUE",
    "TROPPUS": "SUPPORT",
    "RUEVILOJNE": "ENJOLIVEUR",
    "VILOJNE": "ENJOLIV",
    "ELCREVUOC": "COUVERCLE",
    "ESIOTERTNE": "ENTRETOISE",
    "REILAP": "PALIER",
    "TNAVA": "AVANT",
    "ELLEDNOR": "RONDELLE",
    "EETUB": "BUTEE",
    "ELLITSAP": "PASTILLE",
    "RUETCEJORP": "PROJECTEUR",
    "RUETACIDNI": "INDICATEUR",
    "RUETCATNOC": "CONTACTEUR",
    "UAEDNAB": "BANDEAU",
    "RUELCIG": "GICLEUR",
    "EMMARGONOM": "MONOGRAMME",
    "ELLIPUOG": "GOUPILLE",
    "NOHCUOB": "BOUCHON",
    "EILUOP": "POULIE",
    "ELOT": "TOLE",
    "QITPO": "OPTIQUE",
    "ETPURRETN": "INTERRUPTEUR",
    "UETATUMMO": "COMMUTATEUR",
    "TNIO": "JOINT",
}
TASK6_LATIN_PREFIX_MAP: Dict[str, str] = {
    "SBA": "ABS",
    "UAE": "EAU",
    "RGE": "EGR",
    "XEH": "HEX",
    "TA": "AT",
    "ED": "DE",
    "RA": "AR",
    "OC": "CO",
    "EF": "FE",
    "TROPPUS": "SUPPORT",
    "TEHCORC": "CROCHET",
    "EETUB": "BUTEE",
    "ELPUOC": "COUPLE",
    "UAEDNAB": "BANDEAU",
    "ETROP": "PORTE",
    "REIVEL": "LEVIER",
    "EFARGA": "AGRAFE",
    "TNIOJ": "JOINT",
    "ENITALP": "PLATINE",
    "TCATNOC": "CONTACT",
    "TNEMELE": "ELEMENT",
    "TEILAP": "PALETTE",
    "RACL": "RACLEUR",
    "UEF": "FEU",
}
TASK6_LATIN_PREFIX_MAP.update(REVERSED_FRENCH_TERMS)
_TASK6_SQL_PREFIX_PATTERN = r"^(" + "|".join(re.escape(k) for k in sorted(TASK6_LATIN_PREFIX_MAP)) + r")(?:[^A-Za-z]|$)"

# -------------------------------------------------------------------------
# Shared state
# -------------------------------------------------------------------------
_last_report: Dict[str, Any] = {}
_agent_running: bool = False


# =========================================================================
# Helpers
# =========================================================================

async def _get_ils_rate(db: AsyncSession) -> float:
    """Read current ILS/USD rate from system_settings (or fall back to default)."""
    return await get_usd_to_ils_rate(db, fallback=ILS_PER_USD)


def _is_fake_sku(sku: str) -> bool:
    return any(p.match(sku) for p in _FAKE_SKU_PATTERNS)


def _normalize_part_type(raw: str) -> Optional[str]:
    return PART_TYPE_MAP.get(raw.strip().lower())


def _normalize_category(raw: str) -> Optional[str]:
    raw_stripped = raw.strip()
    if raw_stripped in CANONICAL_CATEGORIES:
        return None  # already canonical
    if raw_stripped in CATEGORY_NAME_REMAP:
        return CATEGORY_NAME_REMAP[raw_stripped]
    mapped = CATEGORY_MAP.get(raw_stripped.lower())
    if not mapped:
        return None
    return CATEGORY_NAME_REMAP.get(mapped, mapped)


def _normalize_availability(raw: str) -> Optional[str]:
    return AVAILABILITY_MAP.get(raw.strip().lower())


def _clean_vehicle_model(value: Optional[str]) -> str:
    """Normalize model names extracted from catalog compatibility blobs."""
    return normalize_vehicle_model_name(value)


def _drop_unbalanced_parentheses(value: str) -> str:
    out: List[str] = []
    stack: List[int] = []
    for ch in value:
        if ch == "(":
            stack.append(len(out))
            out.append(ch)
            continue
        if ch == ")":
            if stack:
                stack.pop()
                out.append(ch)
            continue
        out.append(ch)

    for idx in reversed(stack):
        if 0 <= idx < len(out):
            out.pop(idx)
    return "".join(out)


def _looks_like_vehicle_suffix(value: str) -> bool:
    s = re.sub(r"\s+", " ", (value or "").strip())
    if not s:
        return False

    has_year = bool(_YEAR_TOKEN_RE.search(s))
    has_compat = bool(_COMPAT_TOKEN_RE.search(s))
    has_mixed_alnum = bool(re.search(r"(?=.*[A-Za-z])(?=.*\d)", s))
    token_count = len(s.split())
    ascii_letters = len(re.findall(r"[A-Za-z]", s))

    return bool(
        has_year
        or has_compat
        or has_mixed_alnum
        or (ascii_letters >= 3 and 1 <= token_count <= 7)
    )


def _normalize_part_name_punctuation(value: str) -> str:
    s = (value or "").replace(" ", " ")
    s = s.replace("–", "-").replace("—", "-")

    s = re.sub(r"([A-Za-zא-ת0-9])\(", r"\1 (", s)
    s = re.sub(r"\)([A-Za-zא-ת0-9])", r") \1", s)
    s = _EMPTY_PARENS_RE.sub(" ", s)

    for _ in range(2):
        m = _LEADING_TAG_RE.match(s)
        if not m:
            break
        tag = m.group(1).strip()
        rest = s[m.end():].lstrip()
        s = f"{tag} {rest}" if rest else tag

    s = _drop_unbalanced_parentheses(s)

    s = re.sub(r"^[\s'\"`-]+", "", s)
    s = re.sub(r"[\s'\"`-]+$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _strip_trailing_vehicle_suffix(value: str) -> str:
    s = value
    for _ in range(3):
        changed = False

        m = _TRAILING_PAREN_RE.search(s)
        if m and _looks_like_vehicle_suffix(m.group(1)):
            base = s[:m.start()].strip()
            if len(base) >= 3:
                s = base
                changed = True

        m = _TRAILING_DASH_RE.search(s)
        if m and _looks_like_vehicle_suffix(m.group(1)):
            base = s[:m.start()].strip()
            if len(base) >= 3:
                s = base
                changed = True

        if not changed:
            break
        s = _normalize_part_name_punctuation(s)

    return s


def _unwrap_parenthetical_groups(value: str) -> str:
    def _repl(match: re.Match[str]) -> str:
        inner = (match.group(1) or "").strip()
        if not inner:
            return " "
        inner = _NOISE_QUOTE_RE.sub("", inner)
        inner = _YEAR_DANGLING_RANGE_RE.sub(r"\1", inner)
        inner = _TRAILING_TOKEN_DASH_RE.sub(r"\1", inner)
        inner = re.sub(r"\s+", " ", inner).strip(" -.,;:+")
        if not inner:
            return " "
        return f" {inner} "

    return _PARENS_GROUP_RE.sub(_repl, value)


def _normalize_dash_separators(value: str) -> str:
    parts = _DASH_SEP_RE.split(value)
    if len(parts) <= 1:
        return value

    head = (parts[0] or "").strip()
    if not head:
        head = (parts[1] or "").strip() if len(parts) > 1 else ""

    merged = head
    for tail in parts[1:]:
        t = (tail or "").strip()
        if not t:
            continue
        if _looks_like_vehicle_suffix(t):
            continue
        merged = f"{merged} {t}".strip()
    return merged


def _clean_part_name_phase2(value: str) -> str:
    s = value
    s = _NOISE_QUOTE_RE.sub("", s)
    s = _unwrap_parenthetical_groups(s)
    s = _normalize_dash_separators(s)
    s = _YEAR_DANGLING_RANGE_RE.sub(r"\1", s)
    s = _TRAILING_TOKEN_DASH_RE.sub(r"\1", s)
    s = re.sub(r"\s*\+\s*", "+", s)
    s = _normalize_part_name_punctuation(s)
    return s


def _reverse_latin_prefix(name: str) -> str:
    tokens = str(name or "").split()
    fixed: List[str] = []
    in_latin_prefix = True

    for tok in tokens:
        if in_latin_prefix and re.match(r"^[A-Za-z.]+$", tok):
            core = tok.rstrip(".")
            lookup = core.upper()
            mapped = TASK6_LATIN_PREFIX_MAP.get(lookup)
            if mapped:
                suffix = tok[len(core):]
                fixed.append(mapped + suffix)
            else:
                fixed.append(tok)
            continue

        in_latin_prefix = False
        fixed.append(tok)

    return " ".join(fixed)


def _apply_task6_rules(name: str) -> str:
    if not name:
        return name

    value = str(name)
    # Fix Hebrew/Latin concatenation boundaries.
    value = re.sub(r"([א-ת])([A-Za-z0-9])", r"\1 \2", value)
    value = re.sub(r"([A-Za-z0-9])([א-ת])", r"\1 \2", value)
    value = re.sub(r"([א-ת])[\"'`׳’]+([A-Za-z0-9])", r"\1 \2", value)
    value = re.sub(r"([A-Za-z0-9])[\"'`׳’]+([א-ת])", r"\1 \2", value)

    value = _reverse_latin_prefix(value)
    value = re.sub(r" {2,}", " ", value).strip()
    return value


def _clean_part_name_value(name: str) -> str:
    original = str(name or "")
    cleaned = _apply_task6_rules(original)
    cleaned = _normalize_part_name_punctuation(cleaned)
    cleaned = _strip_trailing_vehicle_suffix(cleaned)
    cleaned = _clean_part_name_phase2(cleaned)

    legacy = _NAME_SUFFIX_RE.search(cleaned)
    if legacy:
        base = cleaned[:legacy.start()].strip()
        if len(base) >= 3:
            cleaned = base

    cleaned = _apply_task6_rules(cleaned)
    cleaned = _normalize_part_name_punctuation(cleaned)

    if len(cleaned) < 2:
        return original.strip()
    return cleaned


async def ensure_part_vehicle_fitment_table(db: AsyncSession) -> None:
    """Create the scraped fitment table on demand when runtime code still relies on it."""
    await db.execute(text("CREATE EXTENSION IF NOT EXISTS pgcrypto"))
    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS part_vehicle_fitment (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            part_id UUID NOT NULL REFERENCES parts_catalog(id) ON DELETE CASCADE,
            manufacturer VARCHAR(100) NOT NULL,
            model VARCHAR(100) NOT NULL,
            year_from INTEGER NOT NULL,
            year_to INTEGER NULL,
            engine_type VARCHAR(50) NULL,
            transmission VARCHAR(50) NULL,
            notes TEXT NULL,
            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
        )
    """))
    await db.execute(text("ALTER TABLE part_vehicle_fitment ADD COLUMN IF NOT EXISTS tozeret_cd INTEGER NULL"))
    await db.execute(text("ALTER TABLE part_vehicle_fitment ADD COLUMN IF NOT EXISTS degem_cd INTEGER NULL"))
    await db.execute(text("ALTER TABLE part_vehicle_fitment ADD COLUMN IF NOT EXISTS shnat_yitzur INTEGER NULL"))
    await db.execute(text("ALTER TABLE part_vehicle_fitment ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW()"))

    # Collapse duplicate keys before creating the unique index used by ON CONFLICT inserts.
    await db.execute(text("""
        DELETE FROM part_vehicle_fitment a
        USING part_vehicle_fitment b
        WHERE a.ctid < b.ctid
          AND a.part_id = b.part_id
          AND a.manufacturer = b.manufacturer
          AND a.model = b.model
          AND a.year_from = b.year_from
    """))

    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_fitment_part_id ON part_vehicle_fitment (part_id)"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_fitment_mfr_model ON part_vehicle_fitment (manufacturer, model)"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_fitment_years ON part_vehicle_fitment (year_from, year_to)"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_pvf_tozeret_degem ON part_vehicle_fitment (tozeret_cd, degem_cd, shnat_yitzur)"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_pvf_manufacturer_model ON part_vehicle_fitment (manufacturer, model, year_from, year_to)"))
    await db.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uix_pvf_part_mfr_model_year_from ON part_vehicle_fitment (part_id, manufacturer, model, year_from)"))
    await db.commit()


# =========================================================================
# Task 1 – Clean part names
# =========================================================================

async def clean_part_names(db: AsyncSession) -> Dict[str, Any]:
    """
    Task 6 — clean part names in controlled batches.

    Root-fix goals:
    - split Hebrew/Latin concatenations
    - fix known reversed French prefixes
    - collapse whitespace noise
    """
    t0 = time.monotonic()
    rows_updated = 0
    rows_checked = 0
    batches_run = 0
    no_update_batches = 0
    samples: List[Dict[str, str]] = []

    batch_size = max(50, int(os.getenv("TASK6_PART_NAME_BATCH_SIZE", "340")))
    max_batches = max(1, int(os.getenv("TASK6_PART_NAME_MAX_BATCHES", "2000")))

    select_sql = text(
        """
        SELECT id, name
        FROM parts_catalog
        WHERE is_active = TRUE
          AND name IS NOT NULL
          AND (
              name ~ '[א-ת][A-Za-z0-9]'
              OR name ~ '[A-Za-z0-9][א-ת]'
              OR name ~ '[א-ת]["''`׳’]+[A-Za-z0-9]'
              OR name ~ '[A-Za-z0-9]["''`׳’]+[א-ת]'
              OR name ~ :prefix_pattern
              OR name ~ '\\s{2,}'
              OR name != TRIM(name)
          )
        ORDER BY RANDOM()
        LIMIT :batch_size
        """
    )

    update_sql = text(
        """
        UPDATE parts_catalog
        SET name = :name,
            updated_at = NOW()
        WHERE id = :id
        """
    )

    try:
        for _ in range(max_batches):
            result = await db.execute(
                select_sql,
                {
                    "prefix_pattern": _TASK6_SQL_PREFIX_PATTERN,
                    "batch_size": batch_size,
                },
            )
            rows = result.fetchall()
            if not rows:
                break

            rows_checked += len(rows)
            batches_run += 1

            updates: List[Dict[str, Any]] = []
            for part_id, raw_name in rows:
                source = str(raw_name or "").strip()
                if not source:
                    continue
                cleaned = _clean_part_name_value(source)
                if cleaned and cleaned != source:
                    updates.append({"id": part_id, "name": cleaned})
                    if len(samples) < 25:
                        samples.append({"before": source, "after": cleaned})

            if updates:
                await db.execute(update_sql, updates)
                rows_updated += len(updates)

            await db.commit()

            # If this sampled batch yielded no modifications, avoid spin loops.
            if not updates:
                break

        logger.info(
            "clean_part_names: batches=%d checked=%d updated=%d",
            batches_run,
            rows_checked,
            rows_updated,
        )
    except Exception as exc:
        await db.rollback()
        logger.error("clean_part_names failed: %s", exc)
        return {"task": "clean_part_names", "status": "error", "error": str(exc)}

    return {
        "task": "clean_part_names",
        "status": "ok",
        "batch_size": batch_size,
        "batches_run": batches_run,
        "rows_checked": rows_checked,
        "rows_updated": rows_updated,
        "no_update_batches": no_update_batches,
        "sample_changes": samples,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Delta processing helpers — checkpoint read/write via system_settings
# =========================================================================
# Each delta task stores its last-successful-run timestamp here so that on
# the next cycle it only processes rows modified since then, not the whole table.
# Fallback window when no checkpoint exists: 6 h (covers a missed cycle).

_DELTA_FALLBACK_HOURS = 6


async def _get_task_checkpoint(db: AsyncSession, task_name: str) -> "datetime":
    """Return the last successful run time for *task_name*, or fallback to now-6h.

    Always returns a naive UTC datetime to match parts_catalog.updated_at which is
    'timestamp without time zone' — asyncpg refuses to bind tz-aware datetimes there.
    """
    key = f"delta_checkpoint__{task_name}"
    row = (await db.execute(
        text("SELECT value FROM system_settings WHERE key = :k"),
        {"k": key},
    )).scalar_one_or_none()
    if row:
        try:
            # Strip tzinfo regardless of storage format so asyncpg binds correctly
            return datetime.fromisoformat(row).replace(tzinfo=None)
        except Exception:
            pass
    return datetime.utcnow() - timedelta(hours=_DELTA_FALLBACK_HOURS)


async def _save_task_checkpoint(db: AsyncSession, task_name: str) -> None:
    """Persist the current UTC timestamp as the checkpoint for *task_name*."""
    key = f"delta_checkpoint__{task_name}"
    val = datetime.utcnow().isoformat()  # naive UTC matches updated_at column type
    await db.execute(text("""
        INSERT INTO system_settings (id, key, value, value_type, is_public, updated_at)
        VALUES (gen_random_uuid(), :k, :v, 'string', false, NOW())
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
    """), {"k": key, "v": val})
    await db.commit()


# =========================================================================
# Task 2 – Normalize part types
# =========================================================================

async def normalize_part_types(db: AsyncSession) -> Dict[str, Any]:
    """
    Unify part_type values to one of: "Original", "OEM", "Aftermarket".

    Delta mode: only processes rows modified since the last successful run
    (stored in system_settings as 'delta_checkpoint__normalize_part_types').
    On a 4M-row table under concurrent harvesting this cuts the scan from
    the entire table down to the rows ingested in the last cycle — from
    60+ min timeouts to typically <5 s.
    """
    t0 = time.monotonic()
    catalog_updated = supplier_updated = 0

    since = await _get_task_checkpoint(db, "normalize_part_types")

    when_clauses = "\n            ".join(
        f"WHEN LOWER(TRIM(part_type)) = '{k}' THEN '{v}'"
        for k, v in PART_TYPE_MAP.items()
    )
    case_sql = f"CASE\n            {when_clauses}\n            ELSE part_type\n        END"

    try:
        # BOUNDED BATCHING (2026-07-13): the old single UPDATE matched every delta row
        # — when the checkpoint was stale that was millions of rows locked in ONE tx for
        # 27+ min, serializing all parts_catalog writes (lock storm that blocked the
        # categorizer, importers, enrichment). Now: chunks of BATCH, committed each loop
        # (locks released), FOR UPDATE SKIP LOCKED so it never waits on rows another
        # writer holds, per-batch statement_timeout. Updated rows leave the match set,
        # so no cursor is needed; locked rows are retried a few times then left for the
        # next cycle. Same lock-safe pattern as bmw_oem_dedup / categorize_parts_batch.
        BATCH = 5000
        for table, counter_attr in (("parts_catalog", "catalog_updated"), ("supplier_parts", "supplier_updated")):
            total_n = 0
            empty_streak = 0
            for _ in range(20000):  # hard cap; normally exits on empty match set
                await db.execute(text("SET LOCAL statement_timeout = '30s'"))
                result = await db.execute(
                    text(f"""
                        WITH batch AS (
                            SELECT id FROM {table}
                            WHERE updated_at > :since
                              AND part_type IS NOT NULL
                              AND LOWER(TRIM(part_type)) = ANY(:keys)
                              AND part_type NOT IN ('Original', 'OEM', 'Aftermarket')
                            ORDER BY id
                            LIMIT :batch
                            FOR UPDATE SKIP LOCKED
                        )
                        UPDATE {table} t
                        SET part_type = {case_sql}, updated_at = NOW()
                        FROM batch WHERE t.id = batch.id
                    """),
                    {"since": since, "keys": list(PART_TYPE_MAP.keys()), "batch": BATCH},
                )
                n = result.rowcount
                await db.commit()  # release locks every batch
                total_n += n
                if n == 0:
                    empty_streak += 1
                    if empty_streak >= 3:
                        break  # match set empty (or all locked — next cycle catches)
                    await asyncio.sleep(2)
                    continue
                empty_streak = 0
                await asyncio.sleep(0.2)  # be gentle to the harvester/importers
            if counter_attr == "catalog_updated":
                catalog_updated = total_n
            else:
                supplier_updated = total_n

        await _save_task_checkpoint(db, "normalize_part_types")
        logger.info(
            "normalize_part_types (delta since %s): catalog=%d supplier=%d elapsed=%.1fs",
            since.isoformat(), catalog_updated, supplier_updated, time.monotonic() - t0,
        )
    except Exception as exc:
        await db.rollback()
        logger.error("normalize_part_types failed: %s", exc)
        return {"task": "normalize_part_types", "status": "error", "error": str(exc)}

    return {
        "task": "normalize_part_types",
        "status": "ok",
        "catalog_updated": catalog_updated,
        "supplier_updated": supplier_updated,
        "delta_since": since.isoformat(),
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 3 – Normalize categories
# =========================================================================

async def normalize_categories(db: AsyncSession) -> Dict[str, Any]:
    """
    Map non-canonical category values to canonical Hebrew categories.

    Delta mode: only processes rows where updated_at > last checkpoint.
    Pass 1 (bulk CASE UPDATE) maps known raw values → canonical Hebrew category.
    Pass 2 (batched fallback) sets anything still non-canonical → 'כללי'.
    Both passes are scoped to recently-modified rows only, so the table lock
    time is proportional to ingestion volume, not catalog size.
    """
    t0 = time.monotonic()
    since = await _get_task_checkpoint(db, "normalize_categories")

    # Deadlock resilience (fixed 2026-07-11): this task runs long batched UPDATEs
    # on parts_catalog while the harvester/importers concurrently UPDATE the same
    # rows → DeadlockDetectedError aborted the whole ~20-min task (status=error)
    # ~5×/day. Deadlocks are transient: retry the whole task a few times, and
    # (below) Pass 2 now takes row locks in a consistent id order with FOR UPDATE
    # SKIP LOCKED so it never waits on rows another writer holds.
    _MAX_DEADLOCK_RETRIES = 4
    for _attempt in range(_MAX_DEADLOCK_RETRIES):
        try:
            return await _normalize_categories_once(db, since, t0)
        except Exception as exc:
            try:
                await db.rollback()
            except Exception:
                pass
            if "deadlock" in str(exc).lower() and _attempt < _MAX_DEADLOCK_RETRIES - 1:
                logger.warning("normalize_categories deadlock (attempt %d/%d) — retrying",
                               _attempt + 1, _MAX_DEADLOCK_RETRIES)
                await asyncio.sleep(1.5 * (_attempt + 1))
                continue
            logger.error("normalize_categories failed: %s", exc)
            return {"task": "normalize_categories", "status": "error", "error": str(exc)}


async def _normalize_categories_once(db: AsyncSession, since, t0) -> Dict[str, Any]:
    """One attempt of normalize_categories (wrapped in deadlock retry above)."""
    if True:
        canonical_set = set(CANONICAL_CATEGORIES)
        canonical_sql = ", ".join(f"'{c}'" for c in CANONICAL_CATEGORIES)

        # Build CASE branches (same logic as before)
        branches: List[str] = []
        seen_raw: set = set()
        combined: Dict[str, str] = {}
        for raw, mid in CATEGORY_MAP.items():
            target = CATEGORY_NAME_REMAP.get(mid, mid)
            combined[raw.lower()] = target
        for raw, target in CATEGORY_NAME_REMAP.items():
            if raw not in canonical_set:
                combined.setdefault(raw.lower(), target)
        for raw_lower, target in combined.items():
            if target in canonical_set and raw_lower not in seen_raw:
                seen_raw.add(raw_lower)
                escaped_raw = raw_lower.replace("'", "''")
                escaped_tgt = target.replace("'", "''")
                branches.append(f"WHEN TRIM(LOWER(category)) = '{escaped_raw}' THEN '{escaped_tgt}'")

        rows_mapped = 0
        rows_fallback = 0

        # ── Pass 1: bulk CASE UPDATE — delta scope ─────────────────────────────
        if branches:
            case_sql = "CASE\n  " + "\n  ".join(branches) + "\n  ELSE NULL\nEND"
            result = await db.execute(text(f"""
                UPDATE parts_catalog
                SET    category   = sub.new_cat,
                       updated_at = NOW()
                FROM (
                    SELECT id, {case_sql} AS new_cat
                    FROM   parts_catalog
                    WHERE  updated_at > :since
                      AND  category IS NOT NULL
                      AND  TRIM(category) NOT IN ({canonical_sql})
                ) sub
                WHERE  parts_catalog.id = sub.id
                  AND  sub.new_cat IS NOT NULL
            """), {"since": since})
            rows_mapped = result.rowcount or 0
            await db.commit()
            logger.info("normalize_categories pass1 (delta since %s): mapped=%d elapsed=%.1fs",
                        since.isoformat(), rows_mapped, time.monotonic() - t0)

        # ── Pass 2: fallback → 'כללי' in small batches — delta scope ───────────
        # Capture the max id at pass-start so rows arriving mid-pass (which get
        # updated_at=NOW() and stay in scope) don't keep the loop alive indefinitely.
        cutoff_row = await db.execute(text("SELECT MAX(id) FROM parts_catalog WHERE updated_at > :since"), {"since": since})
        cutoff_id = cutoff_row.scalar()
        batch_size = 5000
        while cutoff_id:
            # ORDER BY id + FOR UPDATE SKIP LOCKED: take row locks in a consistent
            # order and skip rows another writer (harvester/importer) is holding,
            # so this batch never forms a deadlock cycle and never blocks on a
            # contended row. Skipped rows stay in delta scope (updated_at) and are
            # picked up on a later pass/cycle.
            result2 = await db.execute(text(f"""
                WITH batch AS (
                    SELECT id FROM parts_catalog
                    WHERE  updated_at > :since
                      AND  id <= :cutoff_id
                      AND  category IS NOT NULL
                      AND  TRIM(category) NOT IN ({canonical_sql})
                    ORDER BY id
                    LIMIT  {batch_size}
                    FOR UPDATE SKIP LOCKED
                )
                UPDATE parts_catalog
                SET    category   = 'כללי',
                       updated_at = NOW()
                FROM   batch
                WHERE  parts_catalog.id = batch.id
            """), {"since": since, "cutoff_id": cutoff_id})
            n = result2.rowcount or 0
            await db.commit()
            rows_fallback += n
            if n < batch_size:
                break

        await _save_task_checkpoint(db, "normalize_categories")
        logger.info("normalize_categories pass2 (delta): fallback=%d total_elapsed=%.1fs",
                    rows_fallback, time.monotonic() - t0)

    return {
        "task": "normalize_categories",
        "status": "ok",
        "rows_mapped": rows_mapped,
        "rows_fallback": rows_fallback,
        "rows_updated": rows_mapped + rows_fallback,
        "delta_since": since.isoformat(),
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 4 – Normalize availability
# =========================================================================

async def normalize_availability(db: AsyncSession) -> Dict[str, Any]:
    """
    Unify availability strings on supplier_parts to:
    "in_stock" / "out_of_stock" / "on_order"
    """
    t0 = time.monotonic()
    rows_updated = 0

    try:
        result = await db.execute(
            text("SELECT id, availability FROM supplier_parts WHERE availability IS NOT NULL LIMIT 5000")
        )
        rows = result.fetchall()

        for _idx, (row_id, raw_avail) in enumerate(rows):
            if _idx % 100 == 0:
                await asyncio.sleep(0)
            canonical = _normalize_availability(raw_avail)
            if canonical and canonical != raw_avail:
                await db.execute(
                    text(
                        "UPDATE supplier_parts SET availability = :val, "
                        "updated_at = NOW() WHERE id = :id"
                    ),
                    {"val": canonical, "id": row_id},
                )
                rows_updated += 1

        await db.commit()
        logger.info("normalize_availability: updated=%d", rows_updated)
    except Exception as exc:
        await db.rollback()
        logger.error("normalize_availability failed: %s", exc)
        return {"task": "normalize_availability", "status": "error", "error": str(exc)}

    return {
        "task": "normalize_availability",
        "status": "ok",
        "rows_updated": rows_updated,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 5 – Fix base prices
# =========================================================================

async def fix_base_prices(db: AsyncSession) -> Dict[str, Any]:
    """
    Ensure parts_catalog.base_price is not below the cheapest supplier retail price.

    Pricing policy:
      - Domestic (IL) suppliers: cost_excl_vat × 1.18 × 1.45
      - International suppliers: cost_excl_vat × 1.45 (no IL VAT)
      - MARGIN = 1.45 (45% hidden markup — never shown to customers)

    VAT is determined per supplier via suppliers.country = 'IL'.
    supplier_parts.price_ils is always stored excl. VAT.
    """
    MARGIN = 1.45
    t0 = time.monotonic()
    rows_updated = 0

    # Retry up to 3× on deadlock — concurrent freesbe/run_all writes can cause transient deadlocks
    for _attempt in range(3):
        try:
            ils_rate = await _get_ils_rate(db)

            # Single bulk UPDATE via CTE — minimises lock window vs row-by-row loop
            # Pricing policy: base_price = supplier_cost_excl_vat × 1.45 (45% margin, no VAT factor)
            result = await db.execute(
                text(
                    """
                    WITH new_prices AS (
                        SELECT
                            pc.id,
                            ROUND(MIN(
                                CASE
                                    WHEN sp.price_ils IS NOT NULL THEN
                                        sp.price_ils * :margin
                                    WHEN sp.price_usd IS NOT NULL THEN
                                        sp.price_usd * :rate * :margin
                                    ELSE NULL
                                END
                            )::numeric, 2) AS min_retail_ils
                        FROM parts_catalog pc
                        JOIN supplier_parts sp ON sp.part_id = pc.id AND sp.is_available = TRUE
                        JOIN suppliers s ON s.id = sp.supplier_id
                        GROUP BY pc.id
                        HAVING MIN(
                            CASE
                                WHEN sp.price_ils IS NOT NULL THEN
                                    sp.price_ils * :margin
                                WHEN sp.price_usd IS NOT NULL THEN
                                    sp.price_usd * :rate * :margin
                                ELSE NULL
                            END
                        ) IS NOT NULL
                    )
                    UPDATE parts_catalog pc
                    SET base_price = np.min_retail_ils, updated_at = NOW()
                    FROM new_prices np
                    WHERE pc.id = np.id
                      AND (pc.base_price IS NULL OR pc.base_price < np.min_retail_ils)
                    """
                ),
                {"rate": ils_rate, "margin": MARGIN},
            )
            rows_updated = result.rowcount or 0
            await db.commit()
            logger.info("fix_base_prices: updated=%d (rate=%.2f)", rows_updated, ils_rate)
            break  # success

        except Exception as exc:
            await db.rollback()
            exc_str = str(exc)
            if "deadlock" in exc_str.lower() and _attempt < 2:
                wait = 2 ** _attempt
                logger.warning("fix_base_prices deadlock (attempt %d/3), retrying in %ds", _attempt + 1, wait)
                await asyncio.sleep(wait)
                continue
            logger.error("fix_base_prices failed: %s", exc)
            return {"task": "fix_base_prices", "status": "error", "error": exc_str}

    return {
        "task": "fix_base_prices",
        "status": "ok",
        "rows_updated": rows_updated,
        "ils_per_usd_used": ils_rate,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 5b – Normalize base_price from alternate price columns
# =========================================================================

async def normalize_base_price(db: AsyncSession) -> Dict[str, Any]:
    """
    Compute base_price (customer-facing retail) from source price columns.

    Pricing policy: base_price = cost_excl_vat × 1.45. No exceptions.

    Decision tree (applied in priority order):

      1. importer_price_ils > 0  (actual wholesale cost excl. VAT, all IL brands):
         base = importer_price_ils * 1.45

      2. importer=0, online_price_ils > 0  (eBay/international, no IL VAT):
         base = online_price_ils * 1.45

      3. importer=0, online=0, max_price_ils > 0  (IL consumer ref price incl. 18% VAT):
         base = (max_price_ils / 1.18) * 1.45  — divide out VAT first, then apply margin
    """
    MARGIN = float(os.getenv("IL_MARGIN", "1.45"))
    t0 = time.monotonic()
    # Delta scope (added 2026-07-11): base_price only needs recomputing when a
    # SOURCE price column changed — and every importer bumps updated_at when it
    # writes prices (mandatory pipeline rule) — so we only scan rows touched since
    # the last checkpoint instead of the whole 1.9M-row table every cycle. The old
    # full-table scan ran ~30 min holding a snapshot (blocking other work and
    # deadlocking the harvester); delta scans a few thousand rows in seconds. A
    # weekly full pass still catches anything missed (e.g. a MARGIN change). Same
    # pattern as refresh_min_max_prices.
    since = await _get_task_checkpoint(db, "normalize_base_price")
    full_pass = (datetime.utcnow() - since) > timedelta(days=7)
    try:
        # Fail fast instead of deadlocking/blocking on rows the harvester holds;
        # the run_all_tasks central deadlock/lock retry re-runs it next.
        await db.execute(text("SET LOCAL lock_timeout = '2min'"))
        # No-op guard: each UPDATE writes only rows whose base_price would ACTUALLY
        # change (`IS DISTINCT FROM`) — steady state writes ~0.
        # Case 1: importer_price_ils is the actual cost excl. VAT — 45% margin over cost
        r1 = await db.execute(text("""
            UPDATE parts_catalog
            SET base_price = ROUND((importer_price_ils * :margin)::numeric, 2),
                updated_at = NOW()
            WHERE importer_price_ils > 0
              AND is_active = TRUE
              AND (:full_pass OR updated_at > :since)
              AND base_price IS DISTINCT FROM ROUND((importer_price_ils * :margin)::numeric, 2)
        """), {"margin": MARGIN, "full_pass": full_pass, "since": since})
        importer_updated = r1.rowcount

        # Case 2: eBay / international buy price (already excl. VAT) — 45% margin
        r2 = await db.execute(text("""
            UPDATE parts_catalog
            SET base_price = ROUND((online_price_ils * :margin)::numeric, 2),
                updated_at = NOW()
            WHERE (importer_price_ils IS NULL OR importer_price_ils = 0)
              AND online_price_ils > 0
              AND is_active = TRUE
              AND (:full_pass OR updated_at > :since)
              AND base_price IS DISTINCT FROM ROUND((online_price_ils * :margin)::numeric, 2)
        """), {"margin": MARGIN, "full_pass": full_pass, "since": since})
        online_updated = r2.rowcount

        # Case 3: IL importer reference price (incl. 18% VAT) — divide out VAT, then 45% margin
        r3 = await db.execute(text("""
            UPDATE parts_catalog
            SET base_price = ROUND((max_price_ils / 1.18 * :margin)::numeric, 2),
                updated_at = NOW()
            WHERE (importer_price_ils IS NULL OR importer_price_ils = 0)
              AND (online_price_ils IS NULL OR online_price_ils = 0)
              AND max_price_ils > 0
              AND is_active = TRUE
              AND (:full_pass OR updated_at > :since)
              AND base_price IS DISTINCT FROM ROUND((max_price_ils / 1.18 * :margin)::numeric, 2)
        """), {"margin": MARGIN, "full_pass": full_pass, "since": since})
        max_updated = r3.rowcount

        await db.commit()
        await _save_task_checkpoint(db, "normalize_base_price")
        total = online_updated + importer_updated + max_updated
        logger.info(
            "normalize_base_price: online→%d importer→%d il_ref→%d total=%d in %.2fs",
            online_updated, importer_updated, max_updated, total,
            round(time.monotonic() - t0, 2),
        )
    except Exception as exc:
        await db.rollback()
        logger.error("normalize_base_price failed: %s", exc)
        return {"task": "normalize_base_price", "status": "error", "error": str(exc)}

    return {
        "task": "normalize_base_price",
        "status": "ok",
        "online_price_updated": online_updated,
        "importer_price_updated": importer_updated,
        "il_ref_price_updated": max_updated,
        "total_updated": total,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 6 – Flag fake SKUs
# =========================================================================

async def fix_manufacturer_overflow(db: AsyncSession) -> Dict[str, Any]:
    """
    Identify Renault/Chevrolet parts with suspicious OEM prefixes
    and mark them for legitimate OEM lookup.
    """
    t0 = time.monotonic()
    rows_flagged = 0
    try:
        # Renault: should start with digits. 'RE' prefix or non-digit start is suspicious.
        # Chevrolet: often has 'CH' prefix in bad imports.
        result = await db.execute(text("""
            UPDATE parts_catalog
            SET needs_oem_lookup = TRUE,
                updated_at = NOW()
            WHERE (
                (
                  (manufacturer ILIKE '%renault%' OR manufacturer ILIKE '%רנו%')
                  AND (oem_number ~* '^RE' OR oem_number !~ '^[0-9]')
                )
                OR
                (
                  (manufacturer ILIKE '%chevrolet%' OR manufacturer ILIKE '%שברולט%')
                  AND oem_number ~* '^CH'
                )
              )
              AND (needs_oem_lookup IS False OR needs_oem_lookup IS NULL)
              AND is_active = TRUE
            RETURNING id
        """))
        rows_flagged = len(result.fetchall())
        await db.commit()
        logger.info("fix_manufacturer_overflow: flagged=%d", rows_flagged)
    except Exception as exc:
        await db.rollback()
        logger.error("fix_manufacturer_overflow failed: %s", exc)
        return {"task": "fix_manufacturer_overflow", "status": "error", "error": str(exc)}

    return {
        "task": "fix_manufacturer_overflow",
        "status": "ok",
        "rows_flagged": rows_flagged,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


async def flag_fake_skus(db: AsyncSession) -> Dict[str, Any]:
    """
    Set needs_oem_lookup = TRUE on parts whose SKU matches auto-generated
    patterns (e.g., "OIL-TOY-001", "PART-12345").
    """
    t0 = time.monotonic()
    rows_flagged = 0

    try:
        result = await db.execute(
            text(
                "SELECT id, sku FROM parts_catalog "
                "WHERE sku IS NOT NULL AND (needs_oem_lookup IS NULL OR needs_oem_lookup = FALSE)"
            )
        )
        rows = result.fetchall()

        for part_id, sku in rows:
            if _is_fake_sku(sku):
                await db.execute(
                    text(
                        "UPDATE parts_catalog SET needs_oem_lookup = TRUE, "
                        "updated_at = NOW() WHERE id = :id"
                    ),
                    {"id": part_id},
                )
                rows_flagged += 1

        await db.commit()
        logger.info("flag_fake_skus: flagged=%d", rows_flagged)
    except Exception as exc:
        await db.rollback()
        logger.error("flag_fake_skus failed: %s", exc)
        return {"task": "flag_fake_skus", "status": "error", "error": str(exc)}

    return {
        "task": "flag_fake_skus",
        "status": "ok",
        "rows_flagged": rows_flagged,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 7 – Fill car brand metadata
# =========================================================================

async def fill_car_brands(db: AsyncSession) -> Dict[str, Any]:
    """
    Seed il_importer, warranty_years, warranty_km, warranty_notes for
    known Israeli brands. Only updates NULL fields to avoid overriding
    manual edits.

    Root behavior: active brands missing explicit mapping receive
    deterministic fallback metadata.
    """
    t0 = time.monotonic()
    rows_updated = 0
    mapped_updates = 0
    fallback_updates = 0

    try:
        result = await db.execute(
            text("SELECT id, name, is_active FROM car_brands WHERE name IS NOT NULL")
        )
        brands = result.fetchall()

        for brand_id, brand_name, is_active in brands:
            key = brand_name.strip().lower()
            data = BRAND_IMPORTER_MAP.get(key)
            used_fallback = False
            if not data and bool(is_active):
                data = DEFAULT_BRAND_METADATA
                used_fallback = True
            if not data:
                continue

            importer, years, km, notes = data

            updates: List[str] = []
            params: Dict[str, Any] = {"id": brand_id}

            r = await db.execute(
                text(
                    "SELECT il_importer, warranty_years, warranty_km, warranty_notes "
                    "FROM car_brands WHERE id = :id"
                ),
                {"id": brand_id},
            )
            row = r.fetchone()
            if not row:
                continue

            cur_importer, cur_years, cur_km, cur_notes = row

            if cur_importer is None:
                updates.append("il_importer = :importer")
                params["importer"] = importer
            if cur_years is None:
                updates.append("warranty_years = :years")
                params["years"] = years
            if cur_km is None and km is not None:
                updates.append("warranty_km = :km")
                params["km"] = km
            if cur_notes is None and notes:
                updates.append("warranty_notes = :notes")
                params["notes"] = notes

            if updates:
                await db.execute(
                    text(
                        f"UPDATE car_brands SET {', '.join(updates)}, "
                        "updated_at = NOW() WHERE id = :id"
                    ),
                    params,
                )
                rows_updated += 1
                if used_fallback:
                    fallback_updates += 1
                else:
                    mapped_updates += 1

        await db.commit()
        logger.info(
            "fill_car_brands: updated=%d mapped=%d fallback=%d",
            rows_updated,
            mapped_updates,
            fallback_updates,
        )
    except Exception as exc:
        await db.rollback()
        logger.error("fill_car_brands failed: %s", exc)
        return {"task": "fill_car_brands", "status": "error", "error": str(exc)}

    return {
        "task": "fill_car_brands",
        "status": "ok",
        "rows_updated": rows_updated,
        "mapped_updates": mapped_updates,
        "fallback_updates": fallback_updates,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


async def normalize_imported_manufacturers(db: AsyncSession) -> Dict[str, Any]:
    """
    Normalize free-text manufacturer values introduced by imports so search and
    registry mapping remain stable across reruns.
    """
    t0 = time.monotonic()
    catalog_updated = 0
    vehicles_updated = 0
    try:
        from manufacturer_normalization import normalize_manufacturer_name

        cat_rows = (await db.execute(text("""
            SELECT DISTINCT manufacturer
            FROM parts_catalog
            WHERE manufacturer IS NOT NULL
              AND manufacturer <> ''
        """))).fetchall()

        for (raw,) in cat_rows:
            if not raw:
                continue
            canon = normalize_manufacturer_name(raw, raw)
            if canon and canon != raw:
                res = await db.execute(
                    text("""
                        UPDATE parts_catalog
                        SET manufacturer = :canon,
                            updated_at = NOW()
                        WHERE manufacturer = :raw
                    """),
                    {"raw": raw, "canon": canon},
                )
                catalog_updated += res.rowcount or 0

        veh_rows = (await db.execute(text("""
            SELECT DISTINCT manufacturer
            FROM vehicles
            WHERE manufacturer IS NOT NULL
              AND manufacturer <> ''
        """))).fetchall()

        for (raw,) in veh_rows:
            if not raw:
                continue
            canon = normalize_manufacturer_name(raw, raw)
            if canon and canon != raw:
                res = await db.execute(
                    text("""
                        UPDATE vehicles
                        SET manufacturer = :canon
                        WHERE manufacturer = :raw
                    """),
                    {"raw": raw, "canon": canon},
                )
                vehicles_updated += res.rowcount or 0

        # Fix manufacturers based on OEM number prefix
        from manufacturer_normalization import normalize_oem_manufacturer, OEM_PREFIX_TO_MANUFACTURER
        oem_prefix_updated = 0
        for prefix, correct_mfr in OEM_PREFIX_TO_MANUFACTURER.items():
            res = await db.execute(
                text("""
                    UPDATE parts_catalog
                    SET manufacturer = :correct,
                        updated_at = NOW()
                    WHERE oem_number LIKE :prefix
                    AND manufacturer != :correct
                    AND is_active = true
                """),
                {"correct": correct_mfr, "prefix": f"{prefix}%"},
            )
            if res.rowcount:
                logger.info(
                    "normalize_oem_prefix: %s prefix=%s updated=%d",
                    correct_mfr, prefix, res.rowcount
                )
                oem_prefix_updated += res.rowcount

        await db.commit()
        return {
            "task": "normalize_imported_manufacturers",
            "status": "ok",
            "catalog_updated": catalog_updated,
            "vehicles_updated": vehicles_updated,
            "oem_prefix_updated": oem_prefix_updated,
            "elapsed_s": round(time.monotonic() - t0, 2),
        }
    except Exception as exc:
        await db.rollback()
        logger.error("normalize_imported_manufacturers failed: %s", exc)
        return {
            "task": "normalize_imported_manufacturers",
            "status": "error",
            "error": str(exc),
        }


async def sync_manufacturer_registries(db: AsyncSession) -> Dict[str, Any]:
    """
    Keep brand registries deployment-safe and idempotent:
    - canonicalize noisy manufacturers into car_brands
    - ensure baseline truck brands in truck_brands
    - keep logo_url/aliases populated
    """
    t0 = time.monotonic()
    try:
        from clean_manufacturers_registry import sync_manufacturer_registries as _sync
        report = await _sync(db)
        return {
            "task": "sync_manufacturer_registries",
            "status": "ok",
            **report,
            "elapsed_s": round(time.monotonic() - t0, 2),
        }
    except Exception as exc:
        try:
            await db.rollback()
        except Exception:
            pass
        logger.error("sync_manufacturer_registries failed: %s", exc)
        return {
            "task": "sync_manufacturer_registries",
            "status": "error",
            "error": str(exc),
        }


async def sync_models_from_catalog(db: AsyncSession) -> Dict[str, Any]:
    """Backfill vehicles(manufacturer, model, year) from parts_catalog.compatible_vehicles.

    This gives the manufacturer→model hierarchy a richer source even when
    imported vehicle rows are sparse.
    """
    t0 = time.monotonic()
    inserted = 0
    scanned = 0
    skipped = 0
    # Cache manufacturer name → car_brands.id for this run. vehicles.manufacturer_id
    # is a NOT NULL FK to car_brands(id); the insert previously omitted it, which
    # failed the whole task with NotNullViolationError (e.g. "Vw") every cycle.
    # Fixed 2026-07-11: resolve the id (case-insensitive, by name OR alias) and
    # skip rows whose brand is not registered yet rather than aborting.
    _mfr_id_cache: Dict[str, Any] = {}

    rows = (await db.execute(text("""
        SELECT manufacturer, compatible_vehicles
        FROM parts_catalog
        WHERE is_active = TRUE
          AND manufacturer IS NOT NULL
          AND TRIM(manufacturer) <> ''
          AND compatible_vehicles IS NOT NULL
          AND jsonb_typeof(compatible_vehicles) = 'array'
        LIMIT 2000
    """))).fetchall()

    if not rows:
        return {"task": "sync_models_from_catalog", "status": "ok", "scanned": 0, "inserted": 0}

    # Track seen combinations in this run to minimize duplicate DB checks.
    seen_keys = set()
    _yield_every = 50  # yield to event loop every N rows to keep heartbeat alive

    for _row_idx, (manufacturer, compat) in enumerate(rows):
        if _row_idx % _yield_every == 0:
            await asyncio.sleep(0)
        mfr = (manufacturer or "").strip()
        if not mfr:
            continue
        if not isinstance(compat, list):
            continue

        for item in compat:
            if not isinstance(item, dict):
                continue

            raw_model = item.get("model") or item.get("model_year")
            model = _clean_vehicle_model(raw_model)
            if not model:
                continue

            # Prefer structured years when available; else use 0 placeholder.
            y_from = item.get("year_from")
            y_to = item.get("year_to")
            year = 0
            try:
                if isinstance(y_from, int):
                    year = y_from
                elif isinstance(y_from, str) and y_from.isdigit():
                    year = int(y_from)
                elif isinstance(y_to, int):
                    year = y_to
                elif isinstance(y_to, str) and y_to.isdigit():
                    year = int(y_to)
            except Exception:
                year = 0

            key = (mfr.casefold(), model.casefold(), int(year))
            if key in seen_keys:
                continue
            seen_keys.add(key)
            scanned += 1

            exists = (await db.execute(text("""
                SELECT 1
                FROM vehicles
                WHERE LOWER(TRIM(manufacturer)) = LOWER(TRIM(:mfr))
                  AND LOWER(TRIM(model)) = LOWER(TRIM(:model))
                  AND year = :year
                LIMIT 1
            """), {"mfr": mfr, "model": model, "year": year})).fetchone()
            if exists:
                continue

            # Resolve NOT NULL manufacturer_id (FK to car_brands) before insert.
            mkey = mfr.casefold()
            if mkey not in _mfr_id_cache:
                mid_row = (await db.execute(text("""
                    SELECT id FROM car_brands
                    WHERE LOWER(BTRIM(name)) = LOWER(BTRIM(:mfr))
                       OR EXISTS (
                           SELECT 1 FROM unnest(COALESCE(aliases, '{}')) a
                           WHERE LOWER(BTRIM(a)) = LOWER(BTRIM(:mfr))
                       )
                    ORDER BY is_active DESC
                    LIMIT 1
                """), {"mfr": mfr})).fetchone()
                _mfr_id_cache[mkey] = mid_row[0] if mid_row else None
            manufacturer_id = _mfr_id_cache[mkey]
            if manufacturer_id is None:
                # Brand not registered yet — sync_manufacturer_registries adds it;
                # a later cycle will pick this vehicle up. Skip, don't abort.
                skipped += 1
                continue

            await db.execute(text("""
                INSERT INTO vehicles
                    (id, license_plate, manufacturer, manufacturer_id, model, year, vin, created_at)
                VALUES
                    (gen_random_uuid(), NULL, :mfr, :mid, :model, :year, NULL, NOW())
            """), {"mfr": mfr, "mid": manufacturer_id, "model": model, "year": year})
            inserted += 1

    await db.commit()
    return {
        "task": "sync_models_from_catalog",
        "status": "ok",
        "scanned": scanned,
        "inserted": inserted,
        "skipped": skipped,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


async def sync_models_from_catalog_file(db: AsyncSession) -> Dict[str, Any]:
    """Extract manufacturer->model pairs from backend/data/parts_database.xlsx.

    This uses the original catalog workbook (not only already-imported rows),
    so brand model dropdowns get a richer hierarchy (notably Citroen/Peugeot).
    """
    t0 = time.monotonic()
    scanned = 0
    inserted = 0
    hierarchy_rows: set[Tuple[str, str, str, int, int, str]] = set()

    def _add_hierarchy_row(mfr: Optional[str], model: Optional[str], sub_model: Optional[str], year_from: int = 0, year_to: int = 0, source_sheet: str = "derived") -> None:
        canonical_mfr = normalize_manufacturer_name(mfr, mfr)
        if not canonical_mfr or canonical_mfr.strip().lower() in PARTS_BRANDS:
            return
        canonical_model = canonicalize_vehicle_model_for_manufacturer(canonical_mfr, model)
        canonical_sub = normalize_vehicle_submodel_name(sub_model)
        if not canonical_mfr or not canonical_model:
            return
        hierarchy_rows.add((canonical_mfr, canonical_model, canonical_sub or "", int(year_from or 0), int(year_to or 0), source_sheet))

    try:
        import openpyxl
        from pathlib import Path
        
    except Exception as exc:
        return {
            "task": "sync_models_from_catalog_file",
            "status": "error",
            "error": f"dependency_error: {exc}",
        }

    xlsx_path = Path(__file__).resolve().parent / "data" / "parts_database.xlsx"
    if not xlsx_path.exists():
        return {
            "task": "sync_models_from_catalog_file",
            "status": "skipped",
            "reason": "catalog_file_missing",
            "path": str(xlsx_path),
        }

    # Keep consistent with importer sheet mapping.
    sheet_map = {
        "Chevrolet": ("Chevrolet", "A"),
        "Citroen": ("Citroen", "F"),
        "Peugeot": ("Peugeot", "F"),
    }

    def _norm_spaces(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def _extract_years(values: List[Any]) -> List[int]:
        years: set[int] = set()
        for v in values:
            if v is None:
                continue
            s = str(v)
            for m in re.findall(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", s):
                try:
                    yy = int(m)
                except Exception:
                    continue
                if 1990 <= yy <= 2027:
                    years.add(yy)
        return sorted(years)

    def _split_model_submodel(raw: Optional[str], mfr_variants: List[str], row_values: Optional[List[Any]] = None) -> Tuple[str, str, int, int, int]:
        years = _extract_years([raw or ""])
        year_hint = years[0] if years else 0
        year_from = years[0] if years else 0
        year_to = years[-1] if years else 0
        target_mfr = normalize_manufacturer_name(
            mfr_variants[0] if mfr_variants else "",
            mfr_variants[0] if mfr_variants else "",
        )

        m = _clean_vehicle_model(raw)
        if not m:
            return "", "", year_hint, year_from, year_to
        m2 = _norm_spaces(m)

        # Remove leading manufacturer tokens from model labels
        # e.g. "CITROEN C4" -> "C4".
        low = m2.lower()
        for v in sorted({x.lower() for x in mfr_variants if x}, key=len, reverse=True):
            if low.startswith(v + " "):
                m2 = m2[len(v):].strip()
                break

        # Remove trailing noisy qualifiers often present in workbook model fields.
        m2 = re.sub(r"\b(new|basic|accessories|accessory)\b", "", m2, flags=re.IGNORECASE).strip()
        m2 = re.sub(r"\b(19|20)\d{2}\b", "", m2).strip()
        m2 = re.sub(r"\s{2,}", " ", m2).strip()

        if not _clean_vehicle_model(m2):
            return "", "", year_hint, year_from, year_to

        # Split base model and submodel/trim for hierarchy.
        if re.search(r"\s-\s*", m2):
            base, sub = [x.strip() for x in re.split(r"\s-\s*", m2, maxsplit=1)]
        else:
            # Platform/trim variants frequently appear as trailing tokens in workbook rows.
            # Examples: "BERLINGO B9", "BERLINGO K9", "BERLINGO K9 ACC".
            m_code = re.match(
                r"^(?P<base>[A-Za-z0-9\u0590-\u05FF\s]+?)\s+(?P<sub>[A-Z]\d{1,3}(?:\s+[A-Z]{2,8})?)$",
                m2,
                flags=re.IGNORECASE,
            )
            if m_code:
                base, sub = m_code.group("base").strip(), m_code.group("sub").upper()
            else:
                base, sub = m2, ""

        base = canonicalize_vehicle_model_for_manufacturer(target_mfr, base)
        sub = normalize_vehicle_submodel_name(sub)
        if not base:
            return "", "", year_hint, year_from, year_to
        return base, sub, year_hint, year_from, year_to

    def _load_xlsx_rows(path, s_map):
        """Load all sheet rows from the workbook synchronously (runs in a thread)."""
        import openpyxl as _xl
        wb = _xl.load_workbook(path, read_only=True, data_only=True)
        result = {}
        for s_name, (_, s_type) in s_map.items():
            if s_name not in wb.sheetnames:
                continue
            ws = wb[s_name]
            start = 8 if s_type == "F" else 3
            result[s_name] = list(ws.iter_rows(min_row=start, values_only=True))
        wb.close()
        return result

    sheet_rows = await asyncio.to_thread(_load_xlsx_rows, xlsx_path, sheet_map)

    seen_run = set()
    for sheet_name, (raw_mfr, stype) in sheet_map.items():
        if sheet_name not in sheet_rows:
            continue

        canonical_mfr = normalize_manufacturer_name(raw_mfr, raw_mfr)
        mfr_variants = [canonical_mfr, raw_mfr, sheet_name]

        if stype == "F":
            model_idx = 0
        elif stype == "A":
            model_idx = 7
        else:
            continue

        for row in sheet_rows[sheet_name]:
            if model_idx >= len(row):
                continue
            model_raw = row[model_idx]
            model, sub_model, year_hint, year_from, year_to = _split_model_submodel(
                str(model_raw) if model_raw is not None else "",
                mfr_variants,
                None,
            )
            if not model:
                continue

            target_mfr = canonical_mfr
            # Known PSA cross-brand correction: Partner belongs under Peugeot.
            if canonical_mfr.casefold() == "citroen" and model.casefold().startswith("partner"):
                target_mfr = "Peugeot"

            key = (target_mfr.casefold(), model.casefold(), sub_model.casefold(), int(year_hint or 0))
            if key in seen_run:
                continue
            seen_run.add(key)
            scanned += 1
            _add_hierarchy_row(target_mfr, model, sub_model, int(year_from or 0), int(year_to or 0), sheet_name)

            exists = (await db.execute(text("""
                SELECT 1
                FROM vehicles
                WHERE LOWER(TRIM(manufacturer)) = LOWER(TRIM(:mfr))
                  AND LOWER(TRIM(model)) = LOWER(TRIM(:model))
                  AND LOWER(COALESCE(gov_api_data->>'sub_model', '')) = LOWER(:sub_model)
                                    AND year = :year
                LIMIT 1
                        """), {"mfr": target_mfr, "model": model, "sub_model": sub_model or "", "year": int(year_hint or 0)})).fetchone()
            if exists:
                continue

            await db.execute(text("""
                INSERT INTO vehicles
                    (id, license_plate, manufacturer, model, year, vin, gov_api_data, created_at)
                VALUES
                    (gen_random_uuid(), NULL, :mfr, :model, :year, NULL, CAST(:gov AS jsonb), NOW())
            """), {
                "mfr": target_mfr,
                "model": model,
                "year": int(year_hint or 0),
                "gov": json.dumps({"sub_model": sub_model} if sub_model else {}, ensure_ascii=False),
            })
            inserted += 1

    compat_rows = (await db.execute(text("""
        SELECT DISTINCT
            COALESCE(elem->>'make', elem->>'manufacturer') AS manufacturer,
            COALESCE(elem->>'model', elem->>'model_year') AS model,
            COALESCE(elem->>'sub_model', '') AS sub_model,
            COALESCE(elem->>'year_from', '') AS year_from,
            COALESCE(elem->>'year_to', '') AS year_to,
            COALESCE(elem->>'year', '') AS year_hint
        FROM parts_catalog,
             jsonb_array_elements(coalesce(compatible_vehicles, '[]'::jsonb)) AS elem
        WHERE compatible_vehicles IS NOT NULL
          AND jsonb_typeof(compatible_vehicles) = 'array'
          AND COALESCE(elem->>'make', elem->>'manufacturer') IS NOT NULL
          AND COALESCE(elem->>'make', elem->>'manufacturer') <> ''
          AND COALESCE(elem->>'model', elem->>'model_year') IS NOT NULL
          AND COALESCE(elem->>'model', elem->>'model_year') <> ''
    """))).fetchall()
    for mfr, model, sub_model, year_from, year_to, year_hint in compat_rows:
        try:
            yf = int(year_from) if str(year_from).isdigit() else 0
        except Exception:
            yf = 0
        try:
            yt = int(year_to) if str(year_to).isdigit() else 0
        except Exception:
            yt = 0
        if not yf or not yt:
            try:
                yv = int(year_hint) if str(year_hint).isdigit() else 0
            except Exception:
                yv = 0
            if yv:
                yf = yt = yv
        _add_hierarchy_row(mfr, model, sub_model, yf, yt, "compatible_vehicles")

    vehicle_rows = (await db.execute(text("""
        SELECT manufacturer, model, COALESCE(gov_api_data->>'sub_model', '') AS sub_model, year
        FROM vehicles
        WHERE manufacturer IS NOT NULL
          AND manufacturer <> ''
          AND model IS NOT NULL
          AND model <> ''
    """))).fetchall()
    for mfr, model, sub_model, year in vehicle_rows:
        yv = int(year or 0) if isinstance(year, int) else 0
        _add_hierarchy_row(mfr, model, sub_model, yv, yv, "vehicles")

    # Backfill model/sub-model year bounds from vehicles when workbook row lacks year data.
    v_rows = (await db.execute(text("""
        SELECT
            manufacturer,
            model,
            COALESCE(gov_api_data->>'sub_model', '') AS sub_model,
            MIN(year) AS y_from,
            MAX(year) AS y_to
        FROM vehicles
        WHERE year BETWEEN 1990 AND 2027
          AND manufacturer IS NOT NULL
          AND model IS NOT NULL
        GROUP BY manufacturer, model, COALESCE(gov_api_data->>'sub_model', '')
    """))).fetchall()
    exact_years: Dict[Tuple[str, str, str], Tuple[int, int]] = {}
    model_years: Dict[Tuple[str, str], Tuple[int, int]] = {}
    for mfr, mdl, sub, y_from, y_to in v_rows:
        if not mfr or not mdl:
            continue
        mf = str(mfr).strip().casefold()
        md = normalize_vehicle_model_name(str(mdl)).casefold()
        sb = normalize_vehicle_submodel_name(str(sub or "")).casefold()
        if not md:
            continue
        yf = int(y_from or 0)
        yt = int(y_to or 0)
        if yf and yt:
            exact_years[(mf, md, sb)] = (yf, yt)
            prev = model_years.get((mf, md))
            if not prev:
                model_years[(mf, md)] = (yf, yt)
            else:
                model_years[(mf, md)] = (min(prev[0], yf), max(prev[1], yt))

    enriched_rows: set[Tuple[str, str, str, int, int, str]] = set()
    for mfr, model, sub_model, y_from, y_to, source_sheet in hierarchy_rows:
        yf = int(y_from or 0)
        yt = int(y_to or 0)
        rule_span = GENERATION_YEAR_RULES.get((mfr.casefold(), model.casefold(), (sub_model or "").casefold()))
        if rule_span:
            yf, yt = int(rule_span[0]), int(rule_span[1])
        if not yf or not yt:
            ek = (mfr.casefold(), model.casefold(), (sub_model or "").casefold())
            mk = (mfr.casefold(), model.casefold())
            by_sub = exact_years.get(ek)
            by_model = model_years.get(mk)
            # For specific sub-model entries, avoid over-broad model-level spans.
            span = by_sub or (by_model if not (sub_model or "").strip() else None)
            if span:
                yf, yt = int(span[0]), int(span[1])
        enriched_rows.add((mfr, model, sub_model, yf, yt, source_sheet))
    hierarchy_rows = enriched_rows

    # Materialize a dedicated XLS hierarchy table for frontend filters.
    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS vehicle_hierarchy_xls (
            id BIGSERIAL PRIMARY KEY,
            manufacturer TEXT NOT NULL,
            model TEXT NOT NULL,
            sub_model TEXT NOT NULL DEFAULT '',
            year_from INTEGER NOT NULL DEFAULT 0,
            year_to INTEGER NOT NULL DEFAULT 0,
            year_hint INTEGER NOT NULL DEFAULT 0,
            source_sheet TEXT,
            source_tag TEXT NOT NULL DEFAULT 'parts_database.xlsx',
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """))
    await db.execute(text("ALTER TABLE vehicle_hierarchy_xls ADD COLUMN IF NOT EXISTS year_from INTEGER NOT NULL DEFAULT 0"))
    await db.execute(text("ALTER TABLE vehicle_hierarchy_xls ADD COLUMN IF NOT EXISTS year_to INTEGER NOT NULL DEFAULT 0"))
    await db.execute(text("TRUNCATE TABLE vehicle_hierarchy_xls"))

    for mfr, model, sub_model, year_from, year_to, source_sheet in sorted(hierarchy_rows):
        await db.execute(text("""
            INSERT INTO vehicle_hierarchy_xls
                (manufacturer, model, sub_model, year_from, year_to, year_hint, source_sheet, source_tag, updated_at)
            VALUES
                (:mfr, :model, :sub_model, :year_from, :year_to, :year_hint, :source_sheet, 'parts_database.xlsx', NOW())
        """), {
            "mfr": mfr,
            "model": model,
            "sub_model": sub_model,
            "year_from": int(year_from or 0),
            "year_to": int(year_to or 0),
            "year_hint": int(year_from or 0),
            "source_sheet": source_sheet,
        })

    await db.commit()
    return {
        "task": "sync_models_from_catalog_file",
        "status": "ok",
        "scanned": scanned,
        "inserted": inserted,
        "hierarchy_rows": len(hierarchy_rows),
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


async def backfill_catalog_fitment_from_xls(db: AsyncSession) -> Dict[str, Any]:
    """Backfill exact workbook fitment into parts_catalog.compatible_vehicles."""
    t0 = time.monotonic()
    try:
        import openpyxl
        from pathlib import Path
    except Exception as exc:
        return {
            "task": "backfill_catalog_fitment_from_xls",
            "status": "error",
            "error": f"dependency_error: {exc}",
        }

    xlsx_path = Path(__file__).resolve().parent / "data" / "parts_database.xlsx"
    if not xlsx_path.exists():
        return {
            "task": "backfill_catalog_fitment_from_xls",
            "status": "skipped",
            "reason": "catalog_file_missing",
            "path": str(xlsx_path),
        }

    sheet_map = {
        "Chevrolet": ("Chevrolet", "CHEV-", "A"),
        "Citroen": ("Citroen", "CITR-", "F"),
        "Peugeot": ("Peugeot", "PEUG-", "F"),
    }

    def _norm_spaces(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def _extract_years(values: List[Any]) -> List[int]:
        years: set[int] = set()
        for v in values:
            if v is None:
                continue
            for m in re.findall(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", str(v)):
                try:
                    yy = int(m)
                except Exception:
                    continue
                if 1990 <= yy <= 2027:
                    years.add(yy)
        return sorted(years)

    def _split_model_submodel(raw: Optional[str], mfr_variants: List[str]) -> Tuple[str, str, int, int]:
        years = _extract_years([raw or ""])
        year_from = years[0] if years else 0
        year_to = years[-1] if years else 0
        target_mfr = normalize_manufacturer_name(
            mfr_variants[0] if mfr_variants else "",
            mfr_variants[0] if mfr_variants else "",
        )

        model_text = _clean_vehicle_model(raw)
        if not model_text:
            return "", "", year_from, year_to
        model_text = _norm_spaces(model_text)

        low = model_text.lower()
        for v in sorted({x.lower() for x in mfr_variants if x}, key=len, reverse=True):
            if low.startswith(v + " "):
                model_text = model_text[len(v):].strip()
                break

        model_text = re.sub(r"\b(new|basic|accessories|accessory)\b", "", model_text, flags=re.IGNORECASE).strip()
        model_text = re.sub(r"\b(19|20)\d{2}\b", "", model_text).strip()
        model_text = re.sub(r"\s{2,}", " ", model_text).strip()
        if not _clean_vehicle_model(model_text):
            return "", "", year_from, year_to

        if re.search(r"\s-\s*", model_text):
            base, sub = [x.strip() for x in re.split(r"\s-\s*", model_text, maxsplit=1)]
        else:
            m_code = re.match(
                r"^(?P<base>[A-Za-z0-9\u0590-\u05FF\s]+?)\s+(?P<sub>[A-Z]\d{1,3}(?:\s+[A-Z]{2,8})?)$",
                model_text,
                flags=re.IGNORECASE,
            )
            if m_code:
                base, sub = m_code.group("base").strip(), m_code.group("sub").upper()
            else:
                base, sub = model_text, ""

        base = canonicalize_vehicle_model_for_manufacturer(target_mfr, base)
        sub = normalize_vehicle_submodel_name(sub)
        return base, sub, year_from, year_to

    def _parse_fitment_row(row: Any, stype: str) -> Optional[Dict[str, str]]:
        row = list(row)
        if stype == "F":
            if len(row) < 8:
                return None
            model_raw = (str(row[0]).strip() if row[0] is not None else "")
            name = (str(row[6]).strip() if row[6] is not None else "")
            catalog = (str(row[7]).strip() if row[7] is not None else "")
        elif stype == "A":
            if len(row) < 8:
                return None
            model_raw = (str(row[7]).strip() if row[7] is not None else "")
            name = (str(row[2]).strip() if row[2] is not None else "")
            catalog = (str(row[1]).strip() if row[1] is not None else "")
        else:
            return None
        if not model_raw or not name or not catalog:
            return None
        return {"model_raw": model_raw, "name": name, "catalog": catalog}

    hierarchy_rows = (await db.execute(text("""
        SELECT manufacturer, model, sub_model, year_from, year_to
        FROM vehicle_hierarchy_xls
    """))).fetchall()
    hierarchy_map: Dict[Tuple[str, str, str], Tuple[int, int]] = {}
    for mfr, model, sub_model, year_from, year_to in hierarchy_rows:
        key = (
            str(mfr or "").strip().casefold(),
            normalize_vehicle_model_name(str(model or "")).casefold(),
            normalize_vehicle_submodel_name(str(sub_model or "")).casefold(),
        )
        if key[0] and key[1]:
            hierarchy_map[key] = (int(year_from or 0), int(year_to or 0))

    def _load_fitment_xlsx_rows(path, s_map):
        """Load fitment rows from workbook synchronously (runs in a thread)."""
        import openpyxl as _xl
        wb = _xl.load_workbook(path, read_only=True, data_only=True)
        result = {}
        for s_name, (_, _sku_pfx, s_type) in s_map.items():
            if s_name not in wb.sheetnames:
                continue
            ws = wb[s_name]
            start = 8 if s_type == "F" else 3
            result[s_name] = list(ws.iter_rows(min_row=start, values_only=True))
        wb.close()
        return result

    fitment_sheet_rows = await asyncio.to_thread(_load_fitment_xlsx_rows, xlsx_path, sheet_map)

    matched_rows = 0
    updated_parts = 0
    fitment_rows = 0

    for sheet_name, (raw_mfr, sku_prefix, stype) in sheet_map.items():
        if sheet_name not in fitment_sheet_rows:
            continue

        canonical_mfr = normalize_manufacturer_name(raw_mfr, raw_mfr)
        mfr_variants = [canonical_mfr, raw_mfr, sheet_name]

        part_rows = (await db.execute(text("""
            SELECT id, sku, oem_number, name, compatible_vehicles
            FROM parts_catalog
            WHERE manufacturer IS NOT NULL
              AND (
                    LOWER(TRIM(manufacturer)) = LOWER(TRIM(:m0))
                 OR LOWER(TRIM(manufacturer)) = LOWER(TRIM(:m1))
                 OR LOWER(TRIM(manufacturer)) = LOWER(TRIM(:m2))
              )
        """), {"m0": canonical_mfr, "m1": raw_mfr, "m2": sheet_name})).fetchall()

        by_sku: Dict[str, Tuple[str, Any]] = {}
        by_oem: Dict[str, Tuple[str, Any]] = {}
        by_name: Dict[str, Tuple[str, Any]] = {}
        part_compat_by_id: Dict[str, List[Any]] = {}
        for _pk, (part_id, sku, oem_number, name, compat) in enumerate(part_rows):
            if _pk % 500 == 0:
                await asyncio.sleep(0)
            pid = str(part_id)
            part_compat_by_id[pid] = list(compat or []) if isinstance(compat, list) else []
            if sku:
                sku_str = str(sku).strip()
                by_sku[sku_str.upper()] = (pid, compat)
                if sku_str.upper().startswith(sku_prefix):
                    by_sku[sku_str[len(sku_prefix):].upper()] = (pid, compat)
            if oem_number:
                by_oem[str(oem_number).strip().upper()] = (pid, compat)
            if name:
                by_name[str(name).strip()] = (pid, compat)

        part_fitments: Dict[str, List[Dict[str, Any]]] = {}
        seen_rows: set[Tuple[str, str, str]] = set()
        for _xi, row in enumerate(fitment_sheet_rows[sheet_name]):
            if _xi % 200 == 0:
                await asyncio.sleep(0)
            rec = _parse_fitment_row(row, stype)
            if not rec:
                continue

            model, sub_model, year_from, year_to = _split_model_submodel(rec["model_raw"], mfr_variants)
            if not model:
                continue

            target_mfr = canonical_mfr
            if canonical_mfr.casefold() == "citroen" and model.casefold().startswith("partner"):
                target_mfr = "Peugeot"

            row_key = (rec["catalog"].upper(), model.casefold(), sub_model.casefold())
            if row_key in seen_rows:
                continue
            seen_rows.add(row_key)

            match = by_sku.get(rec["catalog"].upper()) or by_oem.get(rec["catalog"].upper()) or by_name.get(rec["name"])
            if not match:
                continue

            span = hierarchy_map.get((target_mfr.casefold(), model.casefold(), sub_model.casefold()))
            if span:
                year_from, year_to = span
            if not year_from or not year_to:
                rule_span = GENERATION_YEAR_RULES.get((target_mfr.casefold(), model.casefold(), sub_model.casefold()))
                if rule_span:
                    year_from, year_to = rule_span

            part_id, _existing = match
            entry: Dict[str, Any] = {
                "manufacturer": target_mfr,
                "model": model,
                "source": "parts_database.xlsx",
            }
            if sub_model:
                entry["sub_model"] = sub_model
            if year_from and year_to and 1990 <= int(year_from) <= int(year_to) <= 2027:
                entry["year_from"] = int(year_from)
                entry["year_to"] = int(year_to)

            part_fitments.setdefault(part_id, []).append(entry)
            matched_rows += 1

        for _xj, (part_id, entries) in enumerate(part_fitments.items()):
            if _xj % 50 == 0:
                await asyncio.sleep(0)
            preserved = [
                item for item in part_compat_by_id.get(part_id, [])
                if not (isinstance(item, dict) and item.get("source") == "parts_database.xlsx")
            ]
            merged: List[Dict[str, Any]] = []
            seen_json = set()
            for item in preserved + entries:
                if not isinstance(item, dict):
                    continue
                key = json.dumps(item, sort_keys=True, ensure_ascii=False)
                if key in seen_json:
                    continue
                seen_json.add(key)
                merged.append(item)

            await db.execute(text("""
                UPDATE parts_catalog
                SET compatible_vehicles = CAST(:compat AS jsonb),
                    updated_at = NOW()
                WHERE id = CAST(:part_id AS uuid)
            """), {
                "part_id": part_id,
                "compat": json.dumps(merged, ensure_ascii=False),
            })
            updated_parts += 1
            fitment_rows += len(entries)

    await db.commit()
    return {
        "task": "backfill_catalog_fitment_from_xls",
        "status": "ok",
        "matched_rows": matched_rows,
        "updated_parts": updated_parts,
        "fitment_rows": fitment_rows,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }



# ---------------------------------------------------------------------------
# BMW chassis-code → fitment backfill
# ---------------------------------------------------------------------------
_BMW_CHASSIS_MAP: dict = {
    # E-series
    "E30": ("3 Series", 1982, 1994), "E36": ("3 Series", 1990, 2002),
    "E46": ("3 Series", 1998, 2006), "E90": ("3 Series", 2004, 2013),
    "E91": ("3 Series", 2004, 2013), "E92": ("3 Series", 2005, 2013),
    "E93": ("3 Series", 2006, 2013), "E34": ("5 Series", 1988, 1996),
    "E39": ("5 Series", 1995, 2004), "E60": ("5 Series", 2003, 2010),
    "E61": ("5 Series", 2003, 2010), "E38": ("7 Series", 1994, 2001),
    "E65": ("7 Series", 2001, 2008), "E66": ("7 Series", 2001, 2008),
    "E70": ("X5",       2006, 2013), "E71": ("X6",       2008, 2014),
    "E72": ("X6",       2009, 2014), "E81": ("1 Series",  2004, 2013),
    "E82": ("1 Series",  2007, 2013), "E84": ("X1",       2009, 2015),
    "E85": ("Z4",       2002, 2008), "E86": ("Z4",       2002, 2008),
    "E87": ("1 Series",  2003, 2013), "E88": ("1 Series",  2007, 2013),
    "E89": ("Z4",       2009, 2016), "E52": ("Z8",       2000, 2003),
    "E63": ("6 Series", 2004, 2010), "E64": ("6 Series", 2004, 2010),
    # F-series (cars)
    "F01": ("7 Series", 2008, 2015), "F02": ("7 Series", 2008, 2015),
    "F06": ("6 Series", 2012, 2018), "F07": ("5 Series GT", 2009, 2017),
    "F10": ("5 Series", 2009, 2017), "F11": ("5 Series", 2009, 2017),
    "F12": ("6 Series", 2011, 2018), "F13": ("6 Series", 2011, 2018),
    "F15": ("X5",       2013, 2018), "F16": ("X6",       2014, 2019),
    "F18": ("5 Series", 2011, 2017), "F20": ("1 Series",  2011, 2019),
    "F21": ("1 Series",  2011, 2019), "F22": ("2 Series", 2013, 2021),
    "F23": ("2 Series", 2013, 2021), "F25": ("X3",       2010, 2017),
    "F26": ("X4",       2013, 2018), "F30": ("3 Series", 2011, 2019),
    "F31": ("3 Series", 2012, 2019), "F32": ("4 Series", 2013, 2021),
    "F33": ("4 Series", 2013, 2021), "F34": ("3 Series GT", 2013, 2019),
    "F36": ("4 Series", 2013, 2021), "F39": ("X2",       2018, 2023),
    "F40": ("1 Series",  2019, 2026), "F44": ("2 Series GC", 2020, 2026),
    "F45": ("2 Series Active", 2014, 2021),
    "F46": ("2 Series Gran", 2014, 2021), "F48": ("X1",   2015, 2022),
    "F49": ("X1",       2015, 2022), "F70": ("7 Series", 2022, 2026),
    "F80": ("M3",       2014, 2020), "F82": ("M4",       2014, 2020),
    "F83": ("M4",       2014, 2020), "F85": ("X5 M",     2014, 2018),
    "F86": ("X6 M",     2014, 2018), "F87": ("M2",       2015, 2021),
    "F90": ("M5",       2017, 2026), "F91": ("M8",       2019, 2026),
    "F92": ("M8",       2019, 2026), "F93": ("M8",       2020, 2026),
    "F95": ("X5 M",     2019, 2026), "F96": ("X6 M",     2020, 2026),
    "F97": ("X3 M",     2019, 2026), "F98": ("X4 M",     2019, 2026),
    # G-series (cars)
    "G01": ("X3",       2017, 2026), "G02": ("X4",       2018, 2026),
    "G05": ("X5",       2018, 2026), "G06": ("X6",       2019, 2026),
    "G07": ("X7",       2018, 2026), "G08": ("iX3",      2020, 2026),
    "G09": ("XM",       2022, 2026),
    "G11": ("7 Series", 2015, 2026), "G12": ("7 Series", 2015, 2026),
    "G15": ("8 Series", 2018, 2026), "G16": ("8 Series GC", 2019, 2026),
    "G20": ("3 Series", 2018, 2026), "G21": ("3 Series", 2019, 2026),
    "G22": ("4 Series", 2020, 2026), "G23": ("4 Series", 2020, 2026),
    "G26": ("i4",       2021, 2026), "G29": ("Z4",       2018, 2026),
    "G30": ("5 Series", 2016, 2026), "G31": ("5 Series", 2016, 2026),
    "G32": ("6 Series GT", 2017, 2026), "G38": ("5 Series", 2017, 2026),
    "G42": ("2 Series", 2021, 2026), "G43": ("2 Series", 2021, 2026),
    "G45": ("X3",       2024, 2026), "G60": ("5 Series", 2023, 2026),
    "G70": ("7 Series", 2022, 2026),
    "G80": ("M3",       2020, 2026), "G81": ("M3",       2021, 2026),
    "G82": ("M4",       2020, 2026), "G83": ("M4",       2021, 2026),
    "G87": ("M2",       2022, 2026),
    # U-series (latest generation)
    "U10": ("X2",         2023, 2026), "U11": ("X1",       2022, 2026),
    "U25": ("X2 Electric",2024, 2026),
    # i / iX
    "I01": ("i3",       2013, 2022), "I12": ("i8",       2014, 2020),
    "I20": ("iX",       2021, 2026),
    # BMW Motorcycles
    "R17":    ("R 1200 GS",  2004, 2013), "R18":    ("R 18",       2021, 2026),
    "R19":    ("R 1300 GS",  2023, 2026),
    "K50":    ("K 1600",     2010, 2026), "K51":    ("K 1600 GT",  2012, 2026),
    "G310R":  ("G 310 R",    2016, 2026),
    "F800GS": ("F 800 GS",   2008, 2018), "F900GS": ("F 900 GS",   2018, 2026),
    "F750GS": ("F 750 GS",   2018, 2026), "F850GS": ("F 850 GS",   2018, 2026),
}

# MINI chassis codes found in BMW-manufacturer parts (BMW OEM parts for MINI vehicles).
# When detected in a BMW part, emit fitment with manufacturer='MINI' not 'BMW'.
_BMW_MINI_PASSTHROUGH: dict = {
    "F54": ("MINI Clubman",     2015, 2024),
    "F55": ("MINI 5-door",      2014, 2023),
    "F56": ("MINI 3-door",      2014, 2023),
    "F57": ("MINI Convertible", 2015, 2023),
    "F60": ("MINI Countryman",  2016, 2024),
    "F66": ("MINI 3-door",      2024, 2026),
}

_BMW_CHASSIS_RE = re.compile(
    r'(?<![A-Za-z0-9])' +
    r'([EFGIU]-?\d{2,3}' +
    r'|[KR]-?\d{2,3}' +
    r'|G310[GR]S?|F[789]00[A-Z]{1,3}|F750GS|F850GS)' +
    r'(?![0-9])',
    re.IGNORECASE,
)


async def backfill_bmw_fitment_from_name_he(db: AsyncSession) -> Dict[str, Any]:
    """
    Parse BMW chassis codes (E46, F30, G20, etc.) from parts_catalog.name_he
    and insert rows into part_vehicle_fitment for every BMW part that has no
    fitment yet.  Afterwards, run merge_catalog_fitment_from_part_vehicle_fitment
    to promote those rows into compatible_vehicles.
    """
    t0 = time.monotonic()
    await ensure_part_vehicle_fitment_table(db)

    # Fetch BMW parts without fitment
    rows = (await db.execute(text("""
        SELECT pc.id, pc.name_he, pc.name
        FROM parts_catalog pc
        WHERE pc.manufacturer = 'BMW'
          AND pc.is_active = TRUE
          AND NOT EXISTS (
              SELECT 1 FROM part_vehicle_fitment pvf WHERE pvf.part_id = pc.id
          )
        ORDER BY pc.id
    """))).fetchall()

    inserted = 0
    skipped = 0
    parts_matched = 0

    for _bi, (part_id, name_he, name_en) in enumerate(rows):
        if _bi % 200 == 0:
            await asyncio.sleep(0)
        text_to_scan = (name_he or "") + " " + (name_en or "")
        raw = {m.upper().replace('-', '') for m in _BMW_CHASSIS_RE.findall(text_to_scan)}
        bmw_mapped  = [(c, _BMW_CHASSIS_MAP[c])       for c in raw if c in _BMW_CHASSIS_MAP]
        mini_mapped = [(c, _BMW_MINI_PASSTHROUGH[c])  for c in raw if c in _BMW_MINI_PASSTHROUGH]
        if not bmw_mapped and not mini_mapped:
            skipped += 1
            continue

        parts_matched += 1
        _MINI_MFR_ID = "47a433bf-4f6f-4f8f-a686-a8c02f7727a8"
        _BMW_MFR_ID  = "caa6ba39-02aa-4394-969d-a15f3f19104c"
        for chassis, (model_series, yf, yt) in bmw_mapped:
            try:
                await db.execute(text("""
                    INSERT INTO part_vehicle_fitment
                        (id, part_id, manufacturer, manufacturer_id,
                         model, year_from, year_to, notes, updated_at)
                    VALUES (gen_random_uuid(), :part_id, 'BMW', :mfr_id,
                            :model, :yf, :yt, :notes, NOW())
                    ON CONFLICT (part_id, manufacturer, model, year_from) DO NOTHING
                """), {
                    "part_id": str(part_id),
                    "mfr_id":  _BMW_MFR_ID,
                    "model":   model_series,
                    "yf":      yf,
                    "yt":      min(yt, 2026),
                    "notes":   f"chassis:{chassis}",
                })
                inserted += 1
            except Exception:
                pass
        for chassis, (model_series, yf, yt) in mini_mapped:
            try:
                await db.execute(text("""
                    INSERT INTO part_vehicle_fitment
                        (id, part_id, manufacturer, manufacturer_id,
                         model, year_from, year_to, notes, updated_at)
                    VALUES (gen_random_uuid(), :part_id, 'MINI', :mfr_id,
                            :model, :yf, :yt, :notes, NOW())
                    ON CONFLICT (part_id, manufacturer, model, year_from) DO NOTHING
                """), {
                    "part_id": str(part_id),
                    "mfr_id":  _MINI_MFR_ID,
                    "model":   model_series,
                    "yf":      yf,
                    "yt":      min(yt, 2026),
                    "notes":   f"chassis:{chassis}",
                })
                inserted += 1
            except Exception:
                pass

        if parts_matched % 500 == 0:
            await db.commit()

    await db.commit()
    logger.info(
        "backfill_bmw_fitment_from_name_he: parts_matched=%d inserted=%d skipped=%d",
        parts_matched, inserted, skipped,
    )
    return {
        "task":          "backfill_bmw_fitment_from_name_he",
        "status":        "ok",
        "parts_matched": parts_matched,
        "inserted":      inserted,
        "skipped_no_code": skipped,
        "elapsed_s":     round(time.monotonic() - t0, 2),
    }


# -------------------------------------------------------------------------
# Ford model patterns (used by backfill_ford_fitment_from_name_he)
# -------------------------------------------------------------------------
_FORD_MODEL_PATTERNS = [
    # F-truck series
    (r'F-?150',                          "F-150",           1975, 2026),
    (r'F-?250',                          "F-250",           1980, 2026),
    (r'F-?350',                          "F-350",           1980, 2026),
    (r'F-?450',                          "F-450",           1999, 2026),
    (r'F-?520',                          "F-520",           2000, 2026),
    (r'F-?530',                          "F-530",           2000, 2026),
    (r'F-?550',                          "F-550",           1999, 2026),
    (r'F-?600',                          "F-600",           2020, 2026),
    # Cars & crossovers
    (r'MACH.?E|MACHE',                   "Mustang Mach-E",  2021, 2026),
    (r'\bMUSTANG|\bMUST\b|\bMUS\s|\bMUS\'', "Mustang",         1964, 2026),
    (r'\bBRONCO',                        "Bronco",           2021, 2026),
    (r'\bEXPLORER|\bEXP\s|\bEXP\'', "Explorer",        1990, 2026),
    (r'\bRANGER',                        "Ranger",           1983, 2026),
    (r'\bEDGE\b|\bEDG\s|\bEDG\'',  "Edge",            2007, 2023),
    (r'\bESCAPE',                        "Escape",           2000, 2026),
    (r'\bFIESTA|פיאסטה',                "Fiesta",          1976, 2023),
    (r'\bFOCUS|\bFOC\s|\bFOC-|\bFOC\'|פוקוס', "Focus", 1998, 2023),
    (r'\bFUSION',                        "Fusion",           2005, 2020),
    (r'\bTRANSIT|\bTRN\s|\bTRN-|טרנזיט', "Transit", 1965, 2026),
    (r'\bMAVERICK',                      "Maverick",         2021, 2026),
    (r'\bEXPEDITION',                    "Expedition",       1997, 2026),
    (r'\bKUGA|קוגה',                    "Kuga",            2008, 2026),
    (r'\bPUMA\b',                       "Puma",             2019, 2026),
    (r'\bTAURUS',                        "Taurus",           1985, 2026),
    (r'\bMONDEO|מונדיאו',             "Mondeo",          1992, 2022),
    (r'\bCONNECT',                       "Transit Connect",  2002, 2026),
    (r'\bTOWN.?CAR',                     "Lincoln Town Car", 1981, 2011),
    (r'\bGALAXY',                        "Galaxy",           1995, 2006),
    (r'\bC-MAX',                         "C-MAX",            2003, 2019),
    (r'\bS-MAX',                         "S-MAX",            2006, 2014),
    (r'\bFLEX\b',                       "Flex",             2009, 2019),
    (r'\bECOSPORT|\bECO\s',            "EcoSport",         2003, 2023),
]


async def backfill_ford_fitment_from_name_he(db: AsyncSession) -> Dict[str, Any]:
    """Extract Ford model names from name_he and insert part_vehicle_fitment rows."""
    t0 = time.monotonic()
    await ensure_part_vehicle_fitment_table(db)

    _FORD_MFR_ID = "73fc77ef-5414-4270-9476-2444d8b7eb41"
    compiled = [
        (re.compile(p, re.IGNORECASE), model, yf, yt)
        for p, model, yf, yt in _FORD_MODEL_PATTERNS
    ]

    rows = (await db.execute(text("""
        SELECT pc.id, pc.name_he, pc.name
        FROM parts_catalog pc
        WHERE pc.manufacturer = 'Ford'
          AND pc.is_active = TRUE
          AND NOT EXISTS (
              SELECT 1 FROM part_vehicle_fitment pvf WHERE pvf.part_id = pc.id
          )
        ORDER BY pc.id
    """))).fetchall()

    inserted = 0
    skipped = 0
    parts_matched = 0

    for _fi, (part_id, name_he, name_en) in enumerate(rows):
        if _fi % 200 == 0:
            await asyncio.sleep(0)
        text_to_scan = (name_he or "") + " " + (name_en or "")
        matched_models: set = set()
        for pattern, model, yf, yt in compiled:
            if pattern.search(text_to_scan):
                matched_models.add((model, yf, yt))

        if not matched_models:
            skipped += 1
            continue

        parts_matched += 1
        for model, yf, yt in matched_models:
            try:
                await db.execute(text("""
                    INSERT INTO part_vehicle_fitment
                        (id, part_id, manufacturer, manufacturer_id,
                         model, year_from, year_to, notes, updated_at)
                    VALUES (gen_random_uuid(), :part_id, 'Ford', :mfr_id,
                            :model, :yf, :yt, :notes, NOW())
                    ON CONFLICT (part_id, manufacturer, model, year_from) DO NOTHING
                """), {
                    "part_id": str(part_id),
                    "mfr_id":  _FORD_MFR_ID,
                    "model":   f"Ford {model}",
                    "yf":      yf,
                    "yt":      yt,
                    "notes":   "extracted:name_he",
                })
                inserted += 1
            except Exception:
                pass

        if parts_matched % 500 == 0:
            await db.commit()

    await db.commit()
    logger.info(
        "backfill_ford_fitment_from_name_he: parts_matched=%d inserted=%d skipped=%d",
        parts_matched, inserted, skipped,
    )
    return {
        "task":             "backfill_ford_fitment_from_name_he",
        "status":           "ok",
        "parts_matched":    parts_matched,
        "inserted":         inserted,
        "skipped_no_match": skipped,
        "elapsed_s":        round(time.monotonic() - t0, 2),
    }


# -------------------------------------------------------------------------
# MINI chassis map + regex (used by backfill_mini_fitment_from_name_he)
# -------------------------------------------------------------------------
_MINI_CHASSIS_MAP: dict = {
    "R50": ("MINI One/Cooper",  2001, 2006),
    "R52": ("MINI Convertible", 2004, 2007),
    "R53": ("MINI Cooper S",    2002, 2006),
    "R55": ("MINI Clubman",     2007, 2014),
    "R56": ("MINI Hatchback",   2007, 2013),
    "R57": ("MINI Convertible", 2008, 2015),
    "R58": ("MINI Coupe",       2011, 2015),
    "R59": ("MINI Roadster",    2012, 2015),
    "R60": ("MINI Countryman",  2010, 2016),
    "R61": ("MINI Paceman",     2012, 2016),
    "F54": ("MINI Clubman",     2015, 2024),
    "F55": ("MINI 5-door",      2014, 2023),
    "F56": ("MINI 3-door",      2014, 2023),
    "F57": ("MINI Convertible", 2015, 2023),
    "F60": ("MINI Countryman",  2016, 2024),
    "F66": ("MINI 3-door",      2024, 2026),
}

_MINI_CHASSIS_RE = re.compile(
    r"(?<![A-Za-z0-9])([RF]-?\d{2,3})(?![0-9])",
    re.IGNORECASE,
)


async def backfill_mini_fitment_from_name_he(db: AsyncSession) -> Dict[str, Any]:
    """Extract MINI chassis codes from name_he/name, insert part_vehicle_fitment rows."""
    t0 = time.monotonic()
    await ensure_part_vehicle_fitment_table(db)

    _MINI_MFR_ID = "47a433bf-4f6f-4f8f-a686-a8c02f7727a8"

    rows = (await db.execute(text("""
        SELECT pc.id, pc.name_he, pc.name
        FROM parts_catalog pc
        WHERE pc.manufacturer = 'MINI'
          AND pc.is_active = TRUE
          AND NOT EXISTS (
              SELECT 1 FROM part_vehicle_fitment pvf WHERE pvf.part_id = pc.id
          )
        ORDER BY pc.id
    """))).fetchall()

    parts_matched = 0
    inserted = 0
    skipped = 0

    for _mi, (part_id, name_he, name_en) in enumerate(rows):
        if _mi % 200 == 0:
            await asyncio.sleep(0)
        text_to_scan = (name_he or "") + " " + (name_en or "")
        raw_codes = {m.upper().replace("-", "") for m in _MINI_CHASSIS_RE.findall(text_to_scan)}
        mapped = [(c, _MINI_CHASSIS_MAP[c]) for c in raw_codes if c in _MINI_CHASSIS_MAP]
        if not mapped:
            skipped += 1
            continue
        parts_matched += 1
        for chassis, (model_series, yf, yt) in mapped:
            result = await db.execute(text("""
                INSERT INTO part_vehicle_fitment
                    (id, part_id, manufacturer, manufacturer_id,
                     model, year_from, year_to, notes, updated_at)
                VALUES (gen_random_uuid(), :part_id, 'MINI', :mfr_id,
                        :model, :yf, :yt, :notes, NOW())
                ON CONFLICT (part_id, manufacturer, model, year_from) DO NOTHING
            """), {
                "part_id": str(part_id),
                "mfr_id":  _MINI_MFR_ID,
                "model":   model_series,
                "yf":      yf,
                "yt":      min(yt, 2026),
                "notes":   f"chassis:{chassis}",
            })
            inserted += result.rowcount

    await db.commit()
    logger.info(
        "backfill_mini_fitment_from_name_he: parts_matched=%d inserted=%d skipped=%d",
        parts_matched, inserted, skipped,
    )
    return {
        "task":            "backfill_mini_fitment_from_name_he",
        "status":          "ok",
        "parts_matched":   parts_matched,
        "inserted":        inserted,
        "skipped_no_code": skipped,
        "elapsed_s":       round(time.monotonic() - t0, 2),
    }


# Jaguar model → (canonical_model_name, year_from, year_to)
# Ordered most-specific first to prevent partial match overlaps.
_JAGUAR_MODEL_MAP: list[tuple[re.Pattern, str, int, int]] = [
    (re.compile(r"(?<![A-Za-z0-9])F[-\s]PACE(?![A-Za-z0-9])", re.IGNORECASE), "F-Pace",  2016, 2025),
    (re.compile(r"(?<![A-Za-z0-9])E[-\s]PACE(?![A-Za-z0-9])", re.IGNORECASE), "E-Pace",  2018, 2024),
    (re.compile(r"(?<![A-Za-z0-9])F[-\s]TYPE(?![A-Za-z0-9])", re.IGNORECASE), "F-Type",  2013, 2023),
    (re.compile(r"(?<![A-Za-z0-9])I[-\s]PACE(?![A-Za-z0-9])", re.IGNORECASE), "I-Pace",  2018, 2024),
    (re.compile(r"(?<![A-Za-z0-9])S[-\s]TYPE(?![A-Za-z0-9])", re.IGNORECASE), "S-Type",  1999, 2008),
    (re.compile(r"(?<![A-Za-z0-9])X[-\s]TYPE(?![A-Za-z0-9])", re.IGNORECASE), "X-Type",  2001, 2010),
    (re.compile(r"(?<![A-Za-z0-9])XKR(?![A-Za-z0-9])",        re.IGNORECASE), "XKR",     2000, 2014),
    (re.compile(r"(?<![A-Za-z0-9])XK8(?![A-Za-z0-9])",        re.IGNORECASE), "XK8",     2000, 2013),
    (re.compile(r"(?<![A-Za-z0-9])XK(?![A-Za-z0-9])",         re.IGNORECASE), "XK",      2000, 2014),
    (re.compile(r"(?<![A-Za-z0-9])XJR(?![A-Za-z0-9])",        re.IGNORECASE), "XJR",     2000, 2007),
    (re.compile(r"(?<![A-Za-z0-9])XJ8(?![A-Za-z0-9])",        re.IGNORECASE), "XJ8",     2000, 2005),
    (re.compile(r"(?<![A-Za-z0-9])XJ6(?![A-Za-z0-9])",        re.IGNORECASE), "XJ6",     2004, 2009),
    (re.compile(r"(?<![A-Za-z0-9])XJ(?![A-Za-z0-9])",         re.IGNORECASE), "XJ",      2000, 2018),
    (re.compile(r"(?<![A-Za-z0-9])XE[-\s]?S(?![A-Za-z0-9])",  re.IGNORECASE), "XE",      2015, 2017),
    (re.compile(r"(?<![A-Za-z0-9])XE(?![A-Za-z0-9])",         re.IGNORECASE), "XE",      2015, 2022),
    (re.compile(r"(?<![A-Za-z0-9])XF[-\s]?[RS]?(?![A-Za-z0-9])", re.IGNORECASE), "XF",   2008, 2018),
]
_JAGUAR_MFR_ID = "fde0f2dc-c6fb-4ab6-b699-765044fbc073"


async def backfill_jaguar_fitment_from_name(db: AsyncSession) -> Dict[str, Any]:
    """
    Parse Jaguar model names (XE, XF, XJ, F-Pace, E-Pace, etc.) from
    parts_catalog.name and name_he, then insert rows into part_vehicle_fitment
    for Jaguar parts that currently have no fitment data.

    Uses Israeli vehicle registry year ranges from vehicle_market_il.
    Mirrors the BMW chassis-code approach — no external API calls needed.
    Processes up to 2000 parts per run; idempotent via ON CONFLICT DO NOTHING.
    """
    t0 = time.monotonic()
    await ensure_part_vehicle_fitment_table(db)

    # Exclude branded merchandise: clothing, accessories, and non-automotive items
    _MERCH_EXCLUDE = re.compile(
        r'\b(shirt|polo|shoe|shoes|boot|jacket|wallet|keyring|key\s+ring|'
        r'mug|notebook|ebook|e-book|badge|scarf|hat\b|bag\b|scale\s+model|'
        r'umbrella|cap\b|sock|cufflink|glove|passport|money\s+clip|suede|'
        r'driving\s+shoe|print\b|sticker|poster|pennant|artwork|miniature)\b',
        re.IGNORECASE,
    )

    rows = (await db.execute(text("""
        SELECT pc.id, pc.name, pc.name_he
        FROM parts_catalog pc
        WHERE pc.manufacturer = 'Jaguar'
          AND pc.is_active = TRUE
          AND (pc.category IS NULL OR pc.category != 'accessories')
          AND NOT EXISTS (
              SELECT 1 FROM part_vehicle_fitment pvf WHERE pvf.part_id = pc.id
          )
        ORDER BY pc.id
        LIMIT 2000
    """))).fetchall()

    inserted = 0
    skipped = 0
    parts_matched = 0

    for idx, (part_id, name, name_he) in enumerate(rows):
        if idx % 50 == 0:
            await asyncio.sleep(0)

        text_to_scan = " ".join(filter(None, [name or "", name_he or ""]))

        # Skip branded merchandise — clothing, lifestyle, non-automotive items
        if _MERCH_EXCLUDE.search(text_to_scan):
            skipped += 1
            continue

        matched_models: list[tuple[str, int, int]] = []
        seen_models: set[str] = set()

        for pattern, model_name, yf, yt in _JAGUAR_MODEL_MAP:
            if pattern.search(text_to_scan) and model_name not in seen_models:
                matched_models.append((model_name, yf, yt))
                seen_models.add(model_name)

        if not matched_models:
            skipped += 1
            continue

        parts_matched += 1
        for model_name, yf, yt in matched_models:
            try:
                await db.execute(text("""
                    INSERT INTO part_vehicle_fitment
                        (id, part_id, manufacturer, manufacturer_id,
                         model, year_from, year_to, notes, updated_at)
                    VALUES (gen_random_uuid(), :part_id, 'Jaguar', :mfr_id,
                            :model, :yf, :yt, 'name_parse:jaguar', NOW())
                    ON CONFLICT (part_id, manufacturer, model, year_from) DO NOTHING
                """), {
                    "part_id": str(part_id),
                    "mfr_id":  _JAGUAR_MFR_ID,
                    "model":   model_name,
                    "yf":      yf,
                    "yt":      yt,
                })
                inserted += 1
            except Exception:
                pass

    with contextlib.suppress(Exception):
        await db.commit()

    return {
        "task":          "backfill_jaguar_fitment_from_name",
        "status":        "ok",
        "scanned":       len(rows),
        "parts_matched": parts_matched,
        "inserted":      inserted,
        "skipped":       skipped,
        "elapsed_s":     round(time.monotonic() - t0, 2),
    }


async def merge_catalog_fitment_from_part_vehicle_fitment(db: AsyncSession) -> Dict[str, Any]:
    """
    Promote part_vehicle_fitment rows into parts_catalog.compatible_vehicles.
    Processes one manufacturer at a time (no full-table fetchall) to avoid OOM.
    Year filter: 1985 <= year_from <= year_to <= 2030.
    """
    t0 = time.monotonic()
    await ensure_part_vehicle_fitment_table(db)

    def _normalize_fitment_json_list(items):
        unique, seen = [], set()
        for item in items:
            if not isinstance(item, dict):
                continue
            key = json.dumps(item, sort_keys=True, ensure_ascii=False)
            if key not in seen:
                seen.add(key)
                unique.append(item)
        unique.sort(key=lambda r: json.dumps(r, sort_keys=True, ensure_ascii=False))
        return unique

    # Step 1: get distinct manufacturers that have fitment data
    mfr_rows = (await db.execute(text("""
        SELECT DISTINCT manufacturer
        FROM part_vehicle_fitment
        WHERE manufacturer IS NOT NULL AND TRIM(manufacturer) <> ''
        ORDER BY manufacturer
    """))).fetchall()

    total_updated  = 0
    total_merged   = 0
    total_scanned  = 0
    mfrs_processed = 0

    for (mfr,) in mfr_rows:
        # Step 2: per-manufacturer JOIN — at most ~100K rows, ~30 MB
        rows = (await db.execute(text("""
            SELECT
                pc.id,
                pc.compatible_vehicles,
                pvf.manufacturer,
                pvf.model,
                pvf.year_from,
                pvf.year_to,
                pvf.engine_type
            FROM parts_catalog pc
            JOIN part_vehicle_fitment pvf ON pvf.part_id = pc.id
            WHERE pvf.manufacturer = :mfr
              AND pvf.model IS NOT NULL
              AND TRIM(pvf.model) <> ''
              AND pvf.year_from IS NOT NULL
              AND pvf.year_to   IS NOT NULL
        """), {"mfr": mfr})).fetchall()

        if not rows:
            continue
        mfrs_processed += 1

        part_existing: Dict[str, list] = {}
        part_fitments: Dict[str, list] = defaultdict(list)

        for i, (part_id, compat, manufacturer, model, year_from, year_to, engine_type) in enumerate(rows):
            if i % 100 == 0:  # unconditional yield before any continue so heartbeat always fires
                await asyncio.sleep(0)
            total_scanned += 1
            pid = str(part_id)
            if pid not in part_existing:
                part_existing[pid] = list(compat or []) if isinstance(compat, list) else []

            canonical_manufacturer = normalize_manufacturer_name(
                str(manufacturer or ""), str(manufacturer or ""))
            canonical_model = canonicalize_vehicle_model_for_manufacturer(
                canonical_manufacturer, model)
            if not canonical_manufacturer or not canonical_model:
                continue

            fitment: Dict[str, Any] = {
                "manufacturer": canonical_manufacturer,
                "model":        canonical_model,
                "source":       "part_vehicle_fitment",
            }
            if engine_type:
                fitment["engine"] = str(engine_type).strip()[:50]

            try:
                yf = int(year_from or 0)
            except Exception:
                yf = 0
            try:
                yt = int(year_to or 0)
            except Exception:
                yt = 0
            if yf and not yt:
                yt = yf
            if yf and yt and 1985 <= yf <= yt <= 2030:
                fitment["year_from"] = yf
                fitment["year_to"]   = yt

            part_fitments[pid].append(fitment)

        # Step 3: update parts_catalog for this manufacturer
        updated_this_mfr = 0
        for j, (part_id, entries) in enumerate(part_fitments.items()):
            if j % 50 == 0:  # unconditional yield before continue so heartbeat always fires
                await asyncio.sleep(0)
            preserved = [
                item for item in part_existing.get(part_id, [])
                if not (isinstance(item, dict) and item.get("source") == "part_vehicle_fitment")
            ]
            merged = _normalize_fitment_json_list(preserved + entries)

            existing_json = json.dumps(
                _normalize_fitment_json_list(list(part_existing.get(part_id, []))),
                sort_keys=True, ensure_ascii=False)
            merged_json = json.dumps(merged, sort_keys=True, ensure_ascii=False)
            if existing_json == merged_json:
                continue

            await db.execute(text("""
                UPDATE parts_catalog
                SET compatible_vehicles = CAST(:compat AS jsonb),
                    updated_at = NOW()
                WHERE id = CAST(:part_id AS uuid)
            """), {"part_id": part_id,
                   "compat":  json.dumps(merged, ensure_ascii=False)})
            updated_this_mfr += 1
            total_merged += len(entries)

        await db.commit()
        await asyncio.sleep(0)  # yield after each manufacturer commit
        total_updated += updated_this_mfr
        logger.info("merge_catalog_fitment: mfr=%s scanned=%d updated=%d",
                    mfr, len(rows), updated_this_mfr)

    return {
        "task":                   "merge_catalog_fitment_from_part_vehicle_fitment",
        "status":                 "ok",
        "manufacturers_processed": mfrs_processed,
        "scanned_rows":           total_scanned,
        "updated_parts":          total_updated,
        "merged_fitment_rows":    total_merged,
        "elapsed_s":              round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 8 – Refresh min / max prices on parts_catalog
# =========================================================================

async def refresh_min_max_prices(db: AsyncSession) -> Dict[str, Any]:
    """
    Recalculate parts_catalog.min_price_ils / max_price_ils from live
    supplier_parts prices (WITH 18% VAT applied).

    DELTA-ONLY (rewritten 2026-07-07): the old version re-aggregated ALL 3.7M
    supplier_parts and re-updated EVERY matching parts_catalog row every cycle —
    a 26-30 min full-table pass that saturated the single virtual disk and made
    the box fragile under any added load. Now it only recomputes parts whose
    supplier_parts changed since the last successful run (checkpoint). A price
    change touches a few thousand rows, not millions. A weekly full pass still
    runs to catch anything missed (e.g. a part going fully unavailable).
    """
    t0 = time.monotonic()
    since = await _get_task_checkpoint(db, "refresh_min_max_prices")

    # Weekly safety full pass: if the checkpoint is >7 days old (or first run),
    # do the complete recompute once, then delta from there.
    full_pass = (datetime.utcnow() - since) > timedelta(days=7)

    try:
        await db.execute(text("SET LOCAL lock_timeout = '10min'"))
        rate = await _get_ils_rate(db)

        if full_pass:
            await db.execute(
                text(
                    """
                    WITH price_agg AS (
                        SELECT part_id,
                               MIN(COALESCE(price_ils, price_usd * :rate)) * :vat AS min_p,
                               MAX(COALESCE(price_ils, price_usd * :rate)) * :vat AS max_p
                        FROM supplier_parts WHERE is_available = TRUE
                        GROUP BY part_id
                    )
                    UPDATE parts_catalog pc
                    SET min_price_ils = pa.min_p, max_price_ils = pa.max_p, updated_at = NOW()
                    FROM price_agg pa WHERE pc.id = pa.part_id
                    """
                ),
                {"rate": rate, "vat": 1 + VAT},
            )
            mode = "full"
        else:
            # Only parts whose supplier_parts changed since the checkpoint.
            await db.execute(
                text(
                    """
                    WITH changed AS (
                        SELECT DISTINCT part_id FROM supplier_parts
                        WHERE updated_at > :since
                    ),
                    price_agg AS (
                        SELECT sp.part_id,
                               MIN(COALESCE(sp.price_ils, sp.price_usd * :rate)) * :vat AS min_p,
                               MAX(COALESCE(sp.price_ils, sp.price_usd * :rate)) * :vat AS max_p
                        FROM supplier_parts sp
                        JOIN changed c ON c.part_id = sp.part_id
                        WHERE sp.is_available = TRUE
                        GROUP BY sp.part_id
                    )
                    UPDATE parts_catalog pc
                    SET min_price_ils = pa.min_p, max_price_ils = pa.max_p, updated_at = NOW()
                    FROM price_agg pa WHERE pc.id = pa.part_id
                    """
                ),
                {"since": since, "rate": rate, "vat": 1 + VAT},
            )
            mode = "delta"

        await db.commit()
        await _save_task_checkpoint(db, "refresh_min_max_prices")
        logger.info("refresh_min_max_prices: done (%s pass)", mode)
    except Exception as exc:
        await db.rollback()
        logger.error("refresh_min_max_prices failed: %s", exc)
        return {"task": "refresh_min_max_prices", "status": "error", "error": str(exc)}

    return {
        "task": "refresh_min_max_prices",
        "status": "ok",
        "mode": mode,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Task 9 – Seed system_settings
# =========================================================================

async def seed_system_settings(db: AsyncSession) -> Dict[str, Any]:
    """
    Ensure required system_settings keys exist (insert if missing, never
    overwrite existing values).
    """
    DEFAULTS = {
        "search_results_per_type": "4",
        "search_type_order": "original,oem,aftermarket",
        "ils_per_usd": str(ILS_PER_USD),
        "vat_rate": "0.18",
    }
    t0 = time.monotonic()
    inserted = 0

    try:
        for key, value in DEFAULTS.items():
            result = await db.execute(
                text("SELECT 1 FROM system_settings WHERE key = :key"),
                {"key": key},
            )
            if not result.fetchone():
                await db.execute(
                    text(
                        "INSERT INTO system_settings (id, key, value, updated_at) "
                        "VALUES (gen_random_uuid(), :key, :value, NOW())"
                    ),
                    {"key": key, "value": value},
                )
                inserted += 1

        await db.commit()
        logger.info("seed_system_settings: inserted=%d", inserted)
    except Exception as exc:
        await db.rollback()
        logger.error("seed_system_settings failed: %s", exc)
        return {"task": "seed_system_settings", "status": "error", "error": str(exc)}

    return {
        "task": "seed_system_settings",
        "status": "ok",
        "inserted": inserted,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# =========================================================================
# Orchestrator – run_all_tasks
# =========================================================================

async def _lookup_oem_spec_task(db: AsyncSession) -> dict:
    from ai_catalog_builder import lookup_oem_spec
    return await lookup_oem_spec(db, limit=500)


async def _enrich_pending_parts_task(db: AsyncSession) -> Dict[str, Any]:
    """Thin wrapper so enrich_pending_parts integrates with TASK_REGISTRY."""
    from ai_catalog_builder import enrich_pending_parts
    return await enrich_pending_parts(db, limit=2000)


async def _trigger_scraper_for_misses_task(db: AsyncSession) -> Dict[str, Any]:
    """Find high-frequency zero-result queries (miss_count >= 3, not yet triggered)
    and fire REX brand discovery for the likely brand."""
    if not _rex_harvest_enabled():
        return {
            "task": "trigger_scraper_for_misses",
            "status": "skipped",
            "reason": "REX_HARVEST_ENABLED=false",
            "triggered": 0,
            "errors": 0,
        }

    from catalog_scraper import run_brand_discovery

    try:
        rows = (await db.execute(
            text("""
                SELECT id, query, normalized_query, vehicle_manufacturer
                FROM search_misses
                WHERE miss_count >= 3
                  AND triggered_scrape = FALSE
                ORDER BY miss_count DESC
                LIMIT 20
            """)
        )).fetchall()
    except Exception as exc:
        if "search_misses" in str(exc).lower() and "does not exist" in str(exc).lower():
            return {
                "task": "trigger_scraper_for_misses",
                "status": "skipped",
                "reason": "search_misses_table_missing",
                "triggered": 0,
                "errors": 0,
            }
        raise

    if not rows:
        return {"task": "trigger_scraper_for_misses", "status": "ok", "triggered": 0, "errors": 0}

    triggered = 0
    errors = 0
    triggered_ids = []

    for row in rows:
        # Prefer explicit vehicle_manufacturer; fall back to first token of query
        brand = (row.vehicle_manufacturer or "").strip()
        if not brand:
            first_token = (row.normalized_query or "").split()
            brand = first_token[0] if first_token else ""
        if not brand:
            continue

        try:
            asyncio.create_task(run_brand_discovery(brands=[brand]))
            triggered += 1
            triggered_ids.append(str(row.id))
        except Exception as e:
            logger.warning("trigger_scraper_for_misses: brand=%s error=%s", brand, e)
            errors += 1

    if triggered_ids:
        for _tid in triggered_ids:
            await db.execute(
                text("UPDATE search_misses SET triggered_scrape = TRUE WHERE id = :tid"),
                {"tid": _tid},
            )
    await db.commit()

    return {
        "task": "trigger_scraper_for_misses",
        "status": "ok",
        "triggered": triggered,
        "errors": errors,
    }


async def _trigger_scraper_for_registry_gaps_task(db: AsyncSession) -> Dict[str, Any]:
    """Queue brand discovery for all known manufacturers under-covered in parts_catalog.

    Pool includes:
    - active car_brands
    - active truck_brands
    - active manufacturers already present in parts_catalog

    We queue only a bounded batch per run for speed + operational safety.
    """
    if not _rex_harvest_enabled():
        return {
            "task": "trigger_scraper_for_registry_gaps",
            "status": "skipped",
            "reason": "REX_HARVEST_ENABLED=false",
            "triggered": 0,
            "target": 0,
            "per_run": 0,
            "brands": [],
        }

    from catalog_scraper import run_brand_discovery

    target = max(1, int(os.getenv("DISCOVERY_TARGET", "120")))
    per_run = min(50, max(1, int(os.getenv("DISCOVERY_PER_RUN", "20"))))
    truck_delay_days = max(0, int(os.getenv("DISCOVERY_TRUCK_DELAY_DAYS", "30")))

    include_trucks = False
    trucks_deferred_reason = "deploy_timestamp_missing"
    deployed_at_iso = (os.getenv("DEPLOYED_AT_ISO", "") or "").strip()
    trucks_unlock_at = None
    if deployed_at_iso:
        try:
            _deploy_ts = datetime.fromisoformat(deployed_at_iso.replace("Z", "+00:00"))
            if _deploy_ts.tzinfo is None:
                _deploy_ts = _deploy_ts.replace(tzinfo=timezone.utc)
            trucks_unlock_at = _deploy_ts + timedelta(days=truck_delay_days)
            include_trucks = datetime.now(timezone.utc) >= trucks_unlock_at
            trucks_deferred_reason = "within_truck_delay_window" if not include_trucks else "delay_window_completed"
        except Exception:
            include_trucks = False
            trucks_deferred_reason = "invalid_deploy_timestamp"

    rows = (await db.execute(
        text(
            """
            WITH manufacturer_pool AS (
                SELECT name
                FROM car_brands
                WHERE is_active = TRUE

                UNION

                SELECT name
                FROM truck_brands
                                WHERE is_active = TRUE
                                    AND :include_trucks = TRUE

                UNION

                SELECT DISTINCT TRIM(manufacturer) AS name
                FROM parts_catalog
                WHERE is_active = TRUE
                  AND manufacturer IS NOT NULL
                  AND TRIM(manufacturer) <> ''
            ),
            part_counts AS (
                SELECT LOWER(TRIM(manufacturer)) AS mkey, COUNT(*) AS cnt
                FROM parts_catalog
                WHERE is_active = TRUE
                  AND manufacturer IS NOT NULL
                GROUP BY LOWER(TRIM(manufacturer))
            )
                        SELECT mp.name, COALESCE(pc.cnt, 0) AS part_count
                        FROM manufacturer_pool mp
            LEFT JOIN part_counts pc
                            ON LOWER(TRIM(mp.name)) = pc.mkey
            WHERE COALESCE(pc.cnt, 0) < :target
                            AND mp.name IS NOT NULL
                            AND TRIM(mp.name) <> ''
                        ORDER BY COALESCE(pc.cnt, 0) ASC, mp.name ASC
            LIMIT :lim
            """
        ),
        {
            "target": target,
            "lim": per_run,
            "include_trucks": include_trucks,
        },
    )).fetchall()

    if not rows:
        return {
            "task": "trigger_scraper_for_registry_gaps",
            "status": "ok",
            "triggered": 0,
            "target": target,
            "per_run": per_run,
            "brands": [],
            "include_trucks": include_trucks,
            "trucks_deferred": not include_trucks,
            "trucks_deferred_reason": trucks_deferred_reason,
            "truck_delay_days": truck_delay_days,
            "deployed_at_iso": deployed_at_iso or None,
            "trucks_unlock_at": trucks_unlock_at.isoformat() if trucks_unlock_at else None,
            "reason": "no_undercovered_brands",
        }

    brands = [r[0] for r in rows if r[0]]
    if not brands:
        return {
            "task": "trigger_scraper_for_registry_gaps",
            "status": "ok",
            "triggered": 0,
            "target": target,
            "per_run": per_run,
            "brands": [],
            "include_trucks": include_trucks,
            "trucks_deferred": not include_trucks,
            "trucks_deferred_reason": trucks_deferred_reason,
            "truck_delay_days": truck_delay_days,
            "deployed_at_iso": deployed_at_iso or None,
            "trucks_unlock_at": trucks_unlock_at.isoformat() if trucks_unlock_at else None,
            "reason": "no_valid_brand_names",
        }

    asyncio.create_task(run_brand_discovery(brands=brands, target=target, per_run=per_run))

    return {
        "task": "trigger_scraper_for_registry_gaps",
        "status": "ok",
        "triggered": len(brands),
        "target": target,
        "per_run": per_run,
        "include_trucks": include_trucks,
        "trucks_deferred": not include_trucks,
        "trucks_deferred_reason": trucks_deferred_reason,
        "truck_delay_days": truck_delay_days,
        "deployed_at_iso": deployed_at_iso or None,
        "trucks_unlock_at": trucks_unlock_at.isoformat() if trucks_unlock_at else None,
        "brands": brands,
        "brand_counts": [{"name": r[0], "part_count": int(r[1] or 0)} for r in rows],
    }


async def _run_image_embedding_batch(rows: list) -> None:
    """Background worker: fetch image bytes, embed via CLIP, write vector to DB.
    Always launched via asyncio.create_task() — never awaited directly."""
    import base64
    from BACKEND_DATABASE_MODELS import async_session_factory as _sf
    from hf_client import hf_clip
    ok = 0
    async with httpx.AsyncClient() as client:
        for row in rows:
            try:
                r = await client.get(row.url, timeout=15.0, follow_redirects=True)
                r.raise_for_status()
                b64 = base64.b64encode(r.content).decode()
                vec = await hf_clip(b64, timeout=30.0)
                async with _sf() as db:
                    await db.execute(
                        text("UPDATE parts_catalog SET image_embedding = CAST(:v AS vector) WHERE id = :id"),
                        {"v": str(vec), "id": str(row.part_id)},
                    )
                    await db.execute(
                        text("UPDATE parts_images SET embedding_generated = TRUE WHERE id = :id"),
                        {"id": str(row.id)},
                    )
                    await db.commit()
                ok += 1
            except Exception as e:
                logger.warning("_run_image_embedding_batch: %s → %s", row.url[:80], e)
    logger.info("_run_image_embedding_batch: %d/%d embedded", ok, len(rows))


async def _generate_image_embeddings_task(db: AsyncSession) -> Dict[str, Any]:
    """Check for parts_images rows pending CLIP embedding; fire background batch.
    Returns immediately without blocking run_all_tasks."""
    if not os.getenv("HF_TOKEN", ""):
        return {"task": "generate_image_embeddings", "status": "ok", "triggered": 0, "note": "HF_TOKEN not set"}

    rows = (await db.execute(
        text("""
            SELECT id, part_id, url
            FROM parts_images
            WHERE embedding_generated = FALSE
              AND url IS NOT NULL
            ORDER BY is_primary DESC, created_at
            LIMIT 20
        """)
    )).fetchall()

    if not rows:
        return {"task": "generate_image_embeddings", "status": "ok", "triggered": 0}

    asyncio.create_task(_run_image_embedding_batch(rows))
    return {"task": "generate_image_embeddings", "status": "ok", "triggered": len(rows)}


async def dedup_catalog_parts(db: AsyncSession) -> Dict[str, Any]:
    """
    Remove duplicate rows in parts_catalog.

    Delta mode: only examines SKUs / (name, manufacturer) pairs where at least
    one row was modified since the last checkpoint. For those keys it checks the
    full catalog for duplicates (a recently-imported part may duplicate an old one).
    This turns a 60+ min full-table window-function scan into a <1 min targeted
    check proportional to ingestion volume, not catalog size.

    Steps:
      1. SKU dedup: null out the older row for any SKU that appears in a recently
         modified row and has a duplicate anywhere in the catalog.
      2. Name/manufacturer dedup: flag older duplicates for OEM lookup review.
    """
    t0 = time.monotonic()
    nulled_skus = 0
    flagged_dupes = 0

    since = await _get_task_checkpoint(db, "dedup_catalog_parts")

    try:
        # ── Step 1: SKU dedup — only check SKUs touched since last run ──────────
        dup_sku_ids = (await db.execute(text("""
            WITH changed_skus AS (
                SELECT DISTINCT sku FROM parts_catalog
                WHERE updated_at > :since AND sku IS NOT NULL
            )
            SELECT id FROM (
                SELECT id,
                       ROW_NUMBER() OVER (PARTITION BY sku ORDER BY created_at DESC) AS rn
                FROM parts_catalog
                WHERE sku IN (SELECT sku FROM changed_skus)
            ) ranked
            WHERE rn > 1
        """), {"since": since})).scalars().all()

        if dup_sku_ids:
            # Sort ids so the UPDATE takes row locks in a consistent (id) order —
            # reduces deadlocks with concurrent writers that also touch these rows
            # (the run_all_tasks central retry is the backstop). 2026-07-11.
            dup_sku_ids = sorted(str(i) for i in dup_sku_ids)
            r1 = await db.execute(
                text("UPDATE parts_catalog SET sku = NULL, updated_at = NOW() WHERE id = ANY(:ids) RETURNING id"),
                {"ids": dup_sku_ids},
            )
            await db.flush()
            nulled_skus = len(r1.fetchall())

        # ── Step 2: name/manufacturer dedup — only pairs touched recently ───────
        has_manufacturer_id = bool((await db.execute(text("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'parts_catalog' AND column_name = 'manufacturer_id'
            )
        """))).scalar())

        if has_manufacturer_id:
            manufacturer_partition = "manufacturer_id"
            manufacturer_filter = "manufacturer_id IS NOT NULL"
            changed_pairs_cte = """
                WITH changed_pairs AS (
                    SELECT DISTINCT lower(name) AS n, manufacturer_id AS m
                    FROM parts_catalog
                    WHERE updated_at > :since AND name IS NOT NULL AND manufacturer_id IS NOT NULL
                )
            """
            dedup_filter = "(lower(name), manufacturer_id) IN (SELECT n, m FROM changed_pairs)"
        else:
            manufacturer_partition = "lower(COALESCE(manufacturer, ''))"
            manufacturer_filter = "manufacturer IS NOT NULL"
            changed_pairs_cte = """
                WITH changed_pairs AS (
                    SELECT DISTINCT lower(name) AS n, lower(COALESCE(manufacturer,'')) AS m
                    FROM parts_catalog
                    WHERE updated_at > :since AND name IS NOT NULL AND manufacturer IS NOT NULL
                )
            """
            dedup_filter = "(lower(name), lower(COALESCE(manufacturer,''))) IN (SELECT n, m FROM changed_pairs)"

        dup_name_ids = (await db.execute(text(f"""
            {changed_pairs_cte}
            SELECT id FROM (
                SELECT id,
                       ROW_NUMBER() OVER (
                           PARTITION BY lower(name), {manufacturer_partition}
                           ORDER BY created_at DESC
                       ) AS rn
                FROM parts_catalog
                WHERE name IS NOT NULL AND {manufacturer_filter}
                  AND {dedup_filter}
            ) ranked
            WHERE rn > 1
        """), {"since": since})).scalars().all()

        if dup_name_ids:
            r2 = await db.execute(
                text("UPDATE parts_catalog SET needs_oem_lookup = TRUE, updated_at = NOW() WHERE id = ANY(:ids) RETURNING id"),
                {"ids": dup_name_ids},
            )
            await db.flush()
            flagged_dupes = len(r2.fetchall())

        await db.commit()
        await _save_task_checkpoint(db, "dedup_catalog_parts")
        logger.info("dedup_catalog_parts (delta since %s): nulled_skus=%d flagged_dupes=%d elapsed=%.1fs",
                    since.isoformat(), nulled_skus, flagged_dupes, time.monotonic() - t0)

    except Exception as exc:
        await db.rollback()
        logger.error("dedup_catalog_parts failed: %s", exc)
        return {"task": "dedup_catalog_parts", "status": "error", "error": str(exc)}

    return {
        "task": "dedup_catalog_parts",
        "status": "ok",
        "nulled_skus": nulled_skus,
        "flagged_dupes": flagged_dupes,
        "delta_since": since.isoformat(),
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# ── populate_supplier_parts constants ───────────────────────────────────────
_BATCH = 500  # rows per DB page when iterating parts_catalog

_DEFAULT_PRICE = 80.0  # ILS fallback when part has no base_price

_CATEGORY_FALLBACK_ILS: Dict[str, float] = {
    "בלמים": 150.0,
    "מתלה": 200.0,
    "היגוי": 180.0,
    "מנוע": 300.0,
    "קירור": 120.0,
    "מערכת דלק": 180.0,
    "מערכת אוויר": 80.0,
    "טורבו": 500.0,
    "פליטה": 250.0,
    "תיבת הילוכים וציר": 400.0,
    "מצמד": 200.0,
    "רצועות תזמון": 90.0,
    "הצתה": 80.0,
    "סינון": 60.0,
    "חשמל ואלקטרוניקה": 120.0,
    "חיישנים": 80.0,
    "מצבר": 250.0,
    "תאורה": 100.0,
    "מזגן וחימום": 150.0,
    "גוף הרכב": 200.0,
    "שמשות ומגבים": 80.0,
    "פנים הרכב": 100.0,
    "גלגלים וצמיגים": 100.0,
    "אטמים וצינורות": 60.0,
    "מערכת בטיחות": 300.0,
    "מערכת היברידית וחשמלי": 500.0,
    "שמנים ונוזלים": 50.0,
    "כלי עבודה ואביזרים": 60.0,
    "כללי": 80.0,
}

_WARRANTY_MAP: Dict[str, int] = {
    "original": 24, "Original": 24,
    "oe_equivalent": 12, "OE Equivalent": 12,
    "aftermarket": 12, "Aftermarket": 12,
    "economy": 6, "Economy": 6,
    "generic": 6,
    "New": 12, "Used": 3, "Remanufactured": 6,
}

# REAL_DATA_ONLY: populate_supplier_parts only links suppliers that have REAL sourced data.
# DO NOT add generic marketplace suppliers (eBay Motors, Motorstore IL) here — they produce
# fabricated price rows that violate the pricing policy. Real eBay/Motorstore rows come from
# ebay_brand_importer.py and scrape_motorstore() respectively.

# Manufacturer-specific = official importers, linked only to their own-brand parts
# price_mult = 1.0 means price_ils = base_price (our 45% margin already applied)
_MANUFACTURER_SUPPLIERS: List[Any] = [
    ("Inbar Group - Land Rover Israel", "LR-IL",  1.00, 0.0, 0.0, 7, "in_stock", True),
    ("Geo Mobility - Zeekr Israel",     "ZEEKR",  1.00, 0.0, 0.0, 7, "in_stock", True),
]

# _UNIVERSAL_SUPPLIERS intentionally empty — no fake marketplace links allowed.
# Rule: supplier_parts rows must come from real scrapers/importers only.
_UNIVERSAL_SUPPLIERS: List[Any] = []
# ────────────────────────────────────────────────────────────────────────────


async def _populate_supplier_parts_task(db: AsyncSession) -> Dict[str, Any]:
    """
    Link manufacturer-direct suppliers to their own-brand parts with computed pricing.
    Idempotent — uses ON CONFLICT DO NOTHING on (supplier_id, part_id).
    Heavy operation; only runs on demand via the admin API, not in run_all_tasks.

    REAL_DATA_ONLY rule: only manufacturer-specific suppliers are linked here.
    Generic marketplace suppliers (eBay Motors, Motorstore IL) must NOT be added to
    _UNIVERSAL_SUPPLIERS — their rows must come from real scrapers/importers only.
    price_ils = base_price (mult=1.0) — 45% margin already in base_price.
    """

    import json
    import uuid as _uuid

    t0 = time.monotonic()

    # ── Load active suppliers ────────────────────────────────────────────────
    rows = (await db.execute(text(
        "SELECT id, name, is_manufacturer, manufacturer_name "
        "FROM suppliers WHERE is_active = TRUE ORDER BY priority"
    ))).mappings().fetchall()
    suppliers: Dict[str, Any] = {r["name"]: dict(r) for r in rows}

    if not suppliers:
        return {"task": "populate_supplier_parts", "status": "error",
                "error": "No active suppliers found — run seed_data.py first"}

    ils_rate: float = ILS_PER_USD  # from module-level constant (overridden by settings)
    total_inserted = 0

    async def _insert_batch(records: list) -> int:
        if not records:
            return 0
        payload = json.dumps(records)
        result = await db.execute(text("""
            INSERT INTO supplier_parts (
                id, supplier_id, part_id, supplier_sku,
                price_usd, price_ils, shipping_cost_usd, shipping_cost_ils,
                availability, warranty_months, estimated_delivery_days,
                is_available, last_checked_at, created_at
            )
            SELECT
                CAST(j->>'id' AS UUID),
                CAST(j->>'supplier_id' AS UUID),
                CAST(j->>'part_id' AS UUID),
                j->>'supplier_sku',
                CAST(j->>'price_usd' AS NUMERIC),
                CAST(j->>'price_ils' AS NUMERIC),
                CAST(j->>'shipping_cost_usd' AS NUMERIC),
                CAST(j->>'shipping_cost_ils' AS NUMERIC),
                j->>'availability',
                CAST(j->>'warranty_months' AS INT),
                CAST(j->>'estimated_delivery_days' AS INT),
                CAST(j->>'is_available' AS BOOLEAN),
                NOW(), NOW()
            FROM json_array_elements(CAST(:payload AS json)) AS j
            ON CONFLICT (supplier_id, part_id) DO NOTHING
        """), {"payload": payload})
        await db.commit()
        return result.rowcount if result.rowcount and result.rowcount > 0 else len(records)

    # ── Pass 1: Universal suppliers → all active parts ───────────────────────
    offset = 0
    while True:
        parts = (await db.execute(text(
            "SELECT id, category, part_type, base_price FROM parts_catalog "
            "WHERE is_active = TRUE ORDER BY id OFFSET :o LIMIT :l"
        ), {"o": offset, "l": _BATCH})).mappings().fetchall()
        if not parts:
            break

        records: list = []
        for part in parts:
            part_id = str(part["id"])
            base = float(part["base_price"] or 0)
            if base <= 1.0:
                base = _CATEGORY_FALLBACK_ILS.get(part["category"] or "כללי", _DEFAULT_PRICE)
            warranty = _WARRANTY_MAP.get(part["part_type"] or "", 12)

            for (s_name, prefix, mult, s_ils, s_usd, days, avail, is_av) in _UNIVERSAL_SUPPLIERS:
                if s_name not in suppliers:
                    continue
                price = round(base * mult, 2)
                records.append({
                    "id": str(_uuid.uuid4()),
                    "supplier_id": str(suppliers[s_name]["id"]),
                    "part_id": part_id,
                    "supplier_sku": f"{prefix}-{part_id}",
                    "price_usd": round(price / ils_rate, 2),
                    "price_ils": price,
                    "shipping_cost_usd": s_usd,
                    "shipping_cost_ils": s_ils,
                    "availability": avail,
                    "warranty_months": warranty,
                    "estimated_delivery_days": days,
                    "is_available": is_av,
                })

        total_inserted += await _insert_batch(records)
        offset += _BATCH

    # ── Pass 2: Manufacturer-direct suppliers → filtered by manufacturer ──────
    for (s_name, prefix, mult, s_ils, s_usd, days, avail, is_av) in _MANUFACTURER_SUPPLIERS:
        if s_name not in suppliers:
            continue
        sup = suppliers[s_name]
        mfr = sup.get("manufacturer_name") if sup.get("is_manufacturer") else None
        where = "AND LOWER(manufacturer) = LOWER(:mfr)" if mfr else ""
        params_base: Dict[str, Any] = {"mfr": mfr} if mfr else {}

        offset = 0
        while True:
            parts = (await db.execute(text(
                f"SELECT id, category, part_type, base_price FROM parts_catalog "
                f"WHERE is_active = TRUE {where} ORDER BY id OFFSET :o LIMIT :l"
            ), {**params_base, "o": offset, "l": _BATCH})).mappings().fetchall()
            if not parts:
                break

            records = []
            for part in parts:
                part_id = str(part["id"])
                base = float(part["base_price"] or 0)
                if base <= 1.0:
                    base = _CATEGORY_FALLBACK_ILS.get(part["category"] or "כללי", _DEFAULT_PRICE)
                warranty = _WARRANTY_MAP.get(part["part_type"] or "", 12)
                price = round(base * mult, 2)
                records.append({
                    "id": str(_uuid.uuid4()),
                    "supplier_id": str(sup["id"]),
                    "part_id": part_id,
                    "supplier_sku": f"{prefix}-{part_id}",
                    "price_usd": round(price / ils_rate, 2),
                    "price_ils": price,
                    "shipping_cost_usd": s_usd,
                    "shipping_cost_ils": s_ils,
                    "availability": avail,
                    "warranty_months": warranty,
                    "estimated_delivery_days": days,
                    "is_available": is_av,
                })

            total_inserted += await _insert_batch(records)
            offset += _BATCH

    # ── Final count ──────────────────────────────────────────────────────────
    total_rows = (await db.execute(text("SELECT COUNT(*) FROM supplier_parts"))).scalar_one()
    parts_covered = (await db.execute(text(
        "SELECT COUNT(*) FROM (SELECT part_id FROM supplier_parts "
        "GROUP BY part_id HAVING COUNT(DISTINCT supplier_id) >= 5) t"
    ))).scalar_one()

    logger.info("populate_supplier_parts: inserted=%d total_rows=%d", total_inserted, total_rows)
    return {
        "task": "populate_supplier_parts",
        "status": "ok",
        "inserted": total_inserted,
        "total_supplier_parts_rows": total_rows,
        "parts_with_5plus_suppliers": parts_covered,
        "elapsed_s": round(time.monotonic() - t0, 2),
    }


# ---------------------------------------------------------------------------
# Task: validate_migrations
# Pre-flight safety check for Alembic migration files.
# Call via: POST /api/v1/admin/db-agent/run/validate_migrations
# ---------------------------------------------------------------------------

import re as _re
from pathlib import Path as _Path


def _check_migration_file(path: _Path) -> Tuple[bool, List[str], List[str]]:
    """
    Scan a single Alembic migration file for unsafe patterns.
    Returns (is_safe, errors, warnings).
    """
    errors: List[str] = []
    warnings: List[str] = []

    content = path.read_text()
    m = _re.search(r"def upgrade\(\):(.*?)(?=def downgrade|$)", content, _re.DOTALL)
    if not m:
        errors.append(f"{path.name}: no upgrade() function found")
        return False, errors, warnings

    code = m.group(1)

    # 1. NOT NULL column without server_default
    for match in _re.finditer(r"sa\.Column\([^)]*nullable=False[^)]*\)", code):
        start = code.rfind("\n", 0, match.start()) + 1
        line = code[start:code.find("\n", match.end())].strip()
        if "server_default" not in line:
            errors.append(f"{path.name}: NOT NULL without server_default → {line[:120]}")

    # 2. Column drops (warn — app code may still reference)
    for match in _re.finditer(r"op\.drop_column\(", code):
        start = code.rfind("\n", 0, match.start()) + 1
        line = code[start:code.find("\n", match.end())].strip()
        warnings.append(f"{path.name}: column drop — verify no live references → {line[:120]}")

    # 3. Table renames (warn)
    if _re.search(r"op\.rename_table\(", code):
        warnings.append(f"{path.name}: table rename — ensure compatibility views exist")

    # 4. Type change without VARCHAR intermediate (warn)
    for match in _re.finditer(r"op\.alter_column\([^,]+,\s*type_=", code):
        start = code.rfind("\n", 0, match.start()) + 1
        line = code[start:code.find("\n", match.end())].strip()
        if "VARCHAR" not in line and "String" not in line:
            warnings.append(f"{path.name}: type change without VARCHAR step → {line[:120]}")

    return len(errors) == 0, errors, warnings


async def _validate_migrations_task(db: AsyncSession) -> Dict[str, Any]:
    """
    Scan all Alembic migration files in both catalog and PII directories
    for patterns that could cause downtime on production deployment.
    Safe read-only operation; does not touch the database.
    """
    import asyncio as _asyncio

    base = _Path(__file__).parent
    dirs = {
        "catalog": base / "alembic" / "versions",
        "pii":     base / "alembic_pii" / "versions",
    }

    all_errors: List[str] = []
    all_warnings: List[str] = []
    files_checked = 0
    files_failed = 0

    def _scan_dirs() -> Tuple[int, int, List[str], List[str]]:
        _checked = 0
        _failed = 0
        _errors: List[str] = []
        _warnings: List[str] = []
        for label, mdir in dirs.items():
            if not mdir.exists():
                _warnings.append(f"{label}: directory not found ({mdir})")
                continue
            for mfile in sorted(mdir.glob("*.py")):
                if mfile.name.startswith("__"):
                    continue
                _checked += 1
                safe, errs, warns = _check_migration_file(mfile)
                _errors.extend(errs)
                _warnings.extend(warns)
                if not safe:
                    _failed += 1
        return _checked, _failed, _errors, _warnings

    files_checked, files_failed, all_errors, all_warnings = await asyncio.to_thread(_scan_dirs)

    status = "ok" if files_failed == 0 else "unsafe"
    logger.info(
        "validate_migrations: checked=%d failed=%d errors=%d warnings=%d",
        files_checked, files_failed, len(all_errors), len(all_warnings),
    )
    return {
        "task": "validate_migrations",
        "status": status,
        "files_checked": files_checked,
        "files_failed": files_failed,
        "errors": all_errors,
        "warnings": all_warnings,
    }


def _has_hebrew(value: str) -> bool:
    return bool(_HE_CHAR_RE.search(value or ""))


def _norm_en_brand(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def _skeleton(value: str) -> str:
    return re.sub(r"[aeiouy]", "", _norm_en_brand(value))


def _translit_hebrew(value: str) -> str:
    out: List[str] = []
    for ch in (value or ""):
        if ch in _HE_TO_LATIN:
            out.append(_HE_TO_LATIN[ch])
        elif "a" <= ch.lower() <= "z" or "0" <= ch <= "9" or ch == " ":
            out.append(ch.lower())
    return re.sub(r"\s+", " ", "".join(out)).strip()


def _load_transport_hebrew_candidates() -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    from run_rex_transport_office_pipeline import (
        _HE_COUNTRY_TOKENS as _TP_COUNTRY_TOKENS,
        _clean_space as _tp_clean_space,
        _norm_key as _tp_norm_key,
        _strip_punct as _tp_strip_punct,
    )

    source_path: Optional[Path] = None
    for p in _TRANSPORT_FREQ_PATHS:
        if p.exists():
            source_path = p
            break
    if source_path is None:
        raise FileNotFoundError("rex_transport_manufacturer_frequency.json not found")

    rows = json.loads(source_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        raise ValueError("manufacturer frequency artifact is not a list")

    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        manufacturer = str((row or {}).get("manufacturer") or "").strip()
        count = int((row or {}).get("count") or 0)
        if not manufacturer or count <= 0:
            continue
        key = _tp_norm_key(manufacturer)
        if not key:
            continue
        g = grouped.setdefault(key, {"total": 0, "variants": {}})
        g["total"] += count
        g["variants"][manufacturer] = g["variants"].get(manufacturer, 0) + count

    candidates: List[Dict[str, Any]] = []
    by_key: Dict[str, Dict[str, Any]] = {}

    for key, group in grouped.items():
        ordered = sorted(group["variants"].items(), key=lambda x: (-x[1], x[0]))
        canonical_raw = ordered[0][0]
        canonical_he = _tp_clean_space(_tp_strip_punct(_TP_COUNTRY_TOKENS.sub(" ", canonical_raw)))
        if not canonical_he or not _has_hebrew(canonical_he):
            continue

        aliases_he: List[str] = []
        for variant, _cnt in ordered[1:]:
            alias = _tp_clean_space(_tp_strip_punct(_TP_COUNTRY_TOKENS.sub(" ", variant)))
            if alias and _has_hebrew(alias) and alias != canonical_he and alias not in aliases_he:
                aliases_he.append(alias)

        latin = _norm_en_brand(_translit_hebrew(canonical_he))
        if not latin:
            continue

        cand = {
            "canonical_key": key,
            "hebrew_name": canonical_he,
            "aliases": aliases_he,
            "total_records": int(group["total"]),
            "latin_norm": latin,
            "latin_skeleton": _skeleton(latin),
        }
        candidates.append(cand)
        by_key.setdefault(key, cand)
        by_key.setdefault(_tp_norm_key(canonical_he), cand)
        for alias in aliases_he:
            by_key.setdefault(_tp_norm_key(alias), cand)

    candidates.sort(key=lambda x: (-x["total_records"], x["hebrew_name"]))
    return candidates, by_key


def _match_hebrew_alias(
    brand_name: str,
    candidates: List[Dict[str, Any]],
    by_key: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    from run_rex_transport_office_pipeline import _norm_key as _tp_norm_key

    en_norm = _norm_en_brand(brand_name)
    if not en_norm:
        return {"accepted": False, "reason": "empty_brand_name"}
    en_skel = _skeleton(en_norm)

    for manual_he in _MANUAL_HEBREW_CANDIDATES.get(en_norm, []):
        manual_he = str(manual_he or "").strip()
        if manual_he:
            return {
                "accepted": True,
                "hebrew_alias": manual_he,
                "confidence": 1.0,
                "margin": 1.0,
                "source": "manual",
            }

    if len(en_norm) <= 3:
        return {
            "accepted": False,
            "reason": "short_brand_auto_block",
        }

    scored: List[Tuple[float, int, Dict[str, Any]]] = []
    for cand in candidates:
        latin = cand["latin_norm"]
        direct = SequenceMatcher(None, en_norm, latin).ratio()
        skel = SequenceMatcher(None, en_skel, cand["latin_skeleton"]).ratio() if en_skel and cand["latin_skeleton"] else 0.0
        score = max(direct, skel * 0.98)
        if len(en_norm) >= 3 and len(latin) >= 3 and en_norm[:3] == latin[:3]:
            score = min(1.0, score + 0.02)
        scored.append((score, int(cand["total_records"]), cand))

    if not scored:
        return {"accepted": False, "reason": "no_candidates"}

    scored.sort(key=lambda x: (-x[0], -x[1], x[2]["hebrew_name"]))
    best_score, _best_count, best = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    margin = best_score - second_score

    accepted = best_score >= _ALIAS_MIN_SCORE and margin >= _ALIAS_MIN_MARGIN
    return {
        "accepted": accepted,
        "hebrew_alias": best["hebrew_name"],
        "confidence": round(best_score, 4),
        "margin": round(margin, 4),
        "source": "auto",
        "reason": None if accepted else "low_confidence",
    }


async def _auto_add_hebrew_brand_aliases_task(db: AsyncSession) -> Dict[str, Any]:
    try:
        candidates, by_key = await asyncio.to_thread(_load_transport_hebrew_candidates)
    except Exception as exc:
        logger.error("auto_add_hebrew_brand_aliases failed to load transport candidates: %s", exc)
        return {
            "task": "auto_add_hebrew_brand_aliases",
            "status": "error",
            "error": str(exc),
        }

    rows = (await db.execute(text(
        """
        SELECT id, name, aliases
        FROM car_brands
        WHERE is_active = TRUE
          AND il_market_priority IS NULL
        ORDER BY name
        """
    ))).fetchall()

    all_active = (await db.execute(text("SELECT name, aliases FROM car_brands WHERE is_active = TRUE"))).fetchall()
    alias_owner: Dict[str, str] = {}
    for owner_name, owner_aliases in all_active:
        for alias in list(owner_aliases or []):
            alias_str = str(alias or "").strip()
            if alias_str and _has_hebrew(alias_str):
                alias_owner.setdefault(alias_str, str(owner_name))

    updated = 0
    skipped_existing_hebrew = 0
    unmatched: List[str] = []
    low_confidence: List[Dict[str, Any]] = []
    matched: List[Dict[str, Any]] = []
    review_queue_upserted = 0

    for brand_id, name, aliases in rows:
        alias_list = list(aliases or [])
        if any(_has_hebrew(str(a)) for a in alias_list):
            skipped_existing_hebrew += 1
            continue

        match = _match_hebrew_alias(str(name), candidates, by_key)
        if not match.get("accepted"):
            if match.get("reason") == "low_confidence":
                low_confidence.append(
                    {
                        "brand_id": str(brand_id),
                        "brand": str(name),
                        "candidate": match.get("hebrew_alias"),
                        "confidence": match.get("confidence", 0.0),
                        "margin": match.get("margin", 0.0),
                        "reason": match.get("reason", "low_confidence"),
                    }
                )
            else:
                unmatched.append(str(name))
            continue

        he_alias = str(match.get("hebrew_alias") or "").strip()
        if not he_alias or he_alias in alias_list:
            continue

        owner = alias_owner.get(he_alias)
        if owner and owner != str(name):
            low_confidence.append(
                {
                    "brand_id": str(brand_id),
                    "brand": str(name),
                    "candidate": he_alias,
                    "confidence": match.get("confidence", 0.0),
                    "margin": match.get("margin", 0.0),
                    "reason": f"alias_conflict_with:{owner}",
                }
            )
            continue

        new_aliases = alias_list + [he_alias]
        await db.execute(
            text("UPDATE car_brands SET aliases = :aliases WHERE id = :id"),
            {"aliases": new_aliases, "id": str(brand_id)},
        )
        updated += 1
        alias_owner[he_alias] = str(name)
        matched.append(
            {
                "brand": str(name),
                "hebrew_alias": he_alias,
                "confidence": match.get("confidence", 0.0),
                "source": match.get("source", "auto"),
            }
        )

    for item in low_confidence:
        candidate_alias = str(item.get("candidate") or "").strip()
        if not candidate_alias:
            continue
        await db.execute(
            text(
                """
                INSERT INTO brand_alias_review_queue (
                    id, brand_id, brand_name, candidate_alias,
                    confidence, margin, reason, source, status,
                    created_at, updated_at
                )
                VALUES (
                    :id, :brand_id, :brand_name, :candidate_alias,
                    :confidence, :margin, :reason, :source, 'pending',
                    NOW(), NOW()
                )
                ON CONFLICT (brand_name, candidate_alias)
                DO UPDATE SET
                    confidence = EXCLUDED.confidence,
                    margin = EXCLUDED.margin,
                    reason = EXCLUDED.reason,
                    source = EXCLUDED.source,
                    status = CASE
                        WHEN brand_alias_review_queue.status = 'approved' THEN 'approved'
                        WHEN brand_alias_review_queue.status = 'rejected' THEN 'rejected'
                        ELSE 'pending'
                    END,
                    updated_at = NOW()
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "brand_id": item.get("brand_id"),
                "brand_name": str(item.get("brand") or "").strip(),
                "candidate_alias": candidate_alias,
                "confidence": float(item.get("confidence") or 0.0),
                "margin": float(item.get("margin") or 0.0),
                "reason": str(item.get("reason") or "low_confidence"),
                "source": "auto_matcher",
            },
        )
        review_queue_upserted += 1

    await db.commit()

    return {
        "task": "auto_add_hebrew_brand_aliases",
        "status": "ok",
        "transport_candidates": len(candidates),
        "brands_scanned": len(rows),
        "aliases_added": updated,
        "skipped_existing_hebrew_alias": skipped_existing_hebrew,
        "unmatched_count": len(unmatched),
        "low_confidence_count": len(low_confidence),
        "review_queue_upserted": review_queue_upserted,
        "matched_top20": matched[:20],
        "low_confidence_top20": sorted(low_confidence, key=lambda x: (-x["confidence"], x["brand"]))[:20],
        "unmatched_top50": unmatched[:50],
    }


async def _sync_transport_market_priority_task(db: AsyncSession) -> Dict[str, Any]:
    from run_rex_transport_office_pipeline import sync_market_priority_to_db

    result = await asyncio.to_thread(sync_market_priority_to_db)
    return {
        "task": "sync_transport_market_priority",
        "status": result.get("status", "ok"),
        **result,
    }


async def validate_watchdog_actions(db: AsyncSession) -> Dict[str, Any]:
    """
    Reads the watchdog's event log, confirms every action was correct, and alerts
    via WhatsApp if an anomaly is detected. Called each run_all_tasks cycle so the
    db_update_agent serves as the authoritative auditor for the watchdog's behaviour.

    Rules validated:
      kill_orphan        → must have details confirming pre-container-start backend_start
      kill_stuck_importer → must have dur_s > 0 (sanity; code guarantees > STUCK_IMPORT_S)
      warn_live_long     → informational only, marked ok
      Anomaly trigger    → any kill action where details suggest it might be a live
                           connection, or an unexpected burst (>10 kills in one cycle)
    """
    try:
        import watchdog_state as _wds
    except ImportError:
        return {"task": "validate_watchdog_actions", "status": "skip", "reason": "watchdog_state not available"}

    events = _wds.drain_unvalidated()
    if not events:
        return {"task": "validate_watchdog_actions", "status": "ok", "validated": 0, "anomalies": 0}

    anomalies = []
    validated = 0
    summary = {"kill_orphan": 0, "kill_zombie": 0, "kill_stuck_importer": 0, "warn_live_long": 0}

    # A same-container zombie query must have been blocking well past the watchdog's
    # ZOMBIE_QUERY_S threshold before it is killed. Kept a little below the code's
    # 2700s so a borderline timing race isn't mislabeled an anomaly.
    ZOMBIE_MIN_DUR_S = 2400

    for evt in events:
        summary[evt.action] = summary.get(evt.action, 0) + 1

        if evt.action == "kill_orphan":
            # Verify the recorded reason confirms orphan origin
            if "predates" not in evt.details and "pre-dates" not in evt.details:
                anomalies.append(
                    f"kill_orphan pid={evt.pid} dur={evt.dur_s}s — details don't confirm orphan: '{evt.details}'"
                )
        elif evt.action == "kill_zombie":
            # Legitimate same-container zombie kill — validate on its OWN terms:
            # it must have been blocking long enough to qualify as a zombie.
            if evt.dur_s < ZOMBIE_MIN_DUR_S:
                anomalies.append(
                    f"kill_zombie pid={evt.pid} dur={evt.dur_s}s — killed too early "
                    f"(< {ZOMBIE_MIN_DUR_S}s zombie threshold)"
                )
        elif evt.action == "kill_stuck_importer":
            if evt.dur_s <= 0:
                anomalies.append(f"kill_stuck_importer pid={evt.pid} has invalid dur_s={evt.dur_s}")

        evt.validated = True
        validated += 1

    # Burst check: >5 kills in one validation cycle is unusual
    total_kills = (summary.get("kill_orphan", 0) + summary.get("kill_zombie", 0)
                   + summary.get("kill_stuck_importer", 0))
    if total_kills > 5:
        anomalies.append(
            f"Unusual kill burst: {total_kills} kills in one cycle "
            f"(orphans={summary['kill_orphan']} zombies={summary['kill_zombie']} "
            f"importers={summary['kill_stuck_importer']})"
        )

    if anomalies:
        owner = os.getenv("OWNER_WHATSAPP_PHONE", "")
        if owner:
            msg = "⚠️ *Watchdog anomaly detected by db_update_agent*:\n" + "\n".join(f"• {a}" for a in anomalies)
            try:
                from social.whatsapp_provider import send_message as _wa_alert
                await _wa_alert(owner, msg)
            except Exception:
                pass
        logger.warning("validate_watchdog_actions anomalies: %s", anomalies)

    stats = _wds.stats()
    logger.info(
        "validate_watchdog_actions: validated=%d anomalies=%d | lifetime: %s",
        validated, len(anomalies), stats,
    )
    return {
        "task": "validate_watchdog_actions",
        "status": "ok" if not anomalies else "warn",
        "validated": validated,
        "anomalies": len(anomalies),
        "anomaly_details": anomalies,
        "lifetime_stats": stats,
    }


TASK_REGISTRY: Dict[str, Any] = {
    "clean_part_names":          clean_part_names,
    "normalize_part_types":      normalize_part_types,
    "normalize_categories":      normalize_categories,
    "dedup_catalog_parts":       dedup_catalog_parts,
    "normalize_availability":    normalize_availability,
    "fix_base_prices":           fix_base_prices,
    "normalize_base_price":      normalize_base_price,
    "fix_manufacturer_overflow": fix_manufacturer_overflow,
    "flag_fake_skus":            flag_fake_skus,
    "fill_car_brands":           fill_car_brands,
    "normalize_imported_manufacturers": normalize_imported_manufacturers,
    "sync_models_from_catalog": sync_models_from_catalog,
    "sync_models_from_catalog_file": sync_models_from_catalog_file,
    "backfill_catalog_fitment_from_xls": backfill_catalog_fitment_from_xls,
    "backfill_bmw_fitment_from_name_he": backfill_bmw_fitment_from_name_he,
    "backfill_mini_fitment_from_name_he": backfill_mini_fitment_from_name_he,
    "backfill_ford_fitment_from_name_he": backfill_ford_fitment_from_name_he,
    "backfill_jaguar_fitment_from_name": backfill_jaguar_fitment_from_name,
    "merge_catalog_fitment_from_part_vehicle_fitment": merge_catalog_fitment_from_part_vehicle_fitment,
    "sync_manufacturer_registries": sync_manufacturer_registries,
    "refresh_min_max_prices":    refresh_min_max_prices,
    "seed_system_settings":      seed_system_settings,
    "lookup_oem_spec": _lookup_oem_spec_task,
    "enrich_pending_parts":      _enrich_pending_parts_task,
    "trigger_scraper_for_registry_gaps": _trigger_scraper_for_registry_gaps_task,
    "trigger_scraper_for_misses": _trigger_scraper_for_misses_task,
    "generate_image_embeddings": _generate_image_embeddings_task,
    "auto_add_hebrew_brand_aliases": _auto_add_hebrew_brand_aliases_task,
    "validate_watchdog_actions": validate_watchdog_actions,
    "sync_transport_market_priority": _sync_transport_market_priority_task,
    # on-demand heavy tasks — NOT included in run_all_tasks
    "populate_supplier_parts":   _populate_supplier_parts_task,
    "validate_migrations":       _validate_migrations_task,
}


async def run_task(task_name: str, db: AsyncSession) -> Dict[str, Any]:
    """Run a single named task and return its report dict."""
    fn = TASK_REGISTRY.get(task_name)
    if fn is None:
        return {
            "task": task_name,
            "status": "error",
            "error": f"Unknown task '{task_name}'. "
                     f"Valid tasks: {list(TASK_REGISTRY.keys())}",
        }
    return await fn(db)


async def run_all_tasks(db: AsyncSession) -> Dict[str, Any]:
    """
    Run all cleaning + normalisation tasks in the recommended order.
    Returns a summary report dict.
    """
    from BACKEND_AUTH_SECURITY import get_redis
    from distributed_lock import acquire_lock
    # ROOT FIX 2026-07-09: was 21600s (6h). The heartbeat thread refreshes this
    # lock every 60s while the cycle runs, so the TTL only needs to outlast the
    # refresh interval — NOT the whole cycle. A 6h TTL meant that when a cycle
    # was killed (e.g. zombie watchdog under load) its lock lingered for up to
    # 6 HOURS, and every subsequent cycle skipped "lock held" the entire time —
    # the db agent was effectively down for 7h. 600s (10 min) TTL + 60s refresh
    # = a dead holder's lock auto-expires in ≤10 min and the next cycle runs.
    lock_ttl_s = int(os.getenv("DB_AGENT_LOCK_TTL_S", "600"))
    job_ttl_s = int(os.getenv("DB_AGENT_JOB_TTL_S", "21600"))
    heartbeat_interval_s = int(os.getenv("DB_AGENT_HEARTBEAT_INTERVAL_S", "60"))
    redis = await get_redis()
    _agent_lock = await acquire_lock(redis, "db_update_agent", ttl_seconds=lock_ttl_s)
    if not _agent_lock:
        return {"status": "skipped", "reason": "db_update_agent already running on another worker",
                "tasks_ok": 0, "tasks_error": 0}
    global _agent_running, _last_report
    _agent_running = True
    started_at = datetime.now(timezone.utc).isoformat()
    t0 = time.monotonic()
    results: List[Dict[str, Any]] = []
    job_id: Optional[str] = None
    task_timeout_s = int(os.getenv("DB_AGENT_TASK_TIMEOUT_S", "1800"))

    # Publish worker start to shared memory so agents can see live status
    try:
        from agents.memory import AgentMemory as _AgentMemory
        _wm = _AgentMemory(db, agent_name="db_update_agent")
        await _wm.write_worker_heartbeat({"status": "starting", "started_at": started_at})
    except Exception:
        pass

    # --- Thread-based heartbeat (immune to asyncio event-loop blocking) ----------
    import threading as _threading
    import psycopg2 as _psycopg2
    import redis as _redis_sync

    _hb_stop = _threading.Event()

    def _heartbeat_thread(jid: str, stop_evt: "_threading.Event") -> None:
        """Daemon thread: bumps last_heartbeat_at and refreshes Redis lock every
        heartbeat_interval_s seconds using plain synchronous clients.  Completely
        independent of the asyncio event loop, so it keeps firing even while the
        main loop is doing CPU-heavy work."""
        lock_key = "autospare:lock:db_update_agent"
        db_url_raw = os.getenv("DATABASE_URL", "")
        # Convert asyncpg URL → psycopg2 dsn
        db_dsn = db_url_raw.replace("postgresql+asyncpg://", "postgresql://")
        redis_url = os.getenv("REDIS_URL", "")
        def _make_db_conn():
            conn = _psycopg2.connect(
                db_dsn,
                connect_timeout=5,
                keepalives=1,
                keepalives_idle=10,
                keepalives_interval=2,
                keepalives_count=3,
                options="-c statement_timeout=8000",
            )
            conn.autocommit = True
            return conn

        try:
            db_conn = _make_db_conn()
        except Exception as exc:
            print(f"[heartbeat_thread] DB connect failed: {exc}")
            db_conn = None
        try:
            rconn = _redis_sync.from_url(redis_url)
        except Exception as exc:
            print(f"[heartbeat_thread] Redis connect failed: {exc}")
            rconn = None

        print(f"[heartbeat_thread] started, interval={heartbeat_interval_s}s, job_id={jid}", flush=True)
        while not stop_evt.wait(timeout=heartbeat_interval_s):
            print(f"[heartbeat_thread] tick, db_conn={'ok' if db_conn else 'none'}", flush=True)
            if db_conn:
                try:
                    with db_conn.cursor() as cur:
                        cur.execute(
                            "UPDATE job_registry SET last_heartbeat_at = NOW() "
                            "WHERE job_id = %s AND status = 'running'",  # noqa: S608
                            (jid,),
                        )
                        updated = cur.rowcount
                    if updated:
                        print(f"[heartbeat_thread] DB updated ok", flush=True)
                    else:
                        print(f"[heartbeat_thread] WARN: 0 rows updated for job_id={jid} (job may be gone/failed)", flush=True)
                except Exception as exc:
                    print(f"[heartbeat_thread] DB update failed: {exc}", flush=True)
                    try:
                        db_conn = _make_db_conn()
                    except Exception as reconn_exc:
                        print(f"[heartbeat_thread] DB reconnect failed: {reconn_exc}", flush=True)
                        db_conn = None
            if rconn:
                try:
                    rconn.set(lock_key, "1", ex=lock_ttl_s, xx=True)
                except Exception as exc:
                    print(f"[heartbeat_thread] Redis refresh failed: {exc}", flush=True)
        if db_conn:
            try:
                db_conn.close()
            except Exception:
                pass
        if rconn:
            try:
                rconn.close()
            except Exception:
                pass

    try:
        try:
            job_id = await job_registry_start(db, "run_all_tasks", ttl_seconds=job_ttl_s)
        except Exception as exc:
            logger.warning("run_all_tasks job_registry_start failed: %s", exc)
            try:
                await db.rollback()
            except Exception:
                pass

        ordered_tasks = [
            "seed_system_settings",
            "fill_car_brands",
            "normalize_imported_manufacturers",
            "sync_models_from_catalog",
            # sync_models_from_catalog_file excluded — XLS already imported; runs on-demand only
            "backfill_catalog_fitment_from_xls",
            # "backfill_bmw_fitment_from_name_he",  # DISABLED: OOM / already complete
            "backfill_mini_fitment_from_name_he",
            # "backfill_ford_fitment_from_name_he",  # DISABLED: OOM / already complete
            # "backfill_jaguar_fitment_from_name",  # DISABLED: OOM / already complete
            # "merge_catalog_fitment_from_part_vehicle_fitment",  # DISABLED: full-catalog scan, OOM per cycle, near-zero updates
            "sync_manufacturer_registries",
            "trigger_scraper_for_registry_gaps",
            "clean_part_names",
            "normalize_part_types",
            "normalize_categories",
            "dedup_catalog_parts",
            "normalize_availability",
            "fix_manufacturer_overflow",
            "flag_fake_skus",
            # "fix_base_prices",  # DISABLED: OOM / already complete
            # "normalize_base_price",  # DISABLED: OOM / already complete
            "refresh_min_max_prices",
            "lookup_oem_spec",
            "enrich_pending_parts",
            "trigger_scraper_for_misses",
            "generate_image_embeddings",
            "validate_watchdog_actions",
        ]

        shared_todos = await get_active_agent_todos(db, "db_update_agent")
        todo_task_names = [name for name in extract_todo_task_names(shared_todos) if name in TASK_REGISTRY]
        if todo_task_names:
            ordered_tasks = todo_task_names + [name for name in ordered_tasks if name not in todo_task_names]

        if job_id:
            _hb_thread = _threading.Thread(
                target=_heartbeat_thread,
                args=(job_id, _hb_stop),
                name="db_agent_heartbeat",
                daemon=True,
            )
            _hb_thread.start()

        for task_name in ordered_tasks:
            t_task = time.monotonic()
            print(f"[run_all_tasks] starting: {task_name}", flush=True)
            logger.info("run_all_tasks → starting: %s", task_name)

            # Central deadlock-retry (added 2026-07-11): several batched-UPDATE
            # tasks (normalize_categories/part_types, dedup_catalog_parts,
            # backfill_catalog_fitment_from_xls) contend with the concurrent
            # harvester on parts_catalog rows → transient DeadlockDetectedError
            # that used to abort the task (status=error) ~5×/day each. Postgres
            # picks a victim and aborts it; retrying the transaction resolves it.
            # Handles BOTH a raised deadlock and a task that catches it and returns
            # {status:error, error:"...deadlock..."}. Timeouts are NOT retried.
            _MAX_TASK_DEADLOCK_RETRIES = 3
            result = None
            for _dl_attempt in range(_MAX_TASK_DEADLOCK_RETRIES):
                try:
                    if task_timeout_s > 0:
                        result = await asyncio.wait_for(run_task(task_name, db), timeout=task_timeout_s)
                    else:
                        result = await run_task(task_name, db)

                except asyncio.TimeoutError:
                    with contextlib.suppress(Exception):
                        await db.rollback()
                    result = {
                        "task": task_name,
                        "status": "error",
                        "error": f"Task timeout after {task_timeout_s}s",
                    }
                    print(f"[run_all_tasks] TIMEOUT: {task_name} after {task_timeout_s}s", flush=True)
                    logger.error("run_all_tasks: task %s timed out after %ss", task_name, task_timeout_s)
                    break

                except Exception as task_exc:
                    with contextlib.suppress(Exception):
                        await db.rollback()
                    if "deadlock" in str(task_exc).lower() and _dl_attempt < _MAX_TASK_DEADLOCK_RETRIES - 1:
                        print(f"[run_all_tasks] DEADLOCK: {task_name} (attempt {_dl_attempt+1}) — retrying", flush=True)
                        logger.warning("run_all_tasks: task %s deadlock (attempt %d) — retrying", task_name, _dl_attempt + 1)
                        await asyncio.sleep(1.0 * (_dl_attempt + 1))
                        continue
                    result = {
                        "task": task_name,
                        "status": "error",
                        "error": str(task_exc)[:400],
                    }
                    print(f"[run_all_tasks] ERROR: {task_name} — {str(task_exc)[:200]}", flush=True)
                    logger.error("run_all_tasks: task %s raised: %s", task_name, task_exc, exc_info=True)
                    break

                # Task returned a result. Retry if it self-reported a deadlock.
                if (
                    isinstance(result, dict)
                    and result.get("status") == "error"
                    and "deadlock" in str(result.get("error", "")).lower()
                    and _dl_attempt < _MAX_TASK_DEADLOCK_RETRIES - 1
                ):
                    with contextlib.suppress(Exception):
                        await db.rollback()
                    print(f"[run_all_tasks] DEADLOCK(result): {task_name} (attempt {_dl_attempt+1}) — retrying", flush=True)
                    logger.warning("run_all_tasks: task %s returned deadlock (attempt %d) — retrying", task_name, _dl_attempt + 1)
                    await asyncio.sleep(1.0 * (_dl_attempt + 1))
                    continue
                break

            elapsed_task = round(time.monotonic() - t_task, 1)
            status = result.get("status", "?")
            print(f"[run_all_tasks] done: {task_name} status={status} elapsed={elapsed_task}s", flush=True)
            results.append(result)
            if result.get("status") == "error":
                logger.warning("run_all_tasks: task %s errored, continuing", task_name)
            if job_id:
                await job_heartbeat(db, job_id)

        total_elapsed = round(time.monotonic() - t0, 2)
        ok_count = sum(1 for r in results if r.get("status") == "ok")
        err_count = len(results) - ok_count

        # Mark todos complete whose task_names all finished ok
        completed_task_names = {r["task"] for r in results if r.get("status") == "ok"}
        todos_completed = 0
        for todo in shared_todos:
            task_names_for_todo = [n for n in extract_todo_task_names([todo]) if n in TASK_REGISTRY]
            if not task_names_for_todo:
                continue
            if all(t in completed_task_names for t in task_names_for_todo):
                try:
                    await db.execute(text(
                        "UPDATE agent_todos SET status = 'completed', completed_at = NOW(), "
                        "progress_pct = 100, updated_at = NOW() "
                        "WHERE id = CAST(:tid AS uuid) AND status != 'completed'"
                    ), {"tid": todo["id"]})
                    todos_completed += 1
                except Exception:
                    pass
        if todos_completed:
            try:
                await db.commit()
            except Exception:
                pass
            logger.info("run_all_tasks: marked %d todos completed", todos_completed)
            print(f"[run_all_tasks] marked {todos_completed} todos completed", flush=True)

        report = {
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "total_elapsed_s": total_elapsed,
            "tasks_ok": ok_count,
            "tasks_error": err_count,
            "shared_todos": [
                {"id": todo["id"], "title": todo["title"], "status": todo["status"]}
                for todo in shared_todos
            ],
            "todo_task_names": todo_task_names,
            "ordered_tasks": ordered_tasks,
            "results": results,
        }

        _last_report = report
        _agent_running = False
        logger.info(
            "run_all_tasks finished: ok=%d err=%d elapsed=%.1fs",
            ok_count,
            err_count,
            total_elapsed,
        )
        # G8 2026-07-20 (owner directive): task errors must reach the owner's WhatsApp,
        # not sit silently in the report/DB until someone checks. Alert once per unique
        # failing-task SET per 24h (Redis-keyed) so a persistent error doesn't nag every
        # 3h cycle but a NEW failure alerts immediately. Quiet-hours-aware via
        # _wa_send_quiet (queued at night, delivered in the morning window).
        if err_count > 0:
            try:
                import hashlib as _hashlib
                _err_tasks = [
                    r for r in results
                    if isinstance(r, dict) and r.get("status") == "error"
                ]
                _sig = _hashlib.sha256(
                    ",".join(sorted(str(r.get("task", "?")) for r in _err_tasks)).encode()
                ).hexdigest()[:16]
                from BACKEND_AUTH_SECURITY import get_redis as _gr
                _r = await _gr()
                _ck = f"autospare:alert_cooldown:dbagent_task_errors:{_sig}"
                if not await _r.exists(_ck):
                    await _r.set(_ck, "1", ex=86400)
                    _lines = [f"🛠️ *db_update_agent — {err_count} משימות נכשלו במחזור האחרון*"]
                    for r in _err_tasks[:8]:
                        _lines.append(f"• {r.get('task','?')}: {str(r.get('error',''))[:120]}")
                    if len(_err_tasks) > 8:
                        _lines.append(f"…ועוד {len(_err_tasks) - 8}")
                    _lines.append(f"({ok_count} משימות הצליחו, {total_elapsed:.0f}s)")
                    _owner = os.getenv("OWNER_WHATSAPP_PHONE", "")
                    if _owner:
                        try:
                            from BACKEND_API_ROUTES import _wa_send_quiet as _waq
                            await _waq(to=_owner, text="\n".join(_lines))
                        except Exception:
                            from social.whatsapp_provider import send_message as _was
                            await _was(to=_owner, text="\n".join(_lines))
            except Exception as _alert_exc:
                logger.warning("run_all_tasks error-alert failed: %s", _alert_exc)
        # Publish final stats to shared memory
        try:
            from agents.memory import AgentMemory as _AgentMemory
            _wm = _AgentMemory(db, agent_name="db_update_agent")
            await _wm.write_worker_heartbeat({
                "status": "completed", "tasks_ok": ok_count, "tasks_error": err_count,
                "elapsed_s": round(total_elapsed, 1),
            })
        except Exception:
            pass
        if job_id:
            _finish_ok = False
            try:
                # Use a fresh session — the run session may be in a bad state after 25 tasks.
                from BACKEND_DATABASE_MODELS import async_session_factory as _asf_finish
                async with _asf_finish() as _finish_db:
                    await job_registry_finish(_finish_db, job_id, status="completed")
                _finish_ok = True
            except Exception as exc:
                logger.warning("run_all_tasks job_registry_finish failed: %s", exc)
                print(f"[run_all_tasks] WARN: job_registry_finish failed: {exc}", flush=True)
            if not _finish_ok:
                # Fallback: use psycopg2 directly (same as heartbeat thread) to avoid orphan records
                try:
                    import psycopg2 as _pg2
                    _fb_dsn = os.getenv("DATABASE_URL", "").replace("postgresql+asyncpg://", "postgresql://")
                    with _pg2.connect(_fb_dsn, connect_timeout=5) as _fc:
                        _fc.autocommit = True
                        with _fc.cursor() as _cur:
                            _cur.execute(
                                "UPDATE job_registry SET status='completed', completed_at=NOW(), "
                                "last_heartbeat_at=NOW() WHERE job_id=%s AND status='running'",
                                (job_id,),
                            )
                    print(f"[run_all_tasks] psycopg2 fallback finish OK for {job_id}", flush=True)
                except Exception as fb_exc:
                    print(f"[run_all_tasks] psycopg2 fallback finish FAILED: {fb_exc}", flush=True)
        _hb_stop.set()  # stop heartbeat only after DB is updated
        await _agent_lock.release()
        return report

    except Exception as exc:
        if job_id:
            try:
                from BACKEND_DATABASE_MODELS import async_session_factory as _asf_finish
                async with _asf_finish() as _finish_db:
                    await job_registry_finish(_finish_db, job_id, status="dead", error_message=str(exc)[:500])
            except Exception as finish_exc:
                print(f"[run_all_tasks] WARN: job_registry_finish(dead) failed: {finish_exc}", flush=True)
        _hb_stop.set()
        _agent_running = False
        await _agent_lock.release()
        raise


def get_last_report() -> Dict[str, Any]:
    """Return the last run report (or empty dict if never run)."""
    return _last_report


def is_running() -> bool:
    return _agent_running


# =========================================================================
# Optional background loop
# =========================================================================

_bg_task: Optional[asyncio.Task] = None


async def _agent_loop(get_db_fn, interval_hours: float = 6.0) -> None:
    """Periodic background loop.  Runs run_all_tasks every `interval_hours`."""
    from resilience import log_job_failure
    # Give container_start.sh time to clear stale Redis locks before first run
    await asyncio.sleep(90)
    logger.info(
        "DB update agent background loop started (interval=%.1fh)", interval_hours
    )
    while True:
        result: dict = {}
        try:
            async for db in get_db_fn():
                result = await run_all_tasks(db)
        except Exception as exc:
            error_msg = str(exc)[:500]
            logger.error("DB update agent loop error: %s", error_msg)
            try:
                from BACKEND_DATABASE_MODELS import pii_session_factory
                async with pii_session_factory() as pii_db:
                    await log_job_failure(
                        pii_db,
                        job_name="run_all_tasks",
                        error=error_msg,
                        payload={},
                        attempts=1,
                    )
            except Exception as dlq_err:
                logger.error("Failed to log run_all_tasks to DLQ: %s", dlq_err)

        # If skipped (lock held), retry in 2 minutes instead of sleeping the full interval
        if result.get("status") == "skipped":
            logger.warning("run_all_tasks skipped (lock held) — will retry in 2m")
            await asyncio.sleep(120)
        else:
            await asyncio.sleep(interval_hours * 3600)


def start_agent_task(get_db_fn, interval_hours: float = 6.0) -> None:
    """
    Call this from the FastAPI startup event to enable the periodic loop.
    ``get_db_fn`` should be the same ``get_db`` dependency used in routes.
    """
    global _bg_task
    _bg_task = asyncio.create_task(
        _agent_loop(get_db_fn, interval_hours),
        name="db_update_agent",
    )
    logger.info("DB update agent background task created")
