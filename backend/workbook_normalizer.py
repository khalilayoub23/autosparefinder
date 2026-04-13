from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional

import openpyxl

from manufacturer_normalization import (
    canonicalize_vehicle_model_for_manufacturer,
    normalize_manufacturer_name,
    normalize_vehicle_model_name,
    normalize_vehicle_submodel_name,
)

RAW_XLSX_FILE = Path(__file__).parent / "data" / "parts_database.xlsx"
NORMALIZED_XLSX_FILE = Path(__file__).parent / "data" / "parts_database.normalized.xlsx"
NORMALIZED_SHEET_NAME = "parts_catalog_import"
NORMALIZED_SOURCE_TAG = "parts_database.normalized.xlsx"

SHEET_TYPES = {
    "Chevrolet": "A",
    "Citroen": "F",
    "Peugeot": "F",
    "Hyundai": "B",
    "Mercedes": "B",
    "Mitsubishi": "B",
    "Smart": "B",
    "GEN": "BM",
    "JAECOO": "BM",
    "ORA": "BM",
    "Porsche": "C",
    "Suzuki": "D",
    "Renault": "E",
}

MULTIBLOCK_SHEETS = {"GEN", "JAECOO", "ORA"}

GENERATION_YEAR_RULES = {
    ("citroen", "berlingo", "b9"): (2008, 2018),
    ("citroen", "berlingo", "k9"): (2018, 2027),
    ("citroen", "berlingo", "k9 acc"): (2018, 2027),
    ("peugeot", "partner", "b9"): (2008, 2018),
    ("peugeot", "partner", "k9"): (2018, 2027),
    ("peugeot", "partner", "k9 acc"): (2018, 2027),
}

HEADER_VALUES = {"זמינות מלאי", "מחיר לצרכן", "תיאור החלק", "מותג", "מספר קטלוגי"}

NORMALIZED_HEADERS = [
    "source_sheet",
    "source_row",
    "catalog_num",
    "sku",
    "oem_number",
    "name",
    "category",
    "manufacturer",
    "manufacturer_id",
    "part_type",
    "description",
    "specifications",
    "compatible_vehicles",
    "base_price",
]


def clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text not in {"-", "None", "nan"} else None


def parse_price(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        text = str(value).replace("₪", "").replace(",", "").strip()
        price = float(text)
        return min(max(price, 0.0), 99_999_999.99)
    except (TypeError, ValueError):
        return 0.0


def parse_manufacturer_fields(raw_mfr: Any, fallback_brand: str) -> tuple[str, int | None]:
    manufacturer_name = normalize_manufacturer_name(fallback_brand, fallback_brand)
    manufacturer_id = None
    mfr = clean(raw_mfr)
    if not mfr:
        return manufacturer_name, manufacturer_id

    try:
        if str(mfr).isdigit():
            manufacturer_id = int(str(mfr))
            return manufacturer_name, manufacturer_id
        num = float(str(mfr))
        if num.is_integer():
            manufacturer_id = int(num)
            return manufacturer_name, manufacturer_id
    except (TypeError, ValueError):
        pass

    manufacturer_name = normalize_manufacturer_name(str(mfr), fallback_brand)
    return manufacturer_name, manufacturer_id


def make_sku(brand: str, catalog_num: Any, row_idx: int) -> str:
    prefix = brand[:4].upper().replace(" ", "_")
    cat = clean(catalog_num)
    if cat:
        cat_clean = str(cat).replace(" ", "").replace("/", "-").strip()
        raw_sku = f"{prefix}-{cat_clean}"
        if len(raw_sku) > 100:
            digest = hashlib.md5(cat_clean.encode()).hexdigest()[:12]
            raw_sku = f"{prefix}-{digest}"
    else:
        raw_sku = f"{prefix}-R{row_idx}"
    return raw_sku[:100]


def _extract_years(text: str) -> List[int]:
    years = set()
    for match in re.findall(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", text or ""):
        year = int(match)
        if 1990 <= year <= 2027:
            years.add(year)
    return sorted(years)


def _build_compatibility(brand: str, vehicle_text: Optional[str]) -> List[Dict[str, Any]]:
    raw_vehicle = clean(vehicle_text)
    if not raw_vehicle:
        return []

    manufacturer = normalize_manufacturer_name(brand, brand)
    years = _extract_years(raw_vehicle)

    model_text = raw_vehicle
    model_text = re.sub(r"\b(19|20)\d{2}\b", "", model_text).strip()
    model_text = re.sub(r"\b(new|basic|accessories|accessory)\b", "", model_text, flags=re.IGNORECASE).strip()
    model_text = re.sub(r"\s{2,}", " ", model_text).strip()

    lower_text = model_text.lower()
    for token in {brand, manufacturer}:
        if token and lower_text.startswith(token.lower() + " "):
            model_text = model_text[len(token):].strip()
            lower_text = model_text.lower()
            break

    if re.search(r"\s-\s*", model_text):
        base_model, sub_model = [part.strip() for part in re.split(r"\s-\s*", model_text, maxsplit=1)]
    else:
        match = re.match(
            r"^(?P<base>[A-Za-z0-9\u0590-\u05FF\s]+?)\s+(?P<sub>[A-Z]\d{1,3}(?:\s+[A-Z]{2,8})?)$",
            model_text,
            flags=re.IGNORECASE,
        )
        if match:
            base_model = match.group("base").strip()
            sub_model = match.group("sub").upper()
        else:
            base_model = model_text
            sub_model = ""

    model = canonicalize_vehicle_model_for_manufacturer(manufacturer, base_model)
    sub = normalize_vehicle_submodel_name(sub_model)
    if not model:
        return []

    if not years:
        span = GENERATION_YEAR_RULES.get((manufacturer.casefold(), model.casefold(), sub.casefold()))
        if span:
            years = [span[0], span[1]]

    entry: Dict[str, Any] = {
        "manufacturer": manufacturer,
        "model": model,
        "source": NORMALIZED_SOURCE_TAG,
    }
    if sub:
        entry["sub_model"] = sub
    if years:
        entry["year_from"] = years[0]
        entry["year_to"] = years[-1]
    return [entry]


def _make_normalized_record(
    *,
    brand: str,
    source_sheet: str,
    source_row: int,
    catalog_num: Any,
    name: Any,
    part_type: Any,
    price: Any,
    manufacturer_raw: Any = None,
    vehicle: Any = None,
    specifications: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    catalog = clean(catalog_num)
    part_name = clean(name)
    if not catalog or catalog in HEADER_VALUES or not part_name or part_name in HEADER_VALUES:
        return None

    manufacturer, manufacturer_id = parse_manufacturer_fields(manufacturer_raw, brand)
    compatible_vehicles = _build_compatibility(manufacturer, vehicle)
    sku = make_sku(brand, catalog, source_row)
    spec_payload = specifications or {}
    if manufacturer:
        spec_payload.setdefault("manufacturer_name", manufacturer)
    if manufacturer_id is not None:
        spec_payload.setdefault("manufacturer_id", manufacturer_id)

    return {
        "source_sheet": source_sheet,
        "source_row": source_row,
        "catalog_num": catalog,
        "sku": sku,
        "oem_number": catalog,
        "name": part_name[:255],
        "category": manufacturer[:100],
        "manufacturer": manufacturer[:100],
        "manufacturer_id": manufacturer_id,
        "part_type": (clean(part_type) or "unknown")[:50],
        "description": part_name[:500],
        "specifications": json.dumps(spec_payload, ensure_ascii=False),
        "compatible_vehicles": json.dumps(compatible_vehicles, ensure_ascii=False),
        "base_price": parse_price(price),
    }


def parse_multiblock_row(sheet_name: str, row: tuple, row_idx: int) -> List[Dict[str, Any]]:
    row = list(row)
    max_col = len(row)

    def g(index: int) -> Any:
        return row[index] if index < max_col else None

    records: List[Dict[str, Any]] = []

    def add(avail_v: Any, price_v: Any, name_v: Any, mfr_v: Any, cat_v: Any) -> None:
        stock = clean(avail_v)
        record = _make_normalized_record(
            brand=sheet_name.strip(),
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=cat_v,
            name=name_v,
            part_type="unknown",
            price=price_v,
            manufacturer_raw=mfr_v,
            specifications={"stock_status": stock} if stock else {},
        )
        if record:
            records.append(record)

    if any(g(i) is not None for i in range(5)):
        add(g(0), g(1), g(2), g(3), g(4))
    if any(g(i) is not None for i in range(5, 10)) and (g(7) is not None or g(9) is not None):
        add(g(5), g(6), g(7), g(8), g(9))

    avail_common = g(5)
    mfr_common = g(8)
    for index in range(10, max_col - 1, 3):
        price_v = g(index)
        name_v = g(index + 1)
        catalog_v = g(index + 2) if index + 2 < max_col else None
        if any(v is not None for v in (price_v, name_v, catalog_v)):
            add(avail_common, price_v, name_v, mfr_common, catalog_v)

    return records


def parse_row(sheet_name: str, row: tuple, row_idx: int) -> Optional[Dict[str, Any]]:
    if not any(row):
        return None

    stype = SHEET_TYPES.get(sheet_name, "A")
    brand = sheet_name.strip()

    def g(index: int) -> Any:
        return row[index] if index < len(row) else None

    if stype == "A":
        stock = clean(g(4))
        warranty = clean(g(6))
        importer = clean(g(0))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(1),
            name=g(2),
            part_type=g(3),
            price=g(5),
            vehicle=g(7),
            specifications={k: v for k, v in {"stock_status": stock, "warranty": warranty, "importer": importer}.items() if v},
        )

    if stype == "F":
        stock = clean(g(1))
        warranty = clean(g(3))
        importer = clean(g(4))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(7),
            name=g(6),
            part_type=g(5),
            price=g(2),
            vehicle=g(0),
            specifications={k: v for k, v in {"stock_status": stock, "warranty": warranty, "importer": importer}.items() if v},
        )

    if stype in {"B", "BM"}:
        stock = clean(g(0))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(4),
            name=g(2),
            part_type="unknown",
            price=0.0,
            manufacturer_raw=g(3),
            specifications={"stock_status": stock} if stock else {},
        )

    if stype == "C":
        stock = clean(g(1))
        warranty = clean(g(2))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(5),
            name=g(4),
            part_type=g(3),
            price=g(0),
            specifications={k: v for k, v in {"stock_status": stock, "warranty": warranty}.items() if v},
        )

    if stype == "D":
        stock = clean(g(1))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(2),
            name=clean(g(5)) or clean(g(4)),
            part_type=g(4),
            price=g(0),
            manufacturer_raw=g(3),
            specifications={"stock_status": stock} if stock else {},
        )

    if stype == "E":
        stock = clean(g(1))
        category_code = clean(g(2))
        return _make_normalized_record(
            brand=brand,
            source_sheet=sheet_name,
            source_row=row_idx,
            catalog_num=g(5),
            name=g(4),
            part_type=g(3),
            price=0.0,
            specifications={k: v for k, v in {"stock_status": stock, "category_code": category_code}.items() if v},
        )

    return None


def iter_normalized_rows(selected_sheets: Optional[Iterable[str]] = None) -> Iterator[Dict[str, Any]]:
    wanted = {sheet.lower() for sheet in (selected_sheets or [])}
    workbook = openpyxl.load_workbook(RAW_XLSX_FILE, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if wanted and sheet_name.lower() not in wanted:
                continue

            sheet = workbook[sheet_name]
            is_multiblock = sheet_name in MULTIBLOCK_SHEETS
            start_row = 1 if is_multiblock else (8 if SHEET_TYPES.get(sheet_name) == "F" else 3)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if is_multiblock:
                    records = parse_multiblock_row(sheet_name, row, row_idx)
                else:
                    parsed = parse_row(sheet_name, row, row_idx)
                    records = [parsed] if parsed else []
                for record in records:
                    yield record
    finally:
        workbook.close()


def build_normalized_workbook(
    output_path: Path = NORMALIZED_XLSX_FILE,
    selected_sheets: Optional[Iterable[str]] = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet(NORMALIZED_SHEET_NAME)
    sheet.append(NORMALIZED_HEADERS)

    for record in iter_normalized_rows(selected_sheets):
        sheet.append([record.get(header, "") for header in NORMALIZED_HEADERS])

    workbook.save(output_path)
    return output_path
